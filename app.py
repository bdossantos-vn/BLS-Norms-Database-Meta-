from __future__ import annotations

import hashlib
import json
import math
import base64
import os
import shutil
import time
import urllib.error
import urllib.request
import zipfile
from contextlib import contextmanager
from difflib import SequenceMatcher
from dataclasses import dataclass
from datetime import datetime
from html import escape
from io import BytesIO
from pathlib import Path
import re

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.legend import LegendEntry
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import (
    CharacterProperties,
    Font as DrawingFont,
    Paragraph,
    ParagraphProperties,
    RegularTextRun,
    RichTextProperties,
)
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components


DEFAULT_DENOMINATOR = "Total answering"
DENOMINATOR_OPTIONS = ["Total answering", "Total sample"]
NOT_AVAILABLE = "Not available"
NOT_TESTED = "Not tested"
NA_NORM_OPTION = "NA"
NO_RESPONDENT_ID_OPTION = "No respondent ID / not available"
APP_DIR = Path(__file__).resolve().parent
SETTINGS_PATH = APP_DIR / "denominator_settings.json"
NORM_MAPPING_PATH = APP_DIR / "norm_mapping_settings.json"
BOX_SCORE_SETTINGS_PATH = APP_DIR / "box_score_settings.json"
QUESTION_TYPE_SETTINGS_PATH = APP_DIR / "question_type_settings.json"
NA_ALIAS_SETTINGS_PATH = APP_DIR / "na_alias_settings.json"
CHANGELOG_PATH = APP_DIR / "CHANGELOG.md"
STATUS_PATH = APP_DIR / "status.md"
LOGO_PATH = APP_DIR / "assets" / "vn_logo.png"


def configured_norm_database_dir() -> Path:
    configured_path = os.environ.get("BLS_NORMS_DATA_DIR")
    if not configured_path:
        return APP_DIR / "norm_database"

    path = Path(configured_path).expanduser()
    return path if path.is_absolute() else APP_DIR / path


NORM_DATABASE_LEGACY_DIR = APP_DIR / "norm_database"
NORM_DATABASE_DIR = configured_norm_database_dir()
NORM_DATABASE_DATASETS_DIR = NORM_DATABASE_DIR / "datasets"
NORM_DATABASE_MANIFEST_PATH = NORM_DATABASE_DIR / "manifest.json"
NORM_DATABASE_WORKBOOK_PATH = NORM_DATABASE_DIR / "saved_norm_tables.xlsx"
NORM_DATABASE_LOCK_PATH = NORM_DATABASE_DIR / ".write.lock"
UPLOADED_DATASETS_DIR = APP_DIR / "uploaded_datasets"
UPLOADED_RAW_DIR = UPLOADED_DATASETS_DIR / "raw_uploads"
UPLOADED_NORM_WORKBOOKS_DIR = UPLOADED_DATASETS_DIR / "norm_workbooks"
UPLOADED_SETTINGS_DIR = UPLOADED_DATASETS_DIR / "norm_settings"
UPLOADED_HISTORY_DIR = UPLOADED_DATASETS_DIR / "norm_history"
UPLOADED_HISTORY_INDEX_PATH = UPLOADED_HISTORY_DIR / "index.json"
UPLOADED_MANIFEST_BACKUP_PATH = UPLOADED_SETTINGS_DIR / "manifest.json"
UPLOADED_NORM_TABLES_BACKUP_PATH = UPLOADED_SETTINGS_DIR / "saved_norm_tables.xlsx"
NORM_DATASET_RESPONDENT_SHEET = "Respondent Data"
NORM_DATASET_RULES_SHEET = "Rules"
NORM_DATASET_NORM_RULES_SHEET = "Norm Rules"
NORM_DATASET_QUESTION_LABELS_SHEET = "Question Labels"
NORM_DATASET_RESPONSE_LABELS_SHEET = "Response Labels"
DUPLICATE_RESPONDENT_ID_OVERLAP_THRESHOLD = 0.8
SIGNIFICANCE_ALPHA = 0.05
NO_LABEL_SHEET = "No labels sheet"
VN_PINK = "FF005C"
VN_CONTROL_GRAY = "C9D0D8"
VN_BLACK = "000000"
VN_WHITE = "FFFFFF"
VN_GRAY_50 = "F8F8FA"
VN_GRAY_200 = "E2E5EA"
VN_GRAY_400 = "8A8F98"
VN_GREEN = "15803D"
VN_SIG_RED = "D92D20"
SMART_TABLES_LAYOUT = "BLS / Smart Tables layout"
RESPONDENT_ROWS_LAYOUT = "Standard respondent table"
DATA_LAYOUT_OPTIONS = [SMART_TABLES_LAYOUT, RESPONDENT_ROWS_LAYOUT]
PAGE_NAMES = [
    "Survey Question Audit",
    "Denominators",
    "Norm tables",
    "Saved datasets",
]
QUESTION_TYPES = [
    "Single-Select",
    "Multi-Select",
    "Scale / Likert",
    "Numeric Data",
    "Open-End Text",
    "Ignore",
]
BOX_SCORE_OPTIONS = ["T2B", "T3B", "B2B", "B3B"]
BOX_SCORE_LABELS = {
    "T2B": "Top 2 Box",
    "T3B": "Top 3 Box",
    "B2B": "Bottom 2 Box",
    "B3B": "Bottom 3 Box",
}
DEFAULT_VARIABLE_BLACKLIST = [
    "StartDate",
    "EndDate",
    "IPAddress",
    "RecipientEmail",
    "RecipientFirstName",
    "RecipientLastName",
    "Status",
    "Duration",
    "Duration (in seconds)",
    "Progress",
    "RecordedDate",
    "ResponseId",
    "ResponseSet",
    "LocationLatitude",
    "LocationLongitude",
    "UserLanguage",
    "Finished",
    "DistributionChannel",
    "ExternalReference",
    "Q_RecaptchaScore",
    "Q_RecaptchaStatus",
    "Q_RecaptchaError",
    "Q_AmbiguousTextPresent",
    "Q_AmbiguousTextQuestions",
    "Q_StraightliningCount",
    "Q_StraightliningPercentage",
    "Q_StraightliningQuestions",
    "Q_UnansweredPercentage",
    "Q_UnansweredQuestions",
]
DEFAULT_BLACKLIST_PREFIXES = [
    "Q_RelevantID",
    "Q_DuplicateRespondent",
]
GROUP_COLUMN_CANDIDATES = [
    "cell",
    "group",
    "condition",
    "treatment",
    "variant",
    "segment",
]
PROJECT_METADATA_VARIABLES = [
    "brand",
    "industry",
    "client",
    "quarter",
    "year",
    "methodology",
    "project",
    "country",
    "c_key",
]
RESPONDENT_ID_CANDIDATES = [
    "ResponseId",
    "Response ID",
    "response_id",
    "respondent_id",
    "Respondent ID",
    "respondentid",
    "resp_id",
    "respid",
    "record_id",
    "case_id",
    "participant_id",
    "panelist_id",
    "uuid",
    "c_key",
]
NORM_FILTER_FIELDS = [
    ("Project", ["project"]),
    ("Brand", ["brand"]),
    ("Client", ["client"]),
    ("Industry", ["industry"]),
    ("Country", ["country"]),
    ("Year", ["year"]),
    ("Quarter", ["quarter"]),
    ("Gender", ["gender", "sex"]),
    ("Age", ["age"]),
]
EXACT_NORM_FILTERS = {"Project", "Brand", "Client", "Industry", "Country", "Year", "Quarter"}
NORM_FILTER_SESSION_PREFIX = "norm_filter_"
UPLOAD_WORKBOOK_SESSION_KEY = "survey_excel_workbook_upload"
ACTIVE_WORKBOOK_NAME_SESSION_KEY = "active_workbook_name"
ACTIVE_WORKBOOK_BYTES_SESSION_KEY = "active_workbook_bytes"
LIKERT_PATTERNS = [
    "strongly disagree",
    "disagree",
    "neutral",
    "agree",
    "strongly agree",
    "very dissatisfied",
    "dissatisfied",
    "satisfied",
    "very satisfied",
]
SCALE_LABEL_HINTS = [
    "agree or disagree",
    "how likely",
    "to what extent",
    "feel about",
    "how interested",
    "brand affinity",
    "brand sentiment",
    "sentiment",
    "interest",
    "affinity",
    "relationship with",
]
SCALE_VALUE_HINTS = [
    "very interested",
    "somewhat interested",
    "moderately interested",
    "not that interested",
    "not very interested",
    "not at all interested",
    "love it",
    "like it",
    "neutral",
    "dislike it",
    "hate it",
    "very likely",
    "quite likely",
    "moderately likely",
    "not that likely",
    "not likely",
    "very unlikely",
    "somewhat likely",
    "somewhat better",
    "about the same",
    "much worse",
    "much better",
    "somewhat worse",
    "dedicated harry potter fan",
    "new to the series but interested",
    "nostalgic toward it",
    "not a fan",
    "leads much more often",
    "leads somewhat more often",
    "follows somewhat more often",
    "follows much more often",
    "follows more often",
]
SCALE_ORDER_PATTERNS = [
    ("love it", 0),
    ("very unlikely", 4),
    ("not at all interested", 4),
    ("not at all likely", 4),
    ("not very interested", 3),
    ("not that interested", 3),
    ("strongly disagree", 4),
    ("dislike it", 3),
    ("hate it", 4),
    ("very likely", 0),
    ("very interested", 0),
    ("strongly agree", 0),
    ("much better", 0),
    ("leads much more often", 0),
    ("quite likely", 1),
    ("somewhat likely", 1),
    ("somewhat interested", 1),
    ("somewhat agree", 1),
    ("somewhat better", 1),
    ("like it", 1),
    ("leads somewhat more often", 1),
    ("about the same", 2),
    ("neutral", 2),
    ("moderately interested", 2),
    ("moderately likely", 2),
    ("neither agree nor disagree", 2),
    ("follows somewhat more often", 2),
    ("somewhat worse", 3),
    ("not that likely", 3),
    ("not likely", 3),
    ("somewhat disagree", 3),
    ("follows much more often", 3),
    ("follows more often", 3),
    ("much worse", 4),
]
HP_INTEREST_ORDER_PATTERNS = [
    ("i am a dedicated harry potter fan", 0),
    ("i enjoyed it in the past and feel nostalgic toward it", 1),
    ("i'm new to the series but interested", 2),
    ("i’m new to the series but interested", 2),
    ("i'm not a fan", 3),
    ("i’m not a fan", 3),
]
EXCLUSIVE_RESPONSE_PATTERNS = [
    "none of the above",
    "none",
    "other",
    "prefer not to say",
    "don't know",
    "dont know",
    "unsure",
    "not applicable",
    "n/a",
]
QUESTION_COLUMN_CANDIDATES = [
    "question",
    "variable",
    "column",
    "field",
    "field_name",
    "question_name",
    "name",
]
VALUE_COLUMN_CANDIDATES = [
    "value",
    "code",
    "response_value",
    "option_value",
    "raw_value",
]
LABEL_COLUMN_CANDIDATES = [
    "label",
    "response_label",
    "option_label",
    "value_label",
    "text",
    "response",
]
QUESTION_LABEL_COLUMN_CANDIDATES = [
    "question_label",
    "question_text",
    "question_title",
    "variable_label",
]


@dataclass
class RespondentSheet:
    dataframe: pd.DataFrame
    question_labels: dict[str, str]
    metadata_rows_removed: int = 0
    metadata_columns_default_na: int = 0
    metadata_columns: list[str] | None = None


def logo_base64() -> str:
    if not LOGO_PATH.exists():
        return ""
    return base64.b64encode(LOGO_PATH.read_bytes()).decode("ascii")


def apply_bls_theme() -> None:
    st.markdown(
        """
        <style>
            :root {
                --vn-black: #000000;
                --vn-white: #ffffff;
                --vn-red: #ff005c;
                --vn-orange: #ff6927;
                --vn-yellow: #ffc227;
                --vn-gray-50: #f8f8fa;
                --vn-gray-100: #f1f2f5;
                --vn-gray-200: #e2e5ea;
                --vn-gray-400: #8a8f98;
            }

            html, body, [class*="css"] {
                font-family: "Proxima Nova", "Avenir Next", "Segoe UI", "Helvetica Neue", Arial, sans-serif;
            }

            .stApp {
                background: var(--vn-white);
                color: var(--vn-black);
                color-scheme: light;
            }

            .block-container {
                padding-top: 1.25rem;
                padding-bottom: 3rem;
                max-width: 1280px;
            }

            h1, h2, h3, h4, h5, h6 {
                color: var(--vn-black);
                letter-spacing: 0;
            }

            [data-testid="stMain"] p,
            [data-testid="stMain"] label,
            [data-testid="stMain"] legend,
            [data-testid="stMain"] span,
            [data-testid="stMain"] div[data-testid="stMarkdownContainer"] * {
                color: var(--vn-black);
            }

            .vn-brand-bar {
                display: flex;
                align-items: center;
                justify-content: space-between;
                gap: 1rem;
                padding: 0.9rem 1rem;
                margin: 0 0 1.25rem 0;
                background: linear-gradient(90deg, #050505 0%, #151515 72%, var(--vn-red) 100%);
                border-radius: 8px;
                box-shadow: 0 10px 26px rgba(0, 0, 0, 0.12);
            }

            .vn-brand-left {
                display: flex;
                align-items: center;
                gap: 1rem;
                min-width: 0;
            }

            .vn-brand-copy {
                display: flex;
                flex-direction: column;
                gap: 0.1rem;
                min-width: 0;
            }

            .vn-brand-kicker {
                color: rgba(255, 255, 255, 0.72) !important;
                font-size: 0.78rem;
                text-transform: uppercase;
                letter-spacing: 0.12em;
            }

            .vn-brand-title {
                color: var(--vn-white) !important;
                font-size: 1.28rem;
                font-weight: 800;
                line-height: 1.15;
            }

            .vn-brand-subtitle {
                color: rgba(255, 255, 255, 0.82) !important;
                font-size: 0.9rem;
                line-height: 1.25;
                text-align: right;
            }

            .stTabs [data-baseweb="tab-list"] {
                gap: 0.45rem;
                border-bottom: 1px solid var(--vn-gray-200);
            }

            .stTabs [data-baseweb="tab"] {
                border-radius: 8px 8px 0 0;
                padding: 0.55rem 0.9rem;
                color: var(--vn-black);
            }

            .stTabs [aria-selected="true"] {
                background: var(--vn-black);
            }

            .stTabs [aria-selected="true"] * {
                color: var(--vn-white) !important;
            }

            .stButton > button,
            .stDownloadButton > button,
            [data-testid="stPopover"] button {
                border-radius: 8px;
                border: 1px solid var(--vn-red);
                color: var(--vn-white) !important;
                background: var(--vn-red);
                font-weight: 700;
                transition: all 0.16s ease;
            }

            .stButton > button:hover,
            .stDownloadButton > button:hover,
            [data-testid="stPopover"] button:hover {
                border-color: var(--vn-red);
                color: var(--vn-white);
                background: linear-gradient(90deg, var(--vn-red) 0%, var(--vn-orange) 100%);
            }

            .stButton > button *,
            .stDownloadButton > button *,
            [data-testid="stPopover"] button * {
                color: var(--vn-white) !important;
                fill: var(--vn-white) !important;
            }

            .stButton > button:disabled,
            .stDownloadButton > button:disabled,
            [data-testid="stPopover"] button:disabled {
                background: var(--vn-gray-400) !important;
                border-color: var(--vn-gray-400) !important;
                color: var(--vn-white) !important;
                opacity: 0.65;
            }

            [data-testid="stMetric"] {
                background: var(--vn-white) !important;
                border: 1px solid var(--vn-gray-200);
                border-radius: 8px;
                padding: 0.9rem 1rem 0.75rem 1rem;
                box-shadow: 0 6px 18px rgba(0, 0, 0, 0.05);
            }

            [data-testid="stMetric"] * {
                color: var(--vn-black) !important;
            }

            [data-testid="stAlert"],
            [data-testid="stExpander"] details,
            details {
                border-radius: 8px !important;
                border-color: var(--vn-gray-200) !important;
            }

            [data-testid="stMain"] div[data-baseweb="select"] > div,
            [data-testid="stMain"] div[data-baseweb="input"] > div,
            [data-testid="stMain"] textarea,
            [data-testid="stMain"] input {
                border-radius: 8px !important;
                background: var(--vn-white) !important;
                color: var(--vn-black) !important;
            }

            [data-testid="stMain"] div[data-baseweb="select"] *,
            [data-testid="stMain"] div[data-baseweb="input"] *,
            [data-testid="stMain"] [data-baseweb="radio"] *,
            [data-testid="stMain"] [data-baseweb="checkbox"] * {
                color: var(--vn-black) !important;
                fill: var(--vn-black) !important;
            }

            [data-testid="stMain"] [data-testid="stFileUploader"] section,
            [data-testid="stMain"] [data-testid="stFileUploaderDropzone"] {
                background: var(--vn-white) !important;
                border: 1px solid var(--vn-gray-200) !important;
                border-radius: 8px !important;
            }

            [data-testid="stMain"] [data-testid="stFileUploader"] section *,
            [data-testid="stMain"] [data-testid="stFileUploaderDropzone"] * {
                color: var(--vn-black) !important;
                fill: var(--vn-black) !important;
            }

            [data-testid="stDataFrame"],
            [data-testid="stDataFrameGlideDataEditor"],
            [data-testid="stDataEditor"],
            [data-testid*="GlideDataEditor"] {
                --gdg-bg-cell: var(--vn-white) !important;
                --gdg-bg-cell-medium: var(--vn-white) !important;
                --gdg-bg-header: var(--vn-gray-50) !important;
                --gdg-bg-header-has-focus: var(--vn-gray-50) !important;
                --gdg-bg-header-hovered: var(--vn-gray-100) !important;
                --gdg-text-dark: var(--vn-black) !important;
                --gdg-text-medium: #333333 !important;
                --gdg-text-light: #666666 !important;
                --gdg-text-header: var(--vn-black) !important;
                --gdg-border-color: var(--vn-gray-200) !important;
                --gdg-horizontal-border-color: var(--vn-gray-200) !important;
                --gdg-accent-color: var(--vn-red) !important;
                --gdg-accent-fg: var(--vn-white) !important;
                color-scheme: light !important;
                background: var(--vn-white) !important;
                color: var(--vn-black) !important;
            }

            .vn-norm-table-wrap {
                width: 100%;
                overflow-x: auto;
                margin: 0.45rem 0 1.45rem 0;
                border: 1px solid var(--vn-gray-200);
                border-radius: 8px;
                background: var(--vn-white);
            }

            .vn-norm-table {
                width: 100%;
                border-collapse: collapse;
                background: var(--vn-white) !important;
                color: var(--vn-black) !important;
                font-size: 0.92rem;
            }

            .vn-norm-table thead th {
                background: var(--vn-gray-50) !important;
                color: var(--vn-black) !important;
                border-bottom: 1px solid var(--vn-gray-200);
                font-weight: 800;
                padding: 0.7rem 0.8rem;
                text-align: left;
                white-space: nowrap;
            }

            .vn-norm-table tbody td {
                background: var(--vn-white) !important;
                color: var(--vn-black) !important;
                border-bottom: 1px solid var(--vn-gray-200);
                padding: 0.65rem 0.8rem;
                vertical-align: top;
            }

            .vn-norm-table tbody tr:last-child td {
                border-bottom: 0;
            }

            .vn-norm-table tbody tr:nth-child(even) td {
                background: var(--vn-gray-50) !important;
            }

            .vn-chart-card {
                width: 100%;
                min-width: 0;
                margin: 0.45rem 0 1.45rem 0;
                padding: 0.9rem 0.95rem 0.8rem 0.95rem;
                border: 1px solid var(--vn-gray-200);
                border-radius: 8px;
                background: var(--vn-white);
                overflow: hidden;
            }

            .vn-chart-title {
                color: var(--vn-black) !important;
                font-size: 0.86rem;
                font-weight: 800;
                margin: 0;
            }

            .vn-chart-header {
                display: flex;
                align-items: center;
                justify-content: space-between;
                gap: 0.75rem;
                margin: 0 0 0.65rem 0;
            }

            .vn-chart-legend {
                display: flex;
                align-items: center;
                flex-wrap: wrap;
                justify-content: flex-end;
                gap: 0.6rem;
            }

            .vn-chart-legend-item {
                display: inline-flex;
                align-items: center;
                gap: 0.28rem;
                color: var(--vn-black) !important;
                font-size: 0.72rem;
                font-weight: 700;
                line-height: 1;
                white-space: nowrap;
            }

            .vn-chart-legend-swatch {
                display: inline-block;
                width: 12px;
                height: 12px;
                border-radius: 3px;
                background: var(--vn-gray-200);
            }

            .vn-chart-legend-swatch.control {
                background: #c9d0d8;
            }

            .vn-chart-legend-swatch.test {
                background: var(--vn-red);
            }

            .vn-chart-legend-bubble {
                display: inline-flex;
                align-items: center;
                justify-content: center;
                width: 15px;
                height: 15px;
                border: 1px solid var(--vn-gray-200);
                border-radius: 999px;
                background: var(--vn-white);
                color: #4c5361 !important;
                font-size: 0.65rem;
                font-weight: 800;
            }

            .vn-chart-scroll {
                display: flex;
                align-items: stretch;
                gap: 0.85rem;
                overflow-x: visible;
                padding: 0.25rem 0.1rem 0.2rem 0.1rem;
            }

            .vn-chart-group {
                position: relative;
                flex: 1 1 86px;
                min-width: 68px;
                max-width: 150px;
            }

            .vn-chart-plot {
                position: relative;
                display: grid;
                grid-template-columns: 1fr 1fr;
                align-items: end;
                gap: 0.35rem;
                height: 190px;
                padding: 0 0.25rem;
                border-bottom: 1px solid var(--vn-gray-200);
            }

            .vn-chart-bar-column {
                display: flex;
                flex-direction: column;
                align-items: center;
                justify-content: flex-end;
                height: 100%;
                min-width: 0;
            }

            .vn-chart-value {
                color: var(--vn-black) !important;
                font-size: 0.88rem;
                font-weight: 800;
                line-height: 1;
                margin-bottom: 0.35rem;
                white-space: nowrap;
            }

            .vn-chart-value.control {
                color: #aeb7c1 !important;
            }

            .vn-chart-value.test {
                color: var(--vn-red) !important;
            }

            .vn-chart-bar {
                width: 100%;
                min-height: 4px;
                border-radius: 4px 4px 0 0;
            }

            .vn-chart-bar.control {
                background: #c9d0d8;
            }

            .vn-chart-bar.test {
                background: var(--vn-red);
            }

            .vn-lift-bubble {
                position: absolute;
                left: 50%;
                bottom: 2.1rem;
                z-index: 3;
                display: flex;
                align-items: center;
                justify-content: center;
                width: 44px;
                height: 44px;
                transform: translateX(-50%);
                border: 1px solid var(--vn-gray-200);
                border-radius: 999px;
                background: var(--vn-white);
                color: #4c5361 !important;
                font-size: 0.95rem;
                font-weight: 800;
                box-shadow: 0 6px 15px rgba(0, 0, 0, 0.08);
            }

            .vn-lift-bubble.positive {
                color: #15803d !important;
            }

            .vn-lift-bubble.negative {
                color: #d92d20 !important;
            }

            .vn-lift-bubble.neutral {
                color: #4c5361 !important;
            }

            .vn-chart-label {
                color: var(--vn-black) !important;
                font-size: 0.76rem;
                font-weight: 700;
                line-height: 1.15;
                min-height: 3.2rem;
                padding-top: 0.45rem;
                text-align: center;
                overflow-wrap: anywhere;
                white-space: normal;
            }

            [data-testid="stCaptionContainer"],
            [data-testid="stCaptionContainer"] * {
                color: var(--vn-gray-400) !important;
            }

            @media (max-width: 720px) {
                .vn-brand-bar {
                    align-items: flex-start;
                    flex-direction: column;
                }

                .vn-brand-subtitle {
                    text-align: left;
                }

                .vn-chart-header {
                    align-items: flex-start;
                    flex-direction: column;
                }

                .vn-chart-legend {
                    justify-content: flex-start;
                }
            }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_bls_header() -> None:
    logo_html = ""
    encoded_logo = logo_base64()
    if encoded_logo:
        logo_html = (
            f'<img src="data:image/png;base64,{encoded_logo}" '
            'style="height:42px; width:auto; display:block;" alt="Viral Nation logo" />'
        )

    st.markdown(
        f"""
        <div class="vn-brand-bar">
            <div class="vn-brand-left">
                {logo_html}
                <div class="vn-brand-copy">
                    <div class="vn-brand-kicker">Viral Nation</div>
                    <div class="vn-brand-title">BLS Norms Database</div>
                </div>
            </div>
            <div class="vn-brand-subtitle">Brand Lift Study Norms</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_page_navigation(page_index: int, review_mode: bool = False) -> None:
    page_sequence = [2, 3] if review_mode else list(range(len(PAGE_NAMES)))
    if page_index in page_sequence:
        sequence_position = page_sequence.index(page_index)
        previous_index = page_sequence[sequence_position - 1] if sequence_position > 0 else -1
        next_index = (
            page_sequence[sequence_position + 1]
            if sequence_position < len(page_sequence) - 1
            else len(PAGE_NAMES)
        )
    else:
        previous_index = -1
        next_index = page_sequence[0]

    previous_label = PAGE_NAMES[previous_index] if previous_index >= 0 else ""
    next_label = PAGE_NAMES[next_index] if next_index < len(PAGE_NAMES) else ""
    previous_disabled = "disabled" if previous_index < 0 else ""
    next_disabled = "disabled" if next_index >= len(PAGE_NAMES) else ""
    previous_button = (
        f"""
            <button class="vn-page-nav-back" type="button" {previous_disabled} data-target-label="{previous_label}" onclick="goToTab({previous_index}, this.dataset.targetLabel)">
                Back{': ' + previous_label if previous_label else ''}
            </button>
        """
        if not previous_disabled
        else ""
    )
    next_button = (
        f"""
            <button class="vn-page-nav-next" type="button" {next_disabled} data-target-label="{next_label}" onclick="goToTab({next_index}, this.dataset.targetLabel)">
                Next{': ' + next_label if next_label else ''}
            </button>
        """
        if not next_disabled
        else ""
    )

    components.html(
        f"""
        <style>
            .vn-page-nav {{
                display: flex;
                gap: 12px;
                align-items: center;
                justify-content: space-between;
                width: 100%;
                box-sizing: border-box;
                padding: 2px 0 8px 0;
                font-family: "Proxima Nova", "Avenir Next", "Segoe UI", Arial, sans-serif;
            }}

            .vn-page-nav button {{
                appearance: none;
                border: 1px solid #ff005c;
                border-radius: 8px;
                background: #ff005c;
                color: #ffffff;
                font-weight: 800;
                font-size: 14px;
                line-height: 1.2;
                padding: 12px 16px;
                min-width: 190px;
                cursor: pointer;
            }}

            .vn-page-nav button:hover:not(:disabled) {{
                background: linear-gradient(90deg, #ff005c 0%, #ff6927 100%);
            }}

            .vn-page-nav button:disabled {{
                background: #8a8f98;
                border-color: #8a8f98;
                cursor: not-allowed;
                opacity: 0.65;
            }}

            .vn-page-nav-next {{
                margin-left: auto;
            }}

            @media (max-width: 720px) {{
                .vn-page-nav {{
                    flex-direction: column;
                }}

                .vn-page-nav button {{
                    width: 100%;
                }}
            }}
        </style>
        <div class="vn-page-nav">
            {previous_button}
            {next_button}
        </div>
        <script>
            function tabText(element) {{
                return (element.textContent || '').replace(/\\s+/g, ' ').trim();
            }}

            function candidateTabs(documentRoot) {{
                const directTabs = Array.from(
                    documentRoot.querySelectorAll('[role="tab"], [data-baseweb="tab"]')
                );
                if (directTabs.length) {{
                    return directTabs;
                }}

                const tabLists = Array.from(documentRoot.querySelectorAll('[role="tablist"]'));
                return tabLists.flatMap((tabList) =>
                    Array.from(tabList.querySelectorAll('button, [tabindex], div'))
                ).filter((element) => tabText(element));
            }}

            function goToTab(index, label) {{
                const documentRoot = window.parent.document;
                const tabs = candidateTabs(documentRoot);
                const target = tabs.find((tab) => tabText(tab) === label)
                    || tabs.find((tab) => tabText(tab).includes(label))
                    || tabs[index];
                if (target) {{
                    target.click();
                    target.scrollIntoView({{ block: 'center', inline: 'nearest' }});
                }}
            }}
        </script>
        """,
        height=62,
    )


def load_denominator_settings() -> dict[str, str]:
    if not SETTINGS_PATH.exists():
        return {}

    try:
        data = json.loads(SETTINGS_PATH.read_text())
    except (json.JSONDecodeError, OSError):
        return {}

    return {
        question: setting
        for question, setting in data.items()
        if setting in DENOMINATOR_OPTIONS
    }


def save_denominator_settings(settings: dict[str, str]) -> None:
    SETTINGS_PATH.write_text(json.dumps(settings, indent=2, sort_keys=True) + "\n")


def load_norm_mapping_settings() -> dict[str, str]:
    if not NORM_MAPPING_PATH.exists():
        return {}

    try:
        data = json.loads(NORM_MAPPING_PATH.read_text())
    except (json.JSONDecodeError, OSError):
        return {}

    return {
        normalize_answer(source) or str(source): normalize_answer(target) or NA_NORM_OPTION
        for source, target in data.items()
    }


def save_norm_mapping_settings(settings: dict[str, str]) -> None:
    NORM_MAPPING_PATH.write_text(json.dumps(settings, indent=2, sort_keys=True) + "\n")


def load_box_score_settings() -> dict[str, list[str]]:
    if not BOX_SCORE_SETTINGS_PATH.exists():
        return {}

    try:
        data = json.loads(BOX_SCORE_SETTINGS_PATH.read_text())
    except (json.JSONDecodeError, OSError):
        return {}

    settings: dict[str, list[str]] = {}
    for question, selected_scores in data.items():
        if not isinstance(selected_scores, list):
            continue
        normalized_question = normalize_answer(question)
        if not normalized_question:
            continue
        settings[normalized_question] = [
            score
            for score in selected_scores
            if score in BOX_SCORE_OPTIONS
        ]

    return settings


def save_box_score_settings(settings: dict[str, list[str]]) -> None:
    clean_settings = {
        question: [
            score
            for score in scores
            if score in BOX_SCORE_OPTIONS
        ]
        for question, scores in settings.items()
    }
    BOX_SCORE_SETTINGS_PATH.write_text(json.dumps(clean_settings, indent=2, sort_keys=True) + "\n")


def load_question_type_settings() -> dict[str, str]:
    if not QUESTION_TYPE_SETTINGS_PATH.exists():
        return {}

    try:
        data = json.loads(QUESTION_TYPE_SETTINGS_PATH.read_text())
    except (json.JSONDecodeError, OSError):
        return {}

    settings: dict[str, str] = {}
    for question, question_type in data.items():
        normalized_question = normalize_answer(question)
        normalized_type = normalize_answer(question_type)
        if normalized_question and normalized_type in QUESTION_TYPES:
            settings[normalized_question] = normalized_type

    return settings


def save_question_type_settings(settings: dict[str, str]) -> None:
    clean_settings = {
        question: question_type
        for question, question_type in settings.items()
        if question_type in QUESTION_TYPES
    }
    QUESTION_TYPE_SETTINGS_PATH.write_text(json.dumps(clean_settings, indent=2, sort_keys=True) + "\n")


def load_na_alias_settings() -> set[str]:
    if not NA_ALIAS_SETTINGS_PATH.exists():
        return set()

    try:
        data = json.loads(NA_ALIAS_SETTINGS_PATH.read_text())
    except (json.JSONDecodeError, OSError):
        return set()

    if not isinstance(data, list):
        return set()

    return {
        alias
        for alias in (normalize_alias_key(value) for value in data)
        if alias
    }


def save_na_alias_settings(alias_keys: set[str]) -> None:
    clean_aliases = sorted(alias for alias in alias_keys if alias)
    NA_ALIAS_SETTINGS_PATH.write_text(json.dumps(clean_aliases, indent=2) + "\n")


def merge_na_alias_settings(alias_keys: set[str]) -> set[str]:
    merged_aliases = load_na_alias_settings() | alias_keys
    save_na_alias_settings(merged_aliases)
    return merged_aliases


def setting_key(question: str) -> str:
    digest = hashlib.sha1(question.encode("utf-8")).hexdigest()[:12]
    return f"denominator_{digest}"


def normalize_answer(value) -> str | None:
    if pd.isna(value):
        return None

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    answer = str(value).strip()
    if re.fullmatch(r"-?\d+\.0", answer):
        return answer[:-2]

    return answer or None


def split_answers(value, split_multi_select: bool, delimiter: str) -> list[str]:
    answer = normalize_answer(value)
    if answer is None:
        return []

    if not split_multi_select or not delimiter:
        return [answer]

    options = [option.strip() for option in answer.split(delimiter)]
    return list(dict.fromkeys(option for option in options if option))


def bls_response_parts(value, extra_delimiter: str = "") -> list[str]:
    normalized_value = normalize_answer(value)
    if normalized_value is None:
        return []

    delimiters = [";", ","]
    if extra_delimiter and extra_delimiter not in delimiters:
        delimiters.append(extra_delimiter)

    split_pattern = "|".join(re.escape(delimiter) for delimiter in delimiters)
    choices: list[str] = []
    for part in re.split(split_pattern, normalized_value):
        normalized_part = normalize_answer(part)
        if normalized_part and normalized_part not in choices:
            choices.append(normalized_part)

    return choices or [normalized_value]


def value_matches_response_option(value, response_option: str, extra_delimiter: str = "") -> bool:
    normalized_value = normalize_answer(value)
    normalized_option = normalize_answer(response_option)
    if normalized_value is None or normalized_option is None:
        return False
    if normalized_value == normalized_option:
        return True
    return normalized_option in bls_response_parts(normalized_value, extra_delimiter)


def infer_multi_select(series: pd.Series) -> bool:
    values = series.dropna().map(normalize_answer).dropna()
    if values.empty:
        return False
    delimiter_ratio = values.map(lambda value: bool(re.search(r"[;,]", value))).mean()
    return float(delimiter_ratio) >= 0.3


def extract_observed_response_options(
    series: pd.Series,
    split_multi_select: bool,
    delimiter: str,
) -> list[str]:
    choices: list[str] = []
    should_split = split_multi_select or infer_multi_select(series)

    for value in series.dropna().tolist():
        if should_split:
            parts = bls_response_parts(value, delimiter if split_multi_select else "")
        else:
            normalized = normalize_answer(value)
            parts = [normalized] if normalized else []

        for part in parts:
            if part not in choices:
                choices.append(part)

    return choices


def has_answer(value) -> bool:
    return normalize_answer(value) is not None


def denominator_for(group_df: pd.DataFrame, question: str, denominator_setting: str) -> int | None:
    if question not in group_df.columns:
        return None

    if denominator_setting == "Total sample":
        return len(group_df)

    if denominator_setting == "Total answering":
        return int(group_df[question].map(has_answer).sum())

    return None


def response_counts(
    group_df: pd.DataFrame,
    question: str,
    split_multi_select: bool,
    delimiter: str,
    response_options: list[str] | None = None,
) -> dict[str, int]:
    counts: dict[str, int] = {}

    if question not in group_df.columns:
        return counts

    if response_options is not None:
        for option in response_options:
            counts[option] = int(
                group_df[question]
                .map(lambda value: value_matches_response_option(value, option, delimiter))
                .sum()
            )
        return counts

    for value in group_df[question]:
        for option in split_answers(value, split_multi_select, delimiter):
            counts[option] = counts.get(option, 0) + 1

    return counts


def option_order(*count_maps: dict[str, int]) -> list[str]:
    seen: dict[str, None] = {}
    for counts in count_maps:
        for option in counts:
            seen.setdefault(option, None)
    return list(seen.keys())


def normalize_sort_text(value) -> str:
    normalized = normalize_answer(value) or ""
    return re.sub(r"\s+", " ", normalized.lower()).strip()


def contains_phrase(normalized: str, phrase: str) -> bool:
    return re.search(rf"(?<![a-z]){re.escape(phrase)}(?![a-z])", normalized) is not None


def is_exclusive_response_label(value) -> bool:
    normalized = normalize_sort_text(value)
    return any(pattern in normalized for pattern in EXCLUSIVE_RESPONSE_PATTERNS)


def count_pattern_hits(values: list[str], patterns: list[str]) -> int:
    return sum(any(pattern in value for pattern in patterns) for value in values)


def scale_choice_score(choice: str) -> tuple[int, int] | None:
    normalized = normalize_sort_text(choice)

    if contains_phrase(normalized, "leads much more often"):
        return (0, 0)
    if contains_phrase(normalized, "leads somewhat more often"):
        return (1, 0)
    if contains_phrase(normalized, "follows somewhat more often"):
        return (3, 0)
    if contains_phrase(normalized, "follows much more often") or contains_phrase(normalized, "follows more often"):
        return (4, 0)

    if contains_phrase(normalized, "i am a dedicated harry potter fan"):
        return (0, 0)
    if contains_phrase(normalized, "feel nostalgic toward it"):
        return (1, 0)
    if contains_phrase(normalized, "new to the series but interested"):
        return (2, 0)
    if contains_phrase(normalized, "not a fan"):
        return (4, 0)

    if contains_phrase(normalized, "love it"):
        return (0, 0)
    if contains_phrase(normalized, "dislike it"):
        return (3, 0)
    if contains_phrase(normalized, "hate it"):
        return (4, 0)
    if contains_phrase(normalized, "like it"):
        return (1, 0)
    if (
        contains_phrase(normalized, "neutral")
        or contains_phrase(normalized, "about the same")
        or contains_phrase(normalized, "neither agree nor disagree")
    ):
        return (2, 0)

    if contains_phrase(normalized, "much better"):
        return (0, 0)
    if contains_phrase(normalized, "somewhat better"):
        return (1, 0)
    if contains_phrase(normalized, "somewhat worse"):
        return (3, 0)
    if contains_phrase(normalized, "much worse"):
        return (4, 0)

    if contains_phrase(normalized, "strongly agree"):
        return (0, 0)
    if contains_phrase(normalized, "somewhat agree"):
        return (1, 0)
    if contains_phrase(normalized, "somewhat disagree"):
        return (3, 0)
    if contains_phrase(normalized, "strongly disagree"):
        return (4, 0)

    positive_weight = None
    if "not at all " in normalized or "very unlikely" in normalized:
        positive_weight = 4
    elif "not that " in normalized or "not very " in normalized or "not likely" in normalized:
        positive_weight = 3
    elif "very " in normalized:
        positive_weight = 0
    elif "quite " in normalized:
        positive_weight = 1
    elif "somewhat " in normalized:
        positive_weight = 1
    elif "moderately " in normalized:
        positive_weight = 2

    if positive_weight is not None and any(token in normalized for token in ["interested", "likely"]):
        return (positive_weight, 1)

    for pattern, score in SCALE_ORDER_PATTERNS:
        if pattern in normalized:
            return (score, 9)

    return None


def count_scored_scale_hits(choices: list[str]) -> int:
    return sum(1 for choice in choices if scale_choice_score(choice) is not None)


def count_hp_interest_hits(choices: list[str]) -> int:
    hits = 0
    for choice in choices:
        normalized = normalize_sort_text(choice)
        if any(pattern in normalized for pattern, _ in HP_INTEREST_ORDER_PATTERNS):
            hits += 1
    return hits


def age_bucket_key(choice: str) -> tuple[int, int] | None:
    normalized = normalize_sort_text(choice)

    under_match = re.search(r"under\s+(\d+)", normalized)
    if under_match:
        return (0, int(under_match.group(1)))

    plus_match = re.search(r"(\d+)\s*\+", normalized)
    if plus_match:
        return (int(plus_match.group(1)), 999)

    range_match = re.search(r"(\d+)\s*-\s*(\d+)", normalized)
    if range_match:
        return (int(range_match.group(1)), int(range_match.group(2)))

    return None


def is_scale_answer_set(display_labels: list[str], question_label: str) -> bool:
    values = [
        normalize_sort_text(label)
        for label in display_labels
        if normalize_sort_text(label) and not is_exclusive_response_label(label)
    ]
    if not values:
        return False

    unique_values = list(dict.fromkeys(values))
    if len(unique_values) > 11:
        return False

    label_lower = normalize_sort_text(question_label)
    if any(token in label_lower for token in SCALE_LABEL_HINTS) and len(unique_values) <= 7:
        return True

    if count_pattern_hits(unique_values, LIKERT_PATTERNS) >= 2:
        return True

    if count_pattern_hits(unique_values, SCALE_VALUE_HINTS) >= 2 and len(unique_values) <= 7:
        return True

    scored_scale_hits = count_scored_scale_hits(unique_values)
    if scored_scale_hits >= max(3, len(unique_values) - 1) and len(unique_values) <= 7:
        return True

    if len(unique_values) in {4, 5} and scored_scale_hits >= 3:
        return True

    return False


def is_numeric_series(series: pd.Series) -> bool:
    coerced = pd.to_numeric(series, errors="coerce")
    non_null_ratio = coerced.notna().mean()
    unique_values = coerced.dropna().nunique()
    return float(non_null_ratio) >= 0.8 and int(unique_values) >= 8


def is_open_text_series(series: pd.Series) -> bool:
    values = series.dropna().map(normalize_answer).dropna()
    if values.empty:
        return False
    unique_ratio = values.nunique() / max(len(values), 1)
    avg_length = values.map(len).mean()
    return float(unique_ratio) >= 0.5 and float(avg_length) >= 15


def guess_question_type(
    df: pd.DataFrame,
    question: str,
    question_labels: dict[str, str],
    response_labels: dict[str, dict[str, str]],
    split_multi_select: bool,
    delimiter: str,
) -> str:
    if is_default_na_metadata_variable(question) or is_project_metadata_variable(question):
        return "Ignore"

    if question not in df.columns:
        return "Single-Select"

    question_label = display_question_label(question, question_labels)
    label_lower = normalize_sort_text(question_label)
    series = df[question]

    if "select all that apply" in label_lower or infer_multi_select(series):
        return "Multi-Select"

    options: list[str] = []
    for option in response_labels.get(question, {}):
        if option not in options:
            options.append(option)
    for option in extract_observed_response_options(series, split_multi_select, delimiter):
        if option not in options:
            options.append(option)

    display_labels = [
        display_response_label(question, option, response_labels)
        for option in options
    ]
    if is_scale_answer_set(display_labels, question_label):
        return "Scale / Likert"

    unique_values = [
        normalize_answer(value)
        for value in series.dropna().tolist()
        if normalize_answer(value)
    ]
    unique_values = list(dict.fromkeys(unique_values))
    numeric_like = pd.to_numeric(pd.Series(unique_values), errors="coerce")
    if numeric_like.notna().all() and 0 < len(unique_values) <= 10:
        return "Scale / Likert"

    if is_numeric_series(series):
        return "Numeric Data"

    if is_open_text_series(series):
        return "Open-End Text"

    return "Single-Select"


def normalize_question_type(value) -> str:
    question_type = normalize_answer(value)
    return question_type if question_type in QUESTION_TYPES else "Single-Select"


def sort_options_by_score(
    options: list[str],
    label_lookup: dict[str, str],
    score_function,
) -> list[str]:
    scored: list[tuple[object, int, str]] = []
    unmatched: list[tuple[int, str]] = []

    for index, option in enumerate(options):
        matched_score = score_function(label_lookup.get(option, option))
        if matched_score is None:
            unmatched.append((index, option))
        else:
            scored.append((matched_score, index, option))

    if not scored:
        return options

    ordered = [option for _, _, option in sorted(scored, key=lambda item: (item[0], item[1]))]
    ordered.extend(option for _, option in unmatched)
    return ordered


def sort_options_by_patterns(
    options: list[str],
    label_lookup: dict[str, str],
    ordered_patterns: list[tuple[str, int]],
) -> list[str]:
    def pattern_score(label: str) -> int | None:
        normalized = normalize_sort_text(label)
        for pattern, score in ordered_patterns:
            if pattern in normalized:
                return score
        return None

    return sort_options_by_score(options, label_lookup, pattern_score)


def anchor_exclusive_options_last(
    options: list[str],
    label_lookup: dict[str, str],
) -> list[str]:
    regular_options: list[str] = []
    exclusive_options: list[str] = []

    for option in options:
        label = label_lookup.get(option, option)
        if is_exclusive_response_label(label):
            exclusive_options.append(option)
        else:
            regular_options.append(option)

    return regular_options + exclusive_options


def sort_response_options(
    question: str,
    options: list[str],
    question_labels: dict[str, str],
    response_labels: dict[str, dict[str, str]],
    question_type: str | None = None,
) -> list[str]:
    if not options:
        return []

    question_type = normalize_question_type(question_type)
    question_label = display_question_label(question, question_labels)
    label_lookup = {
        option: display_response_label(question, option, response_labels)
        for option in options
    }
    display_labels = [label_lookup[option] for option in options]

    age_hits = sum(age_bucket_key(label) is not None for label in display_labels)
    if "age" in normalize_sort_text(question_label) or age_hits >= max(2, len(options) // 2):
        return anchor_exclusive_options_last(
            sort_options_by_score(options, label_lookup, age_bucket_key),
            label_lookup,
        )

    if (
        "relationship with the harry potter series" in normalize_sort_text(question_label)
        or count_hp_interest_hits(display_labels) >= 2
    ):
        return anchor_exclusive_options_last(
            sort_options_by_patterns(options, label_lookup, HP_INTEREST_ORDER_PATTERNS),
            label_lookup,
        )

    if question_type == "Scale / Likert" or is_scale_answer_set(display_labels, question_label):
        return anchor_exclusive_options_last(
            sort_options_by_score(options, label_lookup, scale_choice_score),
            label_lookup,
        )

    return anchor_exclusive_options_last(options, label_lookup)


def response_options_for_question(
    df: pd.DataFrame,
    question: str,
    response_labels: dict[str, dict[str, str]],
    split_multi_select: bool,
    delimiter: str,
    question_labels: dict[str, str] | None = None,
    question_type: str | None = None,
) -> list[str]:
    options: list[str] = []
    question_labels = question_labels or {}
    question_type = normalize_question_type(question_type)

    for option in response_labels.get(question, {}):
        if option not in options:
            options.append(option)

    if question in df.columns:
        for option in extract_observed_response_options(
            df[question],
            split_multi_select,
            delimiter,
        ):
            if option not in options:
                options.append(option)

    if question_type in {"Numeric Data", "Open-End Text", "Ignore"}:
        return []

    return sort_response_options(question, options, question_labels, response_labels, question_type)


def box_score_base_options(
    question: str,
    options: list[str],
    response_labels: dict[str, dict[str, str]],
) -> list[str]:
    return [
        option
        for option in options
        if not is_exclusive_response_label(
            display_response_label(question, option, response_labels)
        )
    ]


def apply_scale_box_score_selection(
    audit_frame: pd.DataFrame,
    selected_scores: list[str],
) -> pd.DataFrame:
    if audit_frame.empty or "Question Type" not in audit_frame.columns:
        return audit_frame

    updated_frame = audit_frame.copy()
    selected_lookup = {
        score
        for score in selected_scores
        if score in BOX_SCORE_OPTIONS
    }
    scale_mask = updated_frame["Question Type"].map(normalize_question_type) == "Scale / Likert"
    if "NA" in updated_frame.columns:
        scale_mask = scale_mask & ~updated_frame["NA"].astype(bool)

    for score in BOX_SCORE_OPTIONS:
        updated_frame.loc[scale_mask, score] = score in selected_lookup

    return updated_frame


def selected_box_score_options(box_score: str, base_options: list[str]) -> list[str]:
    if box_score == "T2B" and len(base_options) >= 2:
        return base_options[:2]
    if box_score == "T3B" and len(base_options) >= 3:
        return base_options[:3]
    if box_score == "B2B" and len(base_options) >= 2:
        return base_options[-2:]
    if box_score == "B3B" and len(base_options) >= 3:
        return base_options[-3:]
    return []


def box_score_response_label(
    question: str,
    box_score: str,
    selected_options: list[str],
    response_labels: dict[str, dict[str, str]],
) -> str:
    label = BOX_SCORE_LABELS.get(box_score, box_score)
    if not selected_options:
        return f"{label} ({NOT_AVAILABLE})"

    selected_labels = [
        display_response_label(question, option, response_labels)
        for option in selected_options
    ]
    return f"{label}: {' + '.join(selected_labels)}"


def box_score_count(
    group_df: pd.DataFrame,
    question: str,
    selected_options: list[str],
    delimiter: str,
) -> int:
    if question not in group_df.columns or not selected_options:
        return 0

    return int(
        group_df[question]
        .map(
            lambda value: any(
                value_matches_response_option(value, option, delimiter)
                for option in selected_options
            )
        )
        .sum()
    )


def box_scores_from_text(*values) -> list[str]:
    selected_scores: list[str] = []
    normalized_values = [
        re.sub(r"[^A-Z0-9]+", " ", (normalize_answer(value) or "").upper()).strip()
        for value in values
    ]
    combined_normalized = " ".join(value for value in normalized_values if value)

    for score in BOX_SCORE_OPTIONS:
        score_words = f"{score[0]} {score[1]} {score[2]}"
        score_found = False
        for normalized in normalized_values:
            normalized_compact = normalized.replace(" ", "")
            if (
                re.search(rf"(?<![A-Z0-9]){score}(?![A-Z0-9])", normalized)
                or re.search(rf"(?<![A-Z0-9]){score}(?![A-Z0-9])", normalized_compact)
                or normalized_compact.startswith(score)
                or normalized_compact.endswith(score)
                or re.search(rf"(?<![A-Z0-9]){score_words}(?![A-Z0-9])", normalized)
            ):
                score_found = True
                break

        if score_found:
            selected_scores.append(score)

    phrase_patterns = {
        "T2B": ["TOP 2 BOX", "TOP TWO BOX"],
        "T3B": ["TOP 3 BOX", "TOP THREE BOX"],
        "B2B": ["BOTTOM 2 BOX", "BOTTOM TWO BOX"],
        "B3B": ["BOTTOM 3 BOX", "BOTTOM THREE BOX"],
    }
    for score, phrases in phrase_patterns.items():
        if score in selected_scores:
            continue
        if any(phrase in combined_normalized for phrase in phrases):
            selected_scores.append(score)

    return selected_scores


def normalize_box_score_list(value) -> list[str]:
    if isinstance(value, str):
        raw_scores = re.split(r"[;,]", value)
    elif isinstance(value, (list, tuple, set)):
        raw_scores = value
    else:
        raw_scores = []

    selected_scores: list[str] = []
    for raw_score in raw_scores:
        score = normalize_answer(raw_score)
        if score in BOX_SCORE_OPTIONS and score not in selected_scores:
            selected_scores.append(score)
    return selected_scores


def box_score_display(scores: list[str] | tuple[str, ...]) -> str:
    return ", ".join(scores) if scores else "Full scale"


def norm_rule_is_included(rule: dict) -> bool:
    include_value = rule.get("Include", True)
    if isinstance(include_value, bool):
        return include_value
    normalized_value = normalize_answer(include_value)
    return (normalized_value or "").lower() not in {"false", "0", "no", "n"}


def load_saved_norm_rule_history() -> list[dict]:
    records = sorted(
        load_norm_database_manifest(),
        key=lambda record: normalize_answer(record.get("saved_at")) or "",
        reverse=True,
    )
    rule_history: list[dict] = []

    for record in records:
        dataset_id = normalize_answer(record.get("dataset_id"))
        if not dataset_id:
            continue

        dataset_path = norm_database_dataset_path(dataset_id)
        if not dataset_path.exists():
            continue

        rules = read_norm_dataset_rules(dataset_path)
        if not rules:
            continue

        question_labels = rules.get("question_labels", {})
        for rule in rules.get("norm_rules", []):
            source_variable = normalize_answer(rule.get("Source variable"))
            norm = normalize_answer(rule.get("Norm / benchmark")) or source_variable
            if (
                not source_variable
                or not norm
                or norm == NA_NORM_OPTION
                or not norm_rule_is_included(rule)
            ):
                continue

            question_label = (
                normalize_answer(question_labels.get(source_variable))
                or source_variable
            )
            rule_history.append(
                {
                    "Dataset ID": dataset_id,
                    "Saved at": record.get("saved_at", NOT_AVAILABLE),
                    "Uploaded file": record.get("uploaded_file", NOT_AVAILABLE),
                    "Source variable": source_variable,
                    "Question Text": question_label,
                    "Norm / benchmark": norm,
                    "Question Type": normalize_question_type(rule.get("Question Type")),
                    "Box scores": normalize_box_score_list(rule.get("Box scores")),
                    "_source_key": normalize_alias_key(source_variable),
                    "_question_text_key": normalize_alias_key(question_label),
                    "_norm_key": normalize_alias_key(norm),
                }
            )

    return rule_history


def prior_norm_rule_match_priority(
    prior_rule: dict,
    question: str,
    question_label: str,
    selected_norm: str,
) -> tuple[int, str] | None:
    question_key = normalize_alias_key(question)
    question_label_key = normalize_alias_key(question_label)
    selected_norm_key = normalize_alias_key(selected_norm)

    if question_key and question_key == prior_rule.get("_source_key"):
        return 0, "Source variable"
    if question_label_key and question_label_key == prior_rule.get("_question_text_key"):
        return 1, "Question text"
    if selected_norm_key and selected_norm_key == prior_rule.get("_norm_key"):
        return 2, "Norm / benchmark"
    return None


def prior_norm_rules_for_audit_row(
    question: str,
    question_label: str,
    selected_norm: str,
    prior_rules: list[dict],
) -> list[dict]:
    matches = []
    for position, prior_rule in enumerate(prior_rules):
        match = prior_norm_rule_match_priority(
            prior_rule,
            question,
            question_label,
            selected_norm,
        )
        if match is None:
            continue

        priority, reason = match
        matches.append(
            {
                **prior_rule,
                "_match_priority": priority,
                "_match_reason": reason,
                "_history_position": position,
            }
        )

    return sorted(
        matches,
        key=lambda rule: (rule["_match_priority"], rule["_history_position"]),
    )


def preferred_prior_norm_rule(
    question: str,
    question_label: str,
    selected_norm: str,
    prior_rules: list[dict],
) -> dict | None:
    matches = prior_norm_rules_for_audit_row(
        question,
        question_label,
        selected_norm,
        prior_rules,
    )
    return matches[0] if matches else None


def norm_audit_prior_rule_issues(
    editor_df: pd.DataFrame,
    prior_rules: list[dict],
) -> list[dict]:
    if editor_df.empty or not prior_rules:
        return []

    issues = []
    for row in editor_df.to_dict(orient="records"):
        question = normalize_answer(row.get("Variable Name"))
        question_label = normalize_answer(row.get("Question Text")) or question
        selected_norm = normalize_answer(row.get("Norm / benchmark"))
        is_na = (
            bool(row.get("NA"))
            or selected_norm == NA_NORM_OPTION
            or normalize_question_type(row.get("Question Type")) == "Ignore"
        )
        if not question or not selected_norm or is_na:
            continue

        current_box_scores = [
            score
            for score in BOX_SCORE_OPTIONS
            if bool(row.get(score))
        ]
        current_signature = tuple(current_box_scores)
        matches = prior_norm_rules_for_audit_row(
            question,
            question_label or question,
            selected_norm,
            prior_rules,
        )
        if not matches:
            continue

        differing_matches = []
        seen_signatures: set[tuple[str, ...]] = set()
        for match in matches:
            prior_signature = tuple(match.get("Box scores", []))
            if prior_signature == current_signature or prior_signature in seen_signatures:
                continue
            differing_matches.append(match)
            seen_signatures.add(prior_signature)

        if not differing_matches:
            continue

        previous_box_scores = [
            box_score_display(match.get("Box scores", []))
            for match in differing_matches[:4]
        ]
        dataset_labels = [
            f"{match.get('Uploaded file', NOT_AVAILABLE)} ({match.get('Saved at', NOT_AVAILABLE)})"
            for match in differing_matches[:4]
        ]
        issues.append(
            {
                "Variable Name": question,
                "Norm / benchmark": selected_norm,
                "Current box score": box_score_display(current_box_scores),
                "Previous saved box score": "; ".join(previous_box_scores),
                "Matched by": ", ".join(
                    dict.fromkeys(match.get("_match_reason", "") for match in differing_matches)
                ),
                "Saved dataset": "; ".join(dataset_labels),
            }
        )

    return issues


def build_norm_catalog(
    candidate_questions: list[str],
    saved_mappings: dict[str, str],
    prior_rules: list[dict] | None = None,
) -> list[str]:
    catalog: list[str] = []

    for target in saved_mappings.values():
        normalized_target = normalize_answer(target)
        if (
            normalized_target
            and normalized_target.upper() != NA_NORM_OPTION
            and normalized_target not in catalog
        ):
            catalog.append(normalized_target)

    for prior_rule in prior_rules or []:
        normalized_target = normalize_answer(prior_rule.get("Norm / benchmark"))
        if (
            normalized_target
            and normalized_target.upper() != NA_NORM_OPTION
            and normalized_target not in catalog
        ):
            catalog.append(normalized_target)

    for question in candidate_questions:
        if is_default_na_metadata_variable(question) or is_project_metadata_variable(question):
            continue
        if question not in catalog:
            catalog.append(question)

    return catalog


def closest_norm_benchmark(
    variable: str,
    question_label: str,
    norm_catalog: list[str],
    question_labels: dict[str, str],
) -> str:
    if variable in norm_catalog:
        return variable

    source_candidates = [variable, question_label]
    best_norm = variable
    best_score = 0.0

    for norm in norm_catalog:
        target_candidates = [norm, question_labels.get(norm, norm)]
        for source_value in source_candidates:
            source_key = normalize_column_name(source_value)
            if not source_key:
                continue
            for target_value in target_candidates:
                target_key = normalize_column_name(target_value)
                if not target_key:
                    continue
                score = SequenceMatcher(None, source_key, target_key).ratio()
                if score > best_score:
                    best_score = score
                    best_norm = norm

    return best_norm if best_score >= 0.55 else variable


def resolve_saved_norm_for_question(
    question: str,
    question_label: str,
    saved_mappings: dict[str, str],
    saved_na_aliases: set[str],
) -> str | None:
    exact_saved_norm = saved_mappings.get(question)
    if exact_saved_norm is not None:
        return exact_saved_norm

    candidate_aliases = {
        alias
        for alias in [
            normalize_alias_key(question),
            normalize_alias_key(question_label),
        ]
        if alias
    }
    if candidate_aliases & saved_na_aliases:
        return NA_NORM_OPTION

    for saved_question, saved_norm in saved_mappings.items():
        normalized_saved_norm = normalize_answer(saved_norm)
        if normalized_saved_norm != NA_NORM_OPTION:
            continue
        if normalize_alias_key(saved_question) in candidate_aliases:
            return NA_NORM_OPTION

    return None


def summarize_response_options(
    df: pd.DataFrame,
    question: str,
    response_labels: dict[str, dict[str, str]],
    split_multi_select: bool,
    delimiter: str,
    question_labels: dict[str, str] | None = None,
    question_type: str | None = None,
    max_options: int = 8,
) -> str:
    options = response_options_for_question(
        df,
        question,
        response_labels,
        split_multi_select,
        delimiter,
        question_labels,
        question_type,
    )
    labels = [display_response_label(question, option, response_labels) for option in options]
    if len(labels) > max_options:
        return " | ".join(labels[:max_options]) + f" | +{len(labels) - max_options} more"
    return " | ".join(labels)


def build_norm_audit_frame(
    df: pd.DataFrame,
    candidate_questions: list[str],
    question_labels: dict[str, str],
    response_labels: dict[str, dict[str, str]],
    saved_mappings: dict[str, str],
    saved_box_scores: dict[str, list[str]],
    saved_question_types: dict[str, str],
    saved_na_aliases: set[str],
    split_multi_select: bool,
    delimiter: str,
    prior_rules: list[dict] | None = None,
) -> pd.DataFrame:
    prior_rules = prior_rules or []
    norm_catalog = build_norm_catalog(candidate_questions, saved_mappings, prior_rules)
    rows = []

    for question in candidate_questions:
        question_label = display_question_label(question, question_labels)
        is_metadata_variable = is_default_na_metadata_variable(question)
        suggested_norm = (
            NA_NORM_OPTION
            if is_metadata_variable
            else closest_norm_benchmark(
                question,
                question_label,
                norm_catalog,
                question_labels,
            )
        )
        saved_norm = resolve_saved_norm_for_question(
            question,
            question_label,
            saved_mappings,
            saved_na_aliases,
        )
        selected_norm = saved_norm if saved_norm is not None else suggested_norm
        is_na = selected_norm == NA_NORM_OPTION
        prior_rule = preferred_prior_norm_rule(
            question,
            question_label,
            selected_norm,
            prior_rules,
        )
        detected_question_type = guess_question_type(
            df,
            question,
            question_labels,
            response_labels,
            split_multi_select,
            delimiter,
        )
        question_type = saved_question_types.get(
            question,
            (prior_rule or {}).get("Question Type", detected_question_type),
        )
        if is_metadata_variable:
            selected_box_scores = []
        elif question in saved_box_scores:
            selected_box_scores = saved_box_scores.get(question, [])
        elif prior_rule is not None:
            selected_box_scores = prior_rule.get("Box scores", [])
        else:
            selected_box_scores = box_scores_from_text(
                selected_norm,
                suggested_norm,
                question,
                question_label,
            )

        rows.append(
            {
                "NA": is_na,
                "Variable Name": question,
                "Question Text": question_label,
                "Question Type": question_type,
                "Suggested norm/benchmark": suggested_norm,
                "Norm / benchmark": selected_norm,
                "T2B": "T2B" in selected_box_scores,
                "T3B": "T3B" in selected_box_scores,
                "B2B": "B2B" in selected_box_scores,
                "B3B": "B3B" in selected_box_scores,
                "Answer Choices Count": len(
                    response_options_for_question(
                        df,
                        question,
                        response_labels,
                        split_multi_select,
                        delimiter,
                        question_labels,
                        question_type,
                    )
                ),
                "Answer Choices": summarize_response_options(
                    df,
                    question,
                    response_labels,
                    split_multi_select,
                    delimiter,
                    question_labels,
                    question_type,
                ),
            }
        )

    return pd.DataFrame(rows)


def normalize_box_score_audit_editor(editor_df: pd.DataFrame) -> dict[str, list[str]]:
    settings: dict[str, list[str]] = {}
    if editor_df.empty:
        return settings

    for row in editor_df.to_dict(orient="records"):
        variable = normalize_answer(row.get("Variable Name"))
        if not variable:
            continue

        if normalize_question_type(row.get("Question Type")) == "Ignore":
            settings[variable] = []
            continue

        selected_scores = [
            score
            for score in BOX_SCORE_OPTIONS
            if bool(row.get(score))
        ]
        settings[variable] = selected_scores

    return settings


def normalize_question_type_audit_editor(editor_df: pd.DataFrame) -> dict[str, str]:
    settings: dict[str, str] = {}
    if editor_df.empty:
        return settings

    for row in editor_df.to_dict(orient="records"):
        variable = normalize_answer(row.get("Variable Name"))
        if not variable:
            continue
        settings[variable] = normalize_question_type(row.get("Question Type"))

    return settings


def normalize_na_aliases_from_audit_editor(editor_df: pd.DataFrame) -> set[str]:
    aliases: set[str] = set()
    if editor_df.empty:
        return aliases

    for row in editor_df.to_dict(orient="records"):
        selected_norm = normalize_answer(row.get("Norm / benchmark"))
        is_na = (
            bool(row.get("NA"))
            or selected_norm == NA_NORM_OPTION
            or normalize_question_type(row.get("Question Type")) == "Ignore"
        )
        if not is_na:
            continue

        for value in [row.get("Variable Name"), row.get("Question Text")]:
            alias = normalize_alias_key(value)
            if alias:
                aliases.add(alias)

    return aliases


def build_saved_norm_review_frame(
    saved_mappings: dict[str, str],
    saved_box_scores: dict[str, list[str]],
    saved_question_types: dict[str, str],
    denominator_settings: dict[str, str],
) -> pd.DataFrame:
    variables = sorted(
        set(saved_mappings)
        | set(saved_box_scores)
        | set(saved_question_types)
        | set(denominator_settings)
    )
    rows = []

    for variable in variables:
        mapped_norm = saved_mappings.get(variable, "")
        is_na = mapped_norm == NA_NORM_OPTION
        box_scores = [
            score
            for score in saved_box_scores.get(variable, [])
            if score in BOX_SCORE_OPTIONS
        ]
        rows.append(
            {
                "Variable Name": variable,
                "Norm / benchmark": mapped_norm or NOT_AVAILABLE,
                "NA": is_na,
                "Question Type": saved_question_types.get(variable, NOT_AVAILABLE),
                "Box scores": ", ".join(box_scores) if box_scores else "None",
                "Denominator": denominator_settings.get(variable, DEFAULT_DENOMINATOR),
            }
        )

    return pd.DataFrame(rows)


def normalize_norm_audit_editor(editor_df: pd.DataFrame) -> dict[str, str]:
    mappings: dict[str, str] = {}
    if editor_df.empty:
        return mappings

    for row in editor_df.to_dict(orient="records"):
        variable = normalize_answer(row.get("Variable Name"))
        selected_norm = normalize_answer(row.get("Norm / benchmark"))
        is_na = (
            bool(row.get("NA"))
            or selected_norm == NA_NORM_OPTION
            or normalize_question_type(row.get("Question Type")) == "Ignore"
        )

        if not variable:
            continue

        mappings[variable] = NA_NORM_OPTION if is_na else (selected_norm or variable)

    return mappings


def included_norm_mappings(audit_mappings: dict[str, str]) -> dict[str, str]:
    return {
        variable: norm
        for variable, norm in audit_mappings.items()
        if norm and norm != NA_NORM_OPTION
    }


def normalize_column_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value).lower()).strip("_")


def normalize_alias_key(value) -> str:
    normalized = normalize_answer(value)
    if not normalized:
        return ""
    return normalize_column_name(normalized)


def make_unique_column_names(raw_values: list[object]) -> list[str]:
    names: list[str] = []
    used_counts: dict[str, int] = {}

    for index, value in enumerate(raw_values, start=1):
        base_name = normalize_answer(value) or f"column_{index}"
        count = used_counts.get(base_name, 0)
        column_name = base_name if count == 0 else f"{base_name}_{count + 1}"
        used_counts[base_name] = count + 1
        names.append(column_name)

    return names


def is_default_na_metadata_variable(column) -> bool:
    blacklist = {value.lower() for value in DEFAULT_VARIABLE_BLACKLIST}
    blacklist_prefixes = [value.lower() for value in DEFAULT_BLACKLIST_PREFIXES]
    normalized = normalize_answer(column) or ""
    normalized_lower = normalized.lower()

    return normalized_lower in blacklist or any(
        normalized_lower.startswith(prefix) for prefix in blacklist_prefixes
    )


def is_project_metadata_variable(column) -> bool:
    return normalize_alias_key(column) in {
        normalize_alias_key(variable)
        for variable in PROJECT_METADATA_VARIABLES
    }


def identify_metadata_columns(columns) -> list[str]:
    return [column for column in columns if is_default_na_metadata_variable(column)]


def default_group_column_index(columns: list[str]) -> int:
    normalized_columns = {
        normalize_column_name(str(column)): index
        for index, column in enumerate(columns)
    }

    for candidate in GROUP_COLUMN_CANDIDATES:
        index = normalized_columns.get(normalize_column_name(candidate))
        if index is not None:
            return index

    for index, column in enumerate(columns):
        if not is_default_na_metadata_variable(column) and not is_project_metadata_variable(column):
            return index

    return 0


def infer_column(columns: list[str], candidates: list[str]) -> str | None:
    normalized_columns = {normalize_column_name(column): column for column in columns}
    for candidate in candidates:
        match = normalized_columns.get(normalize_column_name(candidate))
        if match:
            return match
    return None


def build_response_label_maps(
    label_df: pd.DataFrame | None,
    question_column: str | None,
    value_column: str | None,
    label_column: str | None,
    question_label_column: str | None,
) -> tuple[dict[str, dict[str, str]], dict[str, str]]:
    response_labels: dict[str, dict[str, str]] = {}
    question_labels: dict[str, str] = {}

    if (
        label_df is None
        or label_df.empty
        or not question_column
        or not value_column
        or not label_column
    ):
        return response_labels, question_labels

    for _, row in label_df.iterrows():
        question = normalize_answer(row.get(question_column))
        value = normalize_answer(row.get(value_column))
        label = normalize_answer(row.get(label_column))

        if question_label_column:
            question_label = normalize_answer(row.get(question_label_column))
            if question and question_label:
                question_labels[question] = question_label

        if question and value and label:
            response_labels.setdefault(question, {})[value] = label

    return response_labels, question_labels


def display_response_label(
    question: str,
    value: str,
    response_labels: dict[str, dict[str, str]],
) -> str:
    return response_labels.get(question, {}).get(value, value)


def display_question_label(question: str, question_labels: dict[str, str]) -> str:
    return question_labels.get(question, question)


def display_norm_mapping_label(
    source_question: str,
    mapped_norm: str,
    question_labels: dict[str, str],
) -> str:
    source_label = display_question_label(source_question, question_labels)
    mapped_label = display_question_label(mapped_norm, question_labels)
    if source_question == mapped_norm:
        return source_label
    return f"{source_label} -> {mapped_label}"


def display_norm_variable_mapping_label(source_question: str, mapped_norm: str) -> str:
    if source_question == mapped_norm:
        return source_question
    return f"{source_question} -> {mapped_norm}"


def looks_like_metadata_row(row: pd.Series, header_values: list[str], row_index: int) -> bool:
    values = [normalize_answer(value) or "" for value in row.tolist()]
    non_empty_values = [value for value in values if value]
    if not non_empty_values:
        return True

    import_id_hits = sum("importid" in value.lower() for value in non_empty_values)
    header_overlap = sum(value in header_values for value in non_empty_values)
    metadata_hits = sum(
        any(token in value.lower() for token in ["qualtrics", "metadata", "question text", "survey"])
        for value in non_empty_values
    )

    if row_index == 2 and import_id_hits >= max(1, len(non_empty_values) // 3):
        return True
    if header_overlap >= max(2, len(non_empty_values) // 2):
        return True
    if metadata_hits >= max(2, len(non_empty_values) // 2):
        return True
    return False


def prepare_smart_tables_layout(raw_df: pd.DataFrame) -> RespondentSheet:
    if raw_df.shape[0] < 2:
        raise ValueError("The selected sheet needs at least a variable row and a question-label row.")

    header_values = [normalize_answer(value) or "" for value in raw_df.iloc[0].tolist()]
    if not any(header_values):
        raise ValueError("The first row does not contain usable variable names.")

    column_names = make_unique_column_names(raw_df.iloc[0].tolist())
    question_label_row = raw_df.iloc[1].tolist()
    question_labels = {
        column: normalize_answer(label) or column
        for column, label in zip(column_names, question_label_row)
    }

    df = raw_df.copy()
    df.columns = column_names

    rows_to_drop = [0, 1]
    metadata_rows_removed = 0
    for row_index in range(2, len(df)):
        if looks_like_metadata_row(df.iloc[row_index], header_values, row_index):
            rows_to_drop.append(row_index)
            metadata_rows_removed += 1
        else:
            break

    cleaned_df = df.drop(index=rows_to_drop).reset_index(drop=True)
    metadata_columns = identify_metadata_columns(cleaned_df.columns)
    return RespondentSheet(
        dataframe=cleaned_df,
        question_labels=question_labels,
        metadata_rows_removed=metadata_rows_removed,
        metadata_columns_default_na=len(metadata_columns),
        metadata_columns=metadata_columns,
    )


def safe_rate(count: int, denominator: int | None) -> float | None:
    if denominator is None or denominator <= 0:
        return None
    return count / denominator


def two_proportion_z_test(
    control_count: int,
    control_n: int | None,
    test_count: int,
    test_n: int | None,
) -> str:
    if control_n is None or test_n is None or control_n <= 0 or test_n <= 0:
        return NOT_TESTED

    if control_count > control_n or test_count > test_n:
        return NOT_TESTED

    pooled_n = control_n + test_n
    pooled_rate = (control_count + test_count) / pooled_n
    standard_error = math.sqrt(
        pooled_rate * (1 - pooled_rate) * ((1 / control_n) + (1 / test_n))
    )

    if standard_error == 0:
        return "Not significant"

    control_rate = control_count / control_n
    test_rate = test_count / test_n
    z_score = (test_rate - control_rate) / standard_error
    p_value = math.erfc(abs(z_score) / math.sqrt(2))

    return "Significant" if p_value < SIGNIFICANCE_ALPHA else "Not significant"


def format_percent(value: float | None) -> str:
    if value is None:
        return NOT_AVAILABLE
    return format_percent_points(value * 100)


def round_percentage_points(value: float) -> int:
    if value >= 0:
        return int(math.floor(value + 0.5))
    return int(math.ceil(value - 0.5))


def format_lift_points(points: float | None) -> str:
    if points is None:
        return NOT_AVAILABLE
    return f"{round_percentage_points(points):+d}pts"


def format_percent_points(points: float | None) -> str:
    if points is None:
        return NOT_AVAILABLE
    return f"{round_percentage_points(points):d}%"


def format_lift(control_rate: float | None, test_rate: float | None) -> str:
    if control_rate is None or test_rate is None:
        return NOT_AVAILABLE
    return format_lift_points((test_rate - control_rate) * 100)


def normalize_percent_cell(value: object) -> object:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return value

    text = str(value).strip()
    if not text or text in {NOT_AVAILABLE, NOT_TESTED}:
        return value

    match = re.fullmatch(r"([+-]?\d+(?:\.\d+)?)\s*%", text)
    if not match:
        return value

    return format_percent_points(float(match.group(1)))


def normalize_lift_cell(value: object) -> object:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return value

    if isinstance(value, (int, float)):
        return format_lift_points(float(value))

    text = str(value).strip()
    if not text or text in {NOT_AVAILABLE, NOT_TESTED}:
        return value

    match = re.fullmatch(
        r"([+-]?\d+(?:\.\d+)?)\s*(?:pts?|pp|percentage points)?",
        text,
        flags=re.IGNORECASE,
    )
    if not match:
        return value

    return format_lift_points(float(match.group(1)))


def normalize_lift_output_table(table: pd.DataFrame) -> pd.DataFrame:
    output_table = table.copy()
    for column in ["Control", "Test"]:
        if column in output_table.columns:
            output_table[column] = output_table[column].map(normalize_percent_cell)
    if "Lift" in output_table.columns:
        output_table["Lift"] = output_table["Lift"].map(normalize_lift_cell)
    return output_table


def parse_percent_points(value: object) -> float | None:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None

    if isinstance(value, (int, float)):
        numeric = float(value)
        return numeric * 100 if -1 <= numeric <= 1 else numeric

    text = str(value).strip()
    if not text or text in {NOT_AVAILABLE, NOT_TESTED}:
        return None

    match = re.fullmatch(r"([+-]?\d+(?:\.\d+)?)\s*%", text)
    if not match:
        return None
    return float(match.group(1))


def parse_lift_points(value: object) -> float | None:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text or text in {NOT_AVAILABLE, NOT_TESTED}:
        return None

    match = re.fullmatch(
        r"([+-]?\d+(?:\.\d+)?)\s*(?:pts?|pp|percentage points)?",
        text,
        flags=re.IGNORECASE,
    )
    if not match:
        return None
    return float(match.group(1))


def short_chart_label(label: object, max_length: int = 34) -> str:
    text = normalize_answer(label) or NOT_AVAILABLE
    if len(text) <= max_length:
        return text
    return f"{text[: max_length - 3].rstrip()}..."


def norm_chart_rows(table: pd.DataFrame) -> list[dict]:
    display_table = normalize_lift_output_table(table)
    chart_rows = []
    for row in display_table.to_dict("records"):
        label = normalize_answer(row.get("Response option"))
        if not label or label.lower().startswith("base size"):
            continue

        control_points = parse_percent_points(row.get("Control"))
        test_points = parse_percent_points(row.get("Test"))
        if control_points is None or test_points is None:
            continue

        lift_points = parse_lift_points(row.get("Lift"))
        if lift_points is None:
            lift_points = test_points - control_points

        chart_rows.append(
            {
                "label": label,
                "chart_label": short_chart_label(label),
                "control_points": control_points,
                "test_points": test_points,
                "lift_points": lift_points,
                "significant": (
                    normalize_answer(row.get("Significance result")).lower()
                    == "significant"
                ),
            }
        )

    return chart_rows


def norm_chart_display_rows(table: pd.DataFrame) -> list[dict]:
    chart_rows = norm_chart_rows(table)
    for row in chart_rows:
        row["chart_label"] = row["label"]
    return chart_rows



def calculate_norm_table(
    df: pd.DataFrame,
    question: str,
    group_column: str,
    control_label: str | None,
    test_label: str | None,
    denominator_setting: str,
    split_multi_select: bool,
    delimiter: str,
    response_labels: dict[str, dict[str, str]] | None = None,
    question_labels: dict[str, str] | None = None,
    box_scores: list[str] | None = None,
    question_type: str | None = None,
) -> pd.DataFrame:
    response_labels = response_labels or {}
    question_labels = question_labels or {}
    question_type = normalize_question_type(question_type)
    box_scores = [
        score
        for score in (box_scores or [])
        if score in BOX_SCORE_OPTIONS
    ]
    required_labels_missing = (
        not group_column
        or group_column not in df.columns
        or not control_label
        or not test_label
    )

    group_values = (
        df[group_column].map(normalize_answer)
        if group_column in df.columns
        else pd.Series([], dtype=object)
    )
    control_df = df[group_values == control_label] if not required_labels_missing else pd.DataFrame()
    test_df = df[group_values == test_label] if not required_labels_missing else pd.DataFrame()

    control_n = denominator_for(control_df, question, denominator_setting)
    test_n = denominator_for(test_df, question, denominator_setting)
    options = response_options_for_question(
        df,
        question,
        response_labels,
        split_multi_select,
        delimiter,
        question_labels,
        question_type,
    )
    control_counts = response_counts(
        control_df,
        question,
        split_multi_select,
        delimiter,
        options,
    )
    test_counts = response_counts(
        test_df,
        question,
        split_multi_select,
        delimiter,
        options,
    )
    if not options:
        options = option_order(control_counts, test_counts)

    rows = [
        {
            "Response option": f"Base size ({denominator_setting})",
            "Control": control_n if control_n is not None else NOT_AVAILABLE,
            "Test": test_n if test_n is not None else NOT_AVAILABLE,
            "Lift": NOT_AVAILABLE,
            "Significance result": NOT_TESTED,
        }
    ]

    if not options:
        return pd.DataFrame(rows)

    base_box_options = box_score_base_options(question, options, response_labels)
    if not box_scores:
        for option in options:
            response_option = display_response_label(question, option, response_labels)
            control_count = control_counts.get(option, 0)
            test_count = test_counts.get(option, 0)
            control_rate = safe_rate(control_count, control_n)
            test_rate = safe_rate(test_count, test_n)
            significance = (
                NOT_TESTED
                if required_labels_missing
                else two_proportion_z_test(control_count, control_n, test_count, test_n)
            )

            rows.append(
                {
                    "Response option": response_option,
                    "Control": format_percent(control_rate),
                    "Test": format_percent(test_rate),
                    "Lift": format_lift(control_rate, test_rate),
                    "Significance result": significance,
                }
            )

    for box_score in box_scores:
        selected_options = selected_box_score_options(box_score, base_box_options)
        response_option = box_score_response_label(
            question,
            box_score,
            selected_options,
            response_labels,
        )

        if selected_options:
            control_count = box_score_count(control_df, question, selected_options, delimiter)
            test_count = box_score_count(test_df, question, selected_options, delimiter)
            control_rate = safe_rate(control_count, control_n)
            test_rate = safe_rate(test_count, test_n)
            significance = (
                NOT_TESTED
                if required_labels_missing
                else two_proportion_z_test(control_count, control_n, test_count, test_n)
            )
        else:
            control_rate = None
            test_rate = None
            significance = NOT_TESTED

        rows.append(
            {
                "Response option": response_option,
                "Control": format_percent(control_rate),
                "Test": format_percent(test_rate),
                "Lift": format_lift(control_rate, test_rate),
                "Significance result": significance,
            }
        )

    return pd.DataFrame(rows)


def calculate_norm_count_summary(
    df: pd.DataFrame,
    question: str,
    group_column: str,
    control_label: str | None,
    test_label: str | None,
    denominator_setting: str,
    split_multi_select: bool,
    delimiter: str,
    response_labels: dict[str, dict[str, str]] | None = None,
    question_labels: dict[str, str] | None = None,
    box_scores: list[str] | None = None,
    question_type: str | None = None,
) -> dict:
    response_labels = response_labels or {}
    question_labels = question_labels or {}
    question_type = normalize_question_type(question_type)
    box_scores = [
        score
        for score in (box_scores or [])
        if score in BOX_SCORE_OPTIONS
    ]
    required_labels_missing = (
        not group_column
        or group_column not in df.columns
        or not control_label
        or not test_label
    )

    group_values = (
        df[group_column].map(normalize_answer)
        if group_column in df.columns
        else pd.Series([], dtype=object)
    )
    control_df = df[group_values == control_label] if not required_labels_missing else pd.DataFrame()
    test_df = df[group_values == test_label] if not required_labels_missing else pd.DataFrame()
    control_n = denominator_for(control_df, question, denominator_setting) or 0
    test_n = denominator_for(test_df, question, denominator_setting) or 0

    options = response_options_for_question(
        df,
        question,
        response_labels,
        split_multi_select,
        delimiter,
        question_labels,
        question_type,
    )
    control_counts = response_counts(
        control_df,
        question,
        split_multi_select,
        delimiter,
        options,
    )
    test_counts = response_counts(
        test_df,
        question,
        split_multi_select,
        delimiter,
        options,
    )
    if not options:
        options = option_order(control_counts, test_counts)

    rows = []
    base_box_options = box_score_base_options(question, options, response_labels)
    if not box_scores:
        for option in options:
            rows.append(
                {
                    "Response option": display_response_label(question, option, response_labels),
                    "Control count": control_counts.get(option, 0),
                    "Test count": test_counts.get(option, 0),
                }
            )
    else:
        for box_score in box_scores:
            selected_options = selected_box_score_options(box_score, base_box_options)
            rows.append(
                {
                    "Response option": box_score_response_label(
                        question,
                        box_score,
                        selected_options,
                        response_labels,
                    ),
                    "Control count": (
                        box_score_count(control_df, question, selected_options, delimiter)
                        if selected_options
                        else 0
                    ),
                    "Test count": (
                        box_score_count(test_df, question, selected_options, delimiter)
                        if selected_options
                        else 0
                    ),
                }
            )

    return {
        "Control denominator": control_n,
        "Test denominator": test_n,
        "Denominator": denominator_setting,
        "Rows": rows,
    }


def combined_norm_table_from_counts(
    metric_counts: dict,
) -> pd.DataFrame:
    control_n = metric_counts.get("Control denominator", 0)
    test_n = metric_counts.get("Test denominator", 0)
    denominators = metric_counts.get("Denominators", set())
    denominator_label = (
        next(iter(denominators))
        if len(denominators) == 1
        else "Mixed denominator"
    )

    rows = [
        {
            "Response option": f"Base size ({denominator_label})",
            "Control": control_n,
            "Test": test_n,
            "Lift": NOT_AVAILABLE,
            "Significance result": NOT_TESTED,
        }
    ]

    for response_option, counts in metric_counts.get("Rows", {}).items():
        control_count = counts.get("Control count", 0)
        test_count = counts.get("Test count", 0)
        control_rate = safe_rate(control_count, control_n)
        test_rate = safe_rate(test_count, test_n)
        rows.append(
            {
                "Response option": response_option,
                "Control": format_percent(control_rate),
                "Test": format_percent(test_rate),
                "Lift": format_lift(control_rate, test_rate),
                "Significance result": two_proportion_z_test(
                    control_count,
                    control_n,
                    test_count,
                    test_n,
                ),
            }
        )

    return pd.DataFrame(rows)


def build_combined_saved_norm_tables(
    saved_records: list[dict],
    selected_filters: dict[str, list[str]] | None = None,
) -> tuple[dict[str, pd.DataFrame], dict[str, list[str]]]:
    selected_filters = selected_filters or {}
    combined_counts: dict[str, dict] = {}
    source_variables: dict[str, list[str]] = {}

    for record in saved_records:
        dataset_id = normalize_answer(record.get("dataset_id"))
        if not dataset_id:
            continue

        dataset_path = norm_database_dataset_path(dataset_id)
        rules = read_norm_dataset_rules(dataset_path)
        if rules is None:
            continue

        try:
            source_data = pd.read_excel(
                dataset_path,
                sheet_name=NORM_DATASET_RESPONDENT_SHEET,
                dtype=object,
            )
        except Exception:
            continue

        group_column = rules.get("group_column")
        control_label = rules.get("control_label")
        test_label = rules.get("test_label")
        response_labels = rules.get("response_labels", {})
        question_labels = rules.get("question_labels", {})
        source_data = apply_saved_norm_filters(
            source_data,
            question_labels,
            selected_filters,
        )
        if source_data.empty:
            continue

        for rule in rules.get("norm_rules", []):
            source_variable = normalize_answer(rule.get("Source variable"))
            metric = normalize_answer(rule.get("Norm / benchmark")) or source_variable
            if (
                not source_variable
                or source_variable not in source_data.columns
                or not norm_rule_is_included(rule)
                or metric == NA_NORM_OPTION
            ):
                continue

            denominator = rule.get("Denominator", DEFAULT_DENOMINATOR)
            if denominator not in DENOMINATOR_OPTIONS:
                denominator = DEFAULT_DENOMINATOR
            summary = calculate_norm_count_summary(
                source_data,
                source_variable,
                group_column,
                control_label,
                test_label,
                denominator,
                rules.get("split_multi_select", False),
                rules.get("delimiter", ";"),
                response_labels,
                question_labels,
                normalize_box_score_list(rule.get("Box scores")),
                rule.get("Question Type"),
            )

            metric_counts = combined_counts.setdefault(
                metric,
                {
                    "Control denominator": 0,
                    "Test denominator": 0,
                    "Denominators": set(),
                    "Rows": {},
                },
            )
            metric_counts["Control denominator"] += summary["Control denominator"]
            metric_counts["Test denominator"] += summary["Test denominator"]
            metric_counts["Denominators"].add(summary["Denominator"])
            source_variables.setdefault(metric, [])
            if source_variable not in source_variables[metric]:
                source_variables[metric].append(source_variable)

            for row in summary["Rows"]:
                response_option = row["Response option"]
                row_counts = metric_counts["Rows"].setdefault(
                    response_option,
                    {
                        "Control count": 0,
                        "Test count": 0,
                    },
                )
                row_counts["Control count"] += row["Control count"]
                row_counts["Test count"] += row["Test count"]

    tables = {
        metric: combined_norm_table_from_counts(metric_counts)
        for metric, metric_counts in combined_counts.items()
    }
    return tables, source_variables


def saved_filter_column_for_label(
    df: pd.DataFrame,
    question_labels: dict[str, str],
    label: str,
) -> str | None:
    for field_label, aliases in NORM_FILTER_FIELDS:
        if field_label == label:
            return filter_column_for_field(
                df,
                question_labels,
                aliases,
                exact_match=label in EXACT_NORM_FILTERS,
            )
    return None


def apply_saved_norm_filters(
    df: pd.DataFrame,
    question_labels: dict[str, str],
    selected_filters: dict[str, list[str]],
) -> pd.DataFrame:
    filtered_df = df
    for label, selected_values in selected_filters.items():
        if not selected_values:
            continue

        column = saved_filter_column_for_label(filtered_df, question_labels, label)
        if column is None:
            return filtered_df.iloc[0:0]

        filtered_df = apply_norm_filters(filtered_df, {column: selected_values})
        if filtered_df.empty:
            return filtered_df

    return filtered_df


def saved_norm_filter_options(
    saved_records: list[dict],
) -> tuple[dict[str, list[str]], dict[str, list[str]]]:
    options_by_label: dict[str, set[str]] = {
        label: set()
        for label, _aliases in NORM_FILTER_FIELDS
    }
    columns_by_label: dict[str, set[str]] = {
        label: set()
        for label, _aliases in NORM_FILTER_FIELDS
    }

    for record in saved_records:
        dataset_id = normalize_answer(record.get("dataset_id"))
        if not dataset_id:
            continue

        dataset_path = norm_database_dataset_path(dataset_id)
        rules = read_norm_dataset_rules(dataset_path)
        if rules is None:
            continue

        try:
            source_data = pd.read_excel(
                dataset_path,
                sheet_name=NORM_DATASET_RESPONDENT_SHEET,
                dtype=object,
            )
        except Exception:
            continue

        question_labels = rules.get("question_labels", {})
        for label, _aliases in NORM_FILTER_FIELDS:
            column = saved_filter_column_for_label(source_data, question_labels, label)
            if column is None:
                continue

            columns_by_label[label].add(column)
            options_by_label[label].update(filter_option_values(source_data[column]))

    return (
        {
            label: filter_option_values(pd.Series(list(values), dtype=object))
            for label, values in options_by_label.items()
        },
        {
            label: sorted(values)
            for label, values in columns_by_label.items()
        },
    )


def saved_norm_filter_totals(
    saved_records: list[dict],
    selected_filters: dict[str, list[str]],
) -> dict[str, int]:
    totals = {
        "datasets": 0,
        "rows": 0,
        "control_rows": 0,
        "test_rows": 0,
    }

    for record in saved_records:
        dataset_id = normalize_answer(record.get("dataset_id"))
        if not dataset_id:
            continue

        dataset_path = norm_database_dataset_path(dataset_id)
        rules = read_norm_dataset_rules(dataset_path)
        if rules is None:
            continue

        try:
            source_data = pd.read_excel(
                dataset_path,
                sheet_name=NORM_DATASET_RESPONDENT_SHEET,
                dtype=object,
            )
        except Exception:
            continue

        filtered_data = apply_saved_norm_filters(
            source_data,
            rules.get("question_labels", {}),
            selected_filters,
        )
        if filtered_data.empty:
            continue

        totals["datasets"] += 1
        totals["rows"] += int(len(filtered_data))

        group_column = rules.get("group_column")
        control_label = rules.get("control_label")
        test_label = rules.get("test_label")
        if group_column in filtered_data.columns:
            group_values = filtered_data[group_column].map(normalize_answer)
            totals["control_rows"] += int((group_values == control_label).sum())
            totals["test_rows"] += int((group_values == test_label).sum())

    return totals


def build_norm_tables(
    df: pd.DataFrame,
    norm_questions: list[str],
    included_mappings: dict[str, str],
    group_column: str,
    control_label: str | None,
    test_label: str | None,
    denominator_settings: dict[str, str],
    split_multi_select: bool,
    delimiter: str,
    response_labels: dict[str, dict[str, str]],
    question_labels: dict[str, str],
    box_score_settings: dict[str, list[str]],
    saved_box_score_settings: dict[str, list[str]],
    question_type_settings: dict[str, str],
    saved_question_type_settings: dict[str, str],
) -> dict[str, pd.DataFrame]:
    tables: dict[str, pd.DataFrame] = {}

    for question in norm_questions:
        mapped_norm = included_mappings.get(question, question)
        denominator_setting = denominator_settings.get(question, DEFAULT_DENOMINATOR)
        tables[f"{mapped_norm}__{question}"] = calculate_norm_table(
            df,
            question,
            group_column,
            control_label,
            test_label,
            denominator_setting,
            split_multi_select,
            delimiter,
            response_labels,
            question_labels,
            box_score_settings.get(
                question,
                saved_box_score_settings.get(question, []),
            ),
            question_type_settings.get(
                question,
                saved_question_type_settings.get(question),
            ),
        )

    return tables


def markdown_escape(value: str) -> str:
    return value.replace("|", "\\|").replace("\n", " ")


def append_changelog_entries(changes: list[tuple[str, str | None, str]]) -> None:
    if not changes:
        return

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    entries = [
        f"- {timestamp}: Denominator setting for `{question}` changed from "
        f"`{old or 'Default: Total answering'}` to `{new}`."
        for question, old, new in changes
    ]

    if CHANGELOG_PATH.exists():
        content = CHANGELOG_PATH.read_text()
    else:
        content = "# Changelog\n\n## Unreleased\n"

    marker = "## Unreleased"
    if marker in content:
        content = content.replace(marker, marker + "\n" + "\n".join(entries), 1)
    else:
        content = content.rstrip() + "\n\n## Unreleased\n" + "\n".join(entries) + "\n"

    CHANGELOG_PATH.write_text(content if content.endswith("\n") else content + "\n")


def update_status_denominator_snapshot(settings: dict[str, str]) -> None:
    start_marker = "<!-- denominator-settings:start -->"
    end_marker = "<!-- denominator-settings:end -->"

    rows = ["| Norm question | Denominator |", "| --- | --- |"]
    if settings:
        rows.extend(
            f"| {markdown_escape(question)} | {markdown_escape(setting)} |"
            for question, setting in sorted(settings.items())
        )
    else:
        rows.append("| No saved overrides yet | Total answering default |")

    snapshot = f"{start_marker}\n" + "\n".join(rows) + f"\n{end_marker}"

    if STATUS_PATH.exists():
        content = STATUS_PATH.read_text()
    else:
        content = "# Status\n\n## Denominator Settings\n\n"

    if start_marker in content and end_marker in content:
        before = content.split(start_marker, 1)[0]
        after = content.split(end_marker, 1)[1]
        content = before + snapshot + after
    else:
        content = content.rstrip() + "\n\n## Denominator Settings\n\n" + snapshot + "\n"

    STATUS_PATH.write_text(content if content.endswith("\n") else content + "\n")


def persist_denominator_changes(
    previous_settings: dict[str, str],
    selected_settings: dict[str, str],
) -> None:
    merged_settings = {**previous_settings, **selected_settings}
    changes = [
        (question, previous_settings.get(question), setting)
        for question, setting in selected_settings.items()
        if previous_settings.get(question, DEFAULT_DENOMINATOR) != setting
    ]

    save_denominator_settings(merged_settings)
    append_changelog_entries(changes)
    update_status_denominator_snapshot(merged_settings)


def read_excel_workbook(uploaded_file) -> tuple[bytes, pd.ExcelFile] | tuple[None, None]:
    workbook_bytes = uploaded_file.getvalue()
    try:
        return workbook_bytes, pd.ExcelFile(BytesIO(workbook_bytes))
    except Exception as exc:
        st.error(f"Could not read Excel workbook: {exc}")
        return None, None


def cache_uploaded_workbook_in_session() -> None:
    uploaded_file = st.session_state.get(UPLOAD_WORKBOOK_SESSION_KEY)
    if uploaded_file is None:
        return

    try:
        workbook_bytes = uploaded_file.getvalue()
    except Exception:
        return

    if not workbook_bytes:
        return

    st.session_state[ACTIVE_WORKBOOK_NAME_SESSION_KEY] = uploaded_file.name
    st.session_state[ACTIVE_WORKBOOK_BYTES_SESSION_KEY] = workbook_bytes


def clear_uploaded_workbook_session() -> None:
    for key in [
        ACTIVE_WORKBOOK_NAME_SESSION_KEY,
        ACTIVE_WORKBOOK_BYTES_SESSION_KEY,
    ]:
        st.session_state.pop(key, None)


def active_uploaded_workbook_from_session() -> tuple[str, bytes] | None:
    workbook_name = normalize_answer(
        st.session_state.get(ACTIVE_WORKBOOK_NAME_SESSION_KEY)
    )
    workbook_bytes = st.session_state.get(ACTIVE_WORKBOOK_BYTES_SESSION_KEY)
    if not workbook_name or not workbook_bytes:
        return None
    return workbook_name, workbook_bytes


def read_excel_sheet(workbook_bytes: bytes, sheet_name: str) -> pd.DataFrame | None:
    try:
        return pd.read_excel(BytesIO(workbook_bytes), sheet_name=sheet_name, dtype=object)
    except Exception as exc:
        st.error(f"Could not read Excel sheet `{sheet_name}`: {exc}")
        return None


def auto_detect_response_labels(
    workbook_bytes: bytes,
    excel_file: pd.ExcelFile,
    data_sheet: str,
) -> tuple[str, dict[str, dict[str, str]], dict[str, str]]:
    for sheet_name in excel_file.sheet_names:
        if sheet_name == data_sheet:
            continue

        try:
            label_data = pd.read_excel(
                BytesIO(workbook_bytes),
                sheet_name=sheet_name,
                dtype=object,
            )
        except Exception:
            continue

        if label_data.empty:
            continue

        label_columns = list(label_data.columns)
        question_column = infer_column(label_columns, QUESTION_COLUMN_CANDIDATES)
        value_column = infer_column(label_columns, VALUE_COLUMN_CANDIDATES)
        label_column = infer_column(label_columns, LABEL_COLUMN_CANDIDATES)
        question_label_column = infer_column(label_columns, QUESTION_LABEL_COLUMN_CANDIDATES)

        if not question_column or not value_column or not label_column:
            continue

        label_response_labels, label_question_labels = build_response_label_maps(
            label_data,
            question_column,
            value_column,
            label_column,
            question_label_column,
        )
        if label_response_labels or label_question_labels:
            return sheet_name, label_response_labels, label_question_labels

    return NO_LABEL_SHEET, {}, {}


def project_metadata_column_lookup(columns) -> dict[str, str]:
    normalized_columns = {
        normalize_alias_key(column): column
        for column in columns
        if normalize_alias_key(column)
    }
    return {
        variable: normalized_columns[normalize_alias_key(variable)]
        for variable in PROJECT_METADATA_VARIABLES
        if normalize_alias_key(variable) in normalized_columns
    }


def summarize_metadata_value(series: pd.Series) -> str:
    values = [
        value
        for value in series.map(normalize_answer).dropna().drop_duplicates().tolist()
        if value
    ]
    if not values:
        return NOT_AVAILABLE
    if len(values) <= 3:
        return " | ".join(values)
    return " | ".join(values[:3]) + f" | +{len(values) - 3} more"


def project_metadata_values(df: pd.DataFrame | None) -> dict[str, str]:
    values: dict[str, str] = {}
    if df is None:
        return {
            variable: NOT_AVAILABLE
            for variable in PROJECT_METADATA_VARIABLES
        }

    lookup = project_metadata_column_lookup(df.columns)
    for variable in PROJECT_METADATA_VARIABLES:
        source_column = lookup.get(variable)
        values[variable] = (
            summarize_metadata_value(df[source_column])
            if source_column in df.columns
            else NOT_AVAILABLE
        )
    return values


def norm_filter_key(label: str) -> str:
    return f"{NORM_FILTER_SESSION_PREFIX}{normalize_alias_key(label)}"


def reset_norm_filters() -> None:
    for label, _aliases in NORM_FILTER_FIELDS:
        st.session_state[norm_filter_key(label)] = []


def filter_key_matches_alias(normalized_key: str, aliases: list[str]) -> bool:
    if not normalized_key:
        return False

    key_parts = set(normalized_key.split("_"))
    for alias in aliases:
        target = normalize_alias_key(alias)
        if normalized_key == target or target in key_parts:
            return True
    return False


def filter_column_for_field(
    df: pd.DataFrame,
    question_labels: dict[str, str],
    aliases: list[str],
    exact_match: bool = False,
) -> str | None:
    alias_keys = {normalize_alias_key(alias) for alias in aliases if normalize_alias_key(alias)}

    for column in df.columns:
        if normalize_alias_key(column) in alias_keys:
            return column

    for column in df.columns:
        label = question_labels.get(column)
        if label and normalize_alias_key(label) in alias_keys:
            return column

    if exact_match:
        return None

    for column in df.columns:
        if filter_key_matches_alias(normalize_alias_key(column), aliases):
            return column

    for column in df.columns:
        label = question_labels.get(column)
        if label and filter_key_matches_alias(normalize_alias_key(label), aliases):
            return column

    return None


def filter_option_values(series: pd.Series) -> list[str]:
    values = [
        value
        for value in series.map(normalize_answer).dropna().drop_duplicates().tolist()
        if value
    ]

    def sort_key(value: str) -> tuple[int, int | str, str]:
        if value.isdigit():
            return (0, int(value), value)
        return (1, value.lower(), value)

    return sorted(values, key=sort_key)


def apply_norm_filters(
    df: pd.DataFrame,
    selected_filters: dict[str, list[str]],
) -> pd.DataFrame:
    filtered_df = df
    for column, selected_values in selected_filters.items():
        if column not in filtered_df.columns or not selected_values:
            continue

        selected_value_set = {
            normalize_answer(value)
            for value in selected_values
            if normalize_answer(value)
        }
        if not selected_value_set:
            continue

        filtered_df = filtered_df[
            filtered_df[column].map(normalize_answer).isin(selected_value_set)
        ]

    return filtered_df


def describe_norm_filter(
    label: str,
    column: str,
    selected_values: list[str],
    response_labels: dict[str, dict[str, str]],
) -> str:
    displayed_values = [
        display_response_label(column, value, response_labels)
        for value in selected_values
    ]
    if len(displayed_values) > 3:
        displayed_values = displayed_values[:3] + [f"+{len(selected_values) - 3} more"]
    return f"{label}: {', '.join(displayed_values)}"


def render_norm_filter_controls(
    df: pd.DataFrame,
    group_column: str,
    control_label: str,
    test_label: str,
    question_labels: dict[str, str],
    response_labels: dict[str, dict[str, str]],
) -> pd.DataFrame:
    st.subheader("Filters")
    st.caption(
        "Filter the respondent rows before calculating the norm tables. "
        "Reset returns to total control vs test."
    )

    reset_col, status_col = st.columns([1, 3])
    if reset_col.button("Reset to total control vs test", use_container_width=True):
        reset_norm_filters()

    selected_filters: dict[str, list[str]] = {}
    active_filter_descriptions: list[str] = []
    unavailable_filters: list[str] = []
    filter_columns = st.columns(3)

    for index, (label, aliases) in enumerate(NORM_FILTER_FIELDS):
        key = norm_filter_key(label)
        column = filter_column_for_field(
            df,
            question_labels,
            aliases,
            exact_match=label in EXACT_NORM_FILTERS,
        )
        column_container = filter_columns[index % len(filter_columns)]

        if column is None:
            st.session_state[key] = []
            unavailable_filters.append(label)
            column_container.button(
                f"{label}: Not available",
                disabled=True,
                use_container_width=True,
            )
            continue

        options = filter_option_values(df[column])
        st.session_state[key] = [
            value
            for value in st.session_state.get(key, [])
            if value in options
        ]

        with column_container.popover(label, use_container_width=True):
            selected_values = st.multiselect(
                f"{label} values",
                options,
                key=key,
                format_func=lambda value, source_column=column: display_response_label(
                    source_column,
                    value,
                    response_labels,
                ),
                placeholder="All",
            )
            st.caption(f"Source variable: {column}")

        selected_values = st.session_state.get(key, [])
        if selected_values:
            selected_filters[column] = selected_values
            active_filter_descriptions.append(
                describe_norm_filter(label, column, selected_values, response_labels)
            )

    filtered_df = apply_norm_filters(df, selected_filters)
    group_values = (
        filtered_df[group_column].map(normalize_answer)
        if group_column in filtered_df.columns
        else pd.Series([], dtype=object)
    )
    control_rows = int((group_values == control_label).sum())
    test_rows = int((group_values == test_label).sum())

    if active_filter_descriptions:
        status_col.caption("Active filters: " + " | ".join(active_filter_descriptions))
    else:
        status_col.caption("No filters applied. Tables show total control vs test.")

    metric_cols = st.columns(3)
    metric_cols[0].metric("Filtered rows", f"{len(filtered_df):,}")
    metric_cols[1].metric("Control rows", f"{control_rows:,}")
    metric_cols[2].metric("Test rows", f"{test_rows:,}")

    if unavailable_filters:
        st.caption("Unavailable filters: " + ", ".join(unavailable_filters))
    if filtered_df.empty:
        st.warning("No respondent rows match the selected filters.")

    return filtered_df


def render_saved_norm_filter_controls(saved_records: list[dict]) -> dict[str, list[str]]:
    st.subheader("Filters")
    st.caption(
        "Filter respondent rows across all saved datasets before reviewing norm tables. "
        "Reset returns to total control vs test."
    )

    reset_col, status_col = st.columns([1, 3])
    if reset_col.button("Reset to total control vs test", use_container_width=True):
        reset_norm_filters()

    options_by_label, columns_by_label = saved_norm_filter_options(saved_records)
    selected_filters: dict[str, list[str]] = {}
    active_filter_descriptions: list[str] = []
    unavailable_filters: list[str] = []
    filter_columns = st.columns(3)

    for index, (label, _aliases) in enumerate(NORM_FILTER_FIELDS):
        key = norm_filter_key(label)
        column_container = filter_columns[index % len(filter_columns)]
        options = options_by_label.get(label, [])
        source_columns = columns_by_label.get(label, [])

        if not options:
            st.session_state[key] = []
            unavailable_filters.append(label)
            column_container.button(
                f"{label}: Not available",
                disabled=True,
                use_container_width=True,
            )
            continue

        st.session_state[key] = [
            value
            for value in st.session_state.get(key, [])
            if value in options
        ]

        with column_container.popover(label, use_container_width=True):
            st.multiselect(
                f"{label} values",
                options,
                key=key,
                placeholder="All",
            )
            source_caption = ", ".join(source_columns[:3])
            if len(source_columns) > 3:
                source_caption += f", +{len(source_columns) - 3} more"
            st.caption(f"Source variable: {source_caption}")

        selected_values = st.session_state.get(key, [])
        if selected_values:
            selected_filters[label] = selected_values
            active_filter_descriptions.append(
                f"{label}: {', '.join(selected_values[:3])}"
                + (f", +{len(selected_values) - 3} more" if len(selected_values) > 3 else "")
            )

    if active_filter_descriptions:
        status_col.caption("Active filters: " + " | ".join(active_filter_descriptions))
    else:
        status_col.caption("No filters applied. Tables show total control vs test.")

    if unavailable_filters:
        st.caption("Unavailable filters: " + ", ".join(unavailable_filters))

    return selected_filters


def prompt_for_missing_project_metadata(
    df: pd.DataFrame,
    question_labels: dict[str, str],
) -> tuple[pd.DataFrame, dict[str, str]]:
    lookup = project_metadata_column_lookup(df.columns)
    missing_variables = [
        variable
        for variable in PROJECT_METADATA_VARIABLES
        if variable not in lookup
    ]
    if not missing_variables:
        return df, question_labels

    st.subheader("Project metadata")
    st.caption(
        "These metadata fields are used as norm table filters. Add a value for "
        "any missing field and the app will create that column for this workbook."
    )

    updated_df = df.copy()
    updated_labels = dict(question_labels)
    input_columns = st.columns(3)
    added_variables = []
    for index, variable in enumerate(missing_variables):
        value = input_columns[index % 3].text_input(
            variable,
            key=f"project_metadata_value_{variable}",
            placeholder=f"Enter {variable}",
        )
        normalized_value = normalize_answer(value)
        if normalized_value:
            updated_df[variable] = normalized_value
            updated_labels[variable] = variable
            added_variables.append(variable)

    if added_variables:
        st.caption(
            "Added metadata fields: "
            + ", ".join(added_variables)
            + "."
        )
    else:
        st.info("Missing metadata fields left blank will not be available as filters.")

    return updated_df, updated_labels


def read_respondent_sheet(
    workbook_bytes: bytes,
    sheet_name: str,
    data_layout: str,
) -> RespondentSheet | None:
    try:
        if data_layout == SMART_TABLES_LAYOUT:
            raw_df = pd.read_excel(
                BytesIO(workbook_bytes),
                sheet_name=sheet_name,
                header=None,
                dtype=object,
            )
            return prepare_smart_tables_layout(raw_df)

        df = pd.read_excel(BytesIO(workbook_bytes), sheet_name=sheet_name, dtype=object)
        question_labels = {column: str(column) for column in df.columns}
        metadata_columns = identify_metadata_columns(df.columns)
        return RespondentSheet(
            dataframe=df,
            question_labels=question_labels,
            metadata_columns_default_na=len(metadata_columns),
            metadata_columns=metadata_columns,
        )
    except Exception as exc:
        st.error(f"Could not read respondent data from `{sheet_name}`: {exc}")
        return None


def excel_safe_sheet_name(name: str, used_names: set[str]) -> str:
    safe_name = re.sub(r"[:\\/?*\[\]]", " ", name).strip() or "Norm Table"
    safe_name = re.sub(r"\s+", " ", safe_name)[:31]
    candidate = safe_name
    counter = 2

    while candidate in used_names:
        suffix = f" {counter}"
        candidate = f"{safe_name[:31 - len(suffix)]}{suffix}"
        counter += 1

    used_names.add(candidate)
    return candidate


def excel_cell_value(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def excel_thin_border() -> Border:
    side = Side(style="thin", color=VN_GRAY_200)
    return Border(left=side, right=side, top=side, bottom=side)


def set_excel_column_widths(worksheet, start_col: int, columns: list[str]) -> None:
    width_by_column = {
        "Response option": 34,
        "Control": 13,
        "Test": 13,
        "Lift": 13,
        "Significance result": 22,
    }
    for offset, column in enumerate(columns):
        width = width_by_column.get(str(column), 16)
        worksheet.column_dimensions[get_column_letter(start_col + offset)].width = width


def write_excel_sheet_header(
    worksheet,
    title: str,
    subtitle: str = "",
    end_col: int = 5,
) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    title_cell = worksheet.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=18, color=VN_BLACK)
    title_cell.alignment = Alignment(vertical="center")
    worksheet.row_dimensions[1].height = 28

    if subtitle:
        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
        subtitle_cell = worksheet.cell(row=2, column=1, value=subtitle)
        subtitle_cell.font = Font(size=10, color=VN_GRAY_400)
        subtitle_cell.alignment = Alignment(wrap_text=True, vertical="top")
        worksheet.row_dimensions[2].height = 28

    for col in range(1, end_col + 1):
        accent_cell = worksheet.cell(row=3, column=col)
        accent_cell.fill = PatternFill("solid", fgColor=VN_PINK if col <= 2 else VN_BLACK)
    worksheet.row_dimensions[3].height = 4


def write_section_title(
    worksheet,
    row: int,
    title: str,
    end_col: int = 5,
) -> None:
    worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cell = worksheet.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=14, color=VN_BLACK)
    cell.alignment = Alignment(vertical="center")
    cell.border = Border(bottom=Side(style="medium", color=VN_PINK))
    worksheet.row_dimensions[row].height = 24


def write_norm_table_to_excel(
    worksheet,
    table: pd.DataFrame,
    start_row: int,
    start_col: int = 1,
) -> int:
    columns = [str(column) for column in table.columns]
    border = excel_thin_border()
    header_fill = PatternFill("solid", fgColor=VN_GRAY_50)
    alt_fill = PatternFill("solid", fgColor=VN_GRAY_50)

    for col_offset, column in enumerate(columns):
        cell = worksheet.cell(row=start_row, column=start_col + col_offset, value=column)
        cell.fill = header_fill
        cell.font = Font(bold=True, color=VN_BLACK)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = border

    for row_offset, row in enumerate(table.itertuples(index=False, name=None), start=1):
        excel_row = start_row + row_offset
        fill = alt_fill if row_offset % 2 == 0 else PatternFill(fill_type=None)
        for col_offset, value in enumerate(row):
            column_name = columns[col_offset]
            cell = worksheet.cell(
                row=excel_row,
                column=start_col + col_offset,
                value=excel_cell_value(value),
            )
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if column_name in {"Control", "Test", "Lift"} else "left",
                vertical="center",
                wrap_text=True,
            )
        worksheet.row_dimensions[excel_row].height = 26

    worksheet.row_dimensions[start_row].height = 28
    set_excel_column_widths(worksheet, start_col, columns)
    return start_row + len(table)


def excel_lift_font_color(chart_row: dict) -> str:
    if not chart_row.get("significant"):
        return "4C5361"
    if chart_row.get("lift_points", 0) > 0:
        return VN_GREEN
    if chart_row.get("lift_points", 0) < 0:
        return VN_SIG_RED
    return "4C5361"


def excel_chart_category_label(label: object, max_line_length: int = 18) -> str:
    text = str(label or "").strip()
    if len(text) <= max_line_length:
        return text

    lines: list[str] = []
    current_line = ""
    for word in text.split():
        candidate = f"{current_line} {word}".strip()
        if current_line and len(candidate) > max_line_length:
            lines.append(current_line)
            current_line = word
        else:
            current_line = candidate

    if current_line:
        lines.append(current_line)

    return "\n".join(lines)


def write_chart_source_table(
    worksheet,
    chart_rows: list[dict],
    start_row: int,
    start_col: int,
) -> None:
    headers = ["Response option", "Control", "Test", "", "Lift", "Significance"]
    worksheet.cell(row=start_row - 2, column=start_col, value="Editable chart source").font = Font(
        bold=True,
        size=10,
        color=VN_BLACK,
    )
    worksheet.cell(
        row=start_row - 1,
        column=start_col,
        value="Chart links here so Excel and Google Sheets can edit it.",
    ).font = Font(italic=True, size=9, color=VN_GRAY_400)

    border = excel_thin_border()
    header_fill = PatternFill("solid", fgColor=VN_GRAY_50)
    for offset, header in enumerate(headers):
        cell = worksheet.cell(row=start_row, column=start_col + offset, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True, color=VN_BLACK)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row_offset, chart_row in enumerate(chart_rows, start=1):
        source_row = start_row + row_offset
        lift_points = round_percentage_points(chart_row["lift_points"])
        chart_label = excel_chart_category_label(chart_row["label"])
        values = [
            chart_label,
            chart_row["control_points"] / 100,
            chart_row["test_points"] / 100,
            0,
            f"{lift_points:+d}",
            "Significant" if chart_row["significant"] else "Not significant",
        ]
        for col_offset, value in enumerate(values):
            cell = worksheet.cell(
                source_row,
                start_col + col_offset,
                excel_cell_value(value),
            )
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if col_offset else "left",
                vertical="center",
                wrap_text=True,
            )
            if col_offset in {1, 2}:
                cell.number_format = "0%"
            if col_offset == 4:
                cell.fill = PatternFill("solid", fgColor=VN_WHITE)
                cell.font = Font(bold=True, color=excel_lift_font_color(chart_row))
        worksheet.row_dimensions[source_row].height = max(24, 15 * (chart_label.count("\n") + 1))

    source_widths = [28, 11, 11, 8, 9, 16]
    for offset, width in enumerate(source_widths):
        worksheet.column_dimensions[get_column_letter(start_col + offset)].width = width
    worksheet.column_dimensions[get_column_letter(start_col + 3)].hidden = True


def excel_chart_text(size: int, bold: bool = False, color: str = VN_BLACK) -> RichText:
    def character_properties() -> CharacterProperties:
        return CharacterProperties(
            sz=size,
            b=bold,
            latin=DrawingFont(typeface="Arial"),
            solidFill=color,
        )

    return RichText(
        bodyPr=RichTextProperties(noAutofit=True),
        p=[
            Paragraph(
                pPr=ParagraphProperties(
                    defRPr=character_properties(),
                ),
                r=[RegularTextRun(rPr=character_properties(), t="")],
                endParaRPr=character_properties(),
            )
        ]
    )


def excel_chart_text_xml(size: int, bold: bool = False, color: str = VN_BLACK) -> str:
    bold_value = "1" if bold else "0"
    return (
        "<txPr>"
        '<a:bodyPr xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        "<a:noAutofit/>"
        "</a:bodyPr>"
        '<a:p xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        "<a:pPr>"
        f'<a:defRPr sz="{size}" b="{bold_value}">'
        "<a:solidFill>"
        f'<a:srgbClr val="{color}"/>'
        "</a:solidFill>"
        '<a:latin typeface="Arial"/>'
        "</a:defRPr>"
        "</a:pPr>"
        f'<a:r><a:rPr sz="{size}" b="{bold_value}">'
        "<a:solidFill>"
        f'<a:srgbClr val="{color}"/>'
        "</a:solidFill>"
        '<a:latin typeface="Arial"/>'
        "</a:rPr><a:t></a:t></a:r>"
        f'<a:endParaRPr sz="{size}" b="{bold_value}">'
        "<a:solidFill>"
        f'<a:srgbClr val="{color}"/>'
        "</a:solidFill>"
        '<a:latin typeface="Arial"/>'
        "</a:endParaRPr>"
        "</a:p>"
        "</txPr>"
    )


def patch_excel_chart_text_defaults(workbook_bytes: bytes) -> bytes:
    source = BytesIO(workbook_bytes)
    patched = BytesIO()
    chart_default_text = excel_chart_text_xml(850)

    with zipfile.ZipFile(source, "r") as source_zip:
        with zipfile.ZipFile(patched, "w", zipfile.ZIP_DEFLATED) as patched_zip:
            for item in source_zip.infolist():
                data = source_zip.read(item.filename)
                if item.filename.startswith("xl/charts/chart") and item.filename.endswith(".xml"):
                    xml = data.decode("utf-8")
                    if "</chartSpace>" in xml and "</chart><txPr>" not in xml:
                        xml = xml.replace(
                            "</chartSpace>",
                            f"{chart_default_text}</chartSpace>",
                            1,
                        )
                    data = xml.encode("utf-8")
                patched_zip.writestr(item, data)

    return patched.getvalue()


def add_native_norm_excel_chart(
    worksheet,
    chart_rows: list[dict],
    source_row: int,
    source_col: int,
    anchor_row: int,
) -> None:
    if not chart_rows:
        return

    chart_row_count = len(chart_rows)
    chart = BarChart()
    chart.type = "col"
    chart.title = None
    chart.y_axis.numFmt = "0%"
    chart.y_axis.scaling.min = 0
    chart.y_axis.delete = True
    chart.y_axis.majorGridlines = None
    chart.x_axis.majorGridlines = None
    chart.x_axis.delete = True
    chart.x_axis.majorTickMark = "none"
    chart.x_axis.minorTickMark = "none"
    chart.x_axis.txPr = excel_chart_text(850)
    chart.legend.position = "l"
    chart.legend.overlay = False
    chart.legend.legendEntry = [LegendEntry(idx=2, delete=True)]
    chart.legend.txPr = excel_chart_text(1100, bold=True)
    chart.height = 9.4
    chart.width = max(22.0, min(32.0, 11.0 + chart_row_count * 2.25))
    chart.gapWidth = 60
    chart.overlap = 0
    chart.visible_cells_only = False

    max_points = max(
        max(row["control_points"], row["test_points"])
        for row in chart_rows
    )
    chart.y_axis.scaling.max = min(
        1,
        max(0.12, math.ceil(((max_points + 10) / 100) * 10) / 10),
    )

    data = Reference(
        worksheet,
        min_col=source_col + 1,
        max_col=source_col + 3,
        min_row=source_row,
        max_row=source_row + chart_row_count,
    )
    categories = Reference(
        worksheet,
        min_col=source_col,
        min_row=source_row + 1,
        max_row=source_row + chart_row_count,
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    if len(chart.series) >= 2:
        value_label_style = DataLabelList(
            showVal=True,
            numFmt="0%",
            dLblPos="outEnd",
            txPr=excel_chart_text(1100, bold=True),
        )
        chart.series[0].dLbls = value_label_style
        chart.series[1].dLbls = value_label_style
        chart.series[0].graphicalProperties.solidFill = VN_CONTROL_GRAY
        chart.series[0].graphicalProperties.line.solidFill = VN_CONTROL_GRAY
        chart.series[1].graphicalProperties.solidFill = VN_PINK
        chart.series[1].graphicalProperties.line.solidFill = VN_PINK

    if len(chart.series) >= 3:
        chart.series[2].dLbls = DataLabelList(
            showCatName=True,
            showVal=False,
            dLblPos="b",
            txPr=excel_chart_text(500),
        )
        chart.series[2].graphicalProperties.noFill = True
        chart.series[2].graphicalProperties.line.noFill = True

    worksheet.add_chart(chart, f"A{anchor_row}")


def add_norm_chart_to_excel_sheet(
    worksheet,
    table: pd.DataFrame,
    table_end_row: int,
) -> None:
    chart_rows = norm_chart_rows(table)
    if not chart_rows:
        return

    chart_col = len(table.columns) + 2
    chart_data_row = 6
    write_chart_source_table(worksheet, chart_rows, chart_data_row, chart_col)
    add_native_norm_excel_chart(
        worksheet,
        chart_rows,
        chart_data_row,
        chart_col,
        table_end_row + 3,
    )


def write_all_norms_sheet(
    workbook: Workbook,
    output_tables: dict[str, pd.DataFrame],
) -> None:
    worksheet = workbook.active
    worksheet.title = "All Norms"
    write_excel_sheet_header(
        worksheet,
        "All Norm Tables",
        "Each section mirrors the app view with space between tables for review.",
        end_col=5,
    )
    current_row = 5
    for question, table in output_tables.items():
        write_section_title(worksheet, current_row, str(question), end_col=len(table.columns))
        table_start_row = current_row + 2
        table_end_row = write_norm_table_to_excel(worksheet, table, table_start_row)
        current_row = table_end_row + 4
    worksheet.freeze_panes = "A5"


def write_question_sheet(
    workbook: Workbook,
    sheet_name: str,
    question: str,
    table: pd.DataFrame,
) -> None:
    worksheet = workbook.create_sheet(sheet_name)
    write_excel_sheet_header(
        worksheet,
        str(question),
        "Norm table and editable control-vs-test chart.",
        end_col=11,
    )
    table_start_row = 5
    table_end_row = write_norm_table_to_excel(worksheet, table, table_start_row)
    add_norm_chart_to_excel_sheet(worksheet, table, table_end_row)
    worksheet.freeze_panes = "A5"


def norm_tables_to_excel(tables: dict[str, pd.DataFrame]) -> bytes:
    output = BytesIO()
    used_names = {"All Norms"}
    output_tables = {
        question: normalize_lift_output_table(table)
        for question, table in tables.items()
    }

    workbook = Workbook()
    write_all_norms_sheet(workbook, output_tables)

    for question, table in output_tables.items():
        sheet_name = excel_safe_sheet_name(question, used_names)
        write_question_sheet(workbook, sheet_name, str(question), table)

    workbook.save(output)
    return patch_excel_chart_text_defaults(output.getvalue())


def ensure_uploaded_dataset_dirs() -> None:
    for directory in [
        UPLOADED_RAW_DIR,
        UPLOADED_NORM_WORKBOOKS_DIR,
        UPLOADED_SETTINGS_DIR,
        UPLOADED_HISTORY_DIR,
    ]:
        directory.mkdir(parents=True, exist_ok=True)


def safe_dataset_filename(value: object, fallback: str = "upload") -> str:
    filename = normalize_answer(value) or fallback
    filename = Path(filename).name
    filename = re.sub(r"[^A-Za-z0-9._-]+", "_", filename).strip("._")
    return filename or fallback


def uploaded_dataset_rules_path(dataset_id: str) -> Path:
    return UPLOADED_SETTINGS_DIR / f"{safe_dataset_filename(dataset_id, 'dataset')}_rules.json"


def uploaded_dataset_raw_path(record: dict) -> Path:
    dataset_id = safe_dataset_filename(record.get("dataset_id"), "dataset")
    uploaded_file = safe_dataset_filename(record.get("uploaded_file"), "raw_upload.xlsx")
    suffix = Path(uploaded_file).suffix or ".xlsx"
    stem = Path(uploaded_file).stem or "raw_upload"
    return UPLOADED_RAW_DIR / f"{dataset_id}_{stem}{suffix}"


def uploaded_dataset_workbook_path(dataset_id: str) -> Path:
    return UPLOADED_NORM_WORKBOOKS_DIR / f"{safe_dataset_filename(dataset_id, 'dataset')}.xlsx"


def read_manifest_records(path: Path) -> list[dict]:
    if not path.exists():
        return []
    try:
        data = json.loads(path.read_text())
    except (json.JSONDecodeError, OSError):
        return []
    return data if isinstance(data, list) else []


def norm_history_version_dir(version_id: str) -> Path:
    return UPLOADED_HISTORY_DIR / safe_dataset_filename(version_id, "version")


def load_norm_history_index() -> list[dict]:
    if not UPLOADED_HISTORY_INDEX_PATH.exists():
        return []
    try:
        data = json.loads(UPLOADED_HISTORY_INDEX_PATH.read_text())
    except (json.JSONDecodeError, OSError):
        return []
    entries = data if isinstance(data, list) else []
    return sorted(
        [
            entry
            for entry in entries
            if normalize_answer(entry.get("version_id"))
        ],
        key=lambda entry: normalize_answer(entry.get("created_at")) or "",
        reverse=True,
    )


def save_norm_history_index(entries: list[dict]) -> None:
    ensure_uploaded_dataset_dirs()
    UPLOADED_HISTORY_INDEX_PATH.write_text(
        json.dumps(entries, indent=2, sort_keys=True) + "\n"
    )


def copy_history_files(source_dir: Path, destination_dir: Path, pattern: str = "*") -> None:
    if not source_dir.exists():
        return
    destination_dir.mkdir(parents=True, exist_ok=True)
    for source_path in sorted(source_dir.glob(pattern)):
        if source_path.is_file() and source_path.name != ".DS_Store":
            shutil.copy2(source_path, destination_dir / source_path.name)


def create_norm_database_history_snapshot(
    action: str,
    records: list[dict],
    record: dict | None = None,
    note: str = "",
) -> dict | None:
    ensure_uploaded_dataset_dirs()
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    action_slug = safe_dataset_filename(action, "snapshot")[:32]
    version_base = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{action_slug}"
    version_id = version_base
    counter = 2
    while norm_history_version_dir(version_id).exists():
        version_id = f"{version_base}_{counter}"
        counter += 1

    snapshot_dir = norm_history_version_dir(version_id)
    settings_dir = snapshot_dir / "norm_settings"
    workbooks_dir = snapshot_dir / "norm_workbooks"
    raw_dir = snapshot_dir / "raw_uploads"
    settings_dir.mkdir(parents=True, exist_ok=True)
    workbooks_dir.mkdir(parents=True, exist_ok=True)
    raw_dir.mkdir(parents=True, exist_ok=True)

    try:
        if UPLOADED_MANIFEST_BACKUP_PATH.exists():
            shutil.copy2(UPLOADED_MANIFEST_BACKUP_PATH, settings_dir / "manifest.json")
        else:
            (settings_dir / "manifest.json").write_text(
                json.dumps(records, indent=2, sort_keys=True) + "\n"
            )
        if UPLOADED_NORM_TABLES_BACKUP_PATH.exists():
            shutil.copy2(
                UPLOADED_NORM_TABLES_BACKUP_PATH,
                settings_dir / "saved_norm_tables.xlsx",
            )

        copy_history_files(UPLOADED_NORM_WORKBOOKS_DIR, workbooks_dir, "*.xlsx")
        copy_history_files(UPLOADED_SETTINGS_DIR, settings_dir, "*.json")
        copy_history_files(UPLOADED_RAW_DIR, raw_dir)
    except OSError:
        return None

    dataset_id = normalize_answer(record.get("dataset_id")) if record else ""
    metadata = {
        "version_id": version_id,
        "created_at": created_at,
        "action": action,
        "note": note,
        "dataset_id": dataset_id or NOT_AVAILABLE,
        "dataset_label": saved_dataset_option_label(record) if record else NOT_AVAILABLE,
        "dataset_count": len(records),
    }
    try:
        (snapshot_dir / "metadata.json").write_text(
            json.dumps(metadata, indent=2, sort_keys=True) + "\n"
        )
        save_norm_history_index([metadata, *load_norm_history_index()])
    except OSError:
        return None

    return metadata


def restore_norm_database_history_version(version_id: str) -> tuple[bool, str]:
    safe_version_id = safe_dataset_filename(version_id, "")
    if not safe_version_id:
        return False, "Select a version to restore."

    snapshot_dir = norm_history_version_dir(safe_version_id)
    snapshot_manifest_path = snapshot_dir / "norm_settings" / "manifest.json"
    if not snapshot_manifest_path.exists():
        return False, "The selected version is missing its manifest."

    restored_records = read_manifest_records(snapshot_manifest_path)
    restored_record_ids = {
        normalize_answer(record.get("dataset_id"))
        for record in restored_records
        if normalize_answer(record.get("dataset_id"))
    }

    try:
        with norm_database_write_lock():
            current_records = load_norm_database_manifest()
            backup_norm_database_to_uploaded_datasets(current_records)
            create_norm_database_history_snapshot(
                "Before restore",
                current_records,
                note=f"Automatic safety snapshot before restoring {safe_version_id}.",
            )

            ensure_norm_database_dirs()
            for dataset_path in NORM_DATABASE_DATASETS_DIR.glob("*.xlsx"):
                if dataset_path.stem not in restored_record_ids:
                    dataset_path.unlink(missing_ok=True)

            missing_workbooks = []
            for record in restored_records:
                dataset_id = normalize_answer(record.get("dataset_id"))
                if not dataset_id:
                    continue
                source_workbook_path = (
                    snapshot_dir
                    / "norm_workbooks"
                    / f"{safe_dataset_filename(dataset_id, 'dataset')}.xlsx"
                )
                if not source_workbook_path.exists():
                    missing_workbooks.append(dataset_id)
                    continue
                shutil.copy2(source_workbook_path, norm_database_dataset_path(dataset_id))

            if missing_workbooks:
                return (
                    False,
                    "The selected version is missing workbook files for: "
                    + ", ".join(missing_workbooks),
                )

            save_norm_database_manifest(restored_records)
            refresh_norm_database_workbook(restored_records)
            backup_norm_database_to_uploaded_datasets(restored_records)
            create_norm_database_history_snapshot(
                "Restore version",
                restored_records,
                note=f"Restored version {safe_version_id}.",
            )
    except (OSError, TimeoutError) as exc:
        return False, str(exc)

    commit_ok, commit_message = github_autocommit_uploaded_datasets(
        f"Restore BLS norms database version {safe_version_id}"
    )
    return True, save_message_with_github_status(
        "Saved datasets restored from version history.",
        commit_ok,
        commit_message,
    )


def restore_norm_database_from_uploaded_datasets_if_needed() -> None:
    if not UPLOADED_MANIFEST_BACKUP_PATH.exists():
        return

    try:
        NORM_DATABASE_DATASETS_DIR.mkdir(parents=True, exist_ok=True)
        backup_records = read_manifest_records(UPLOADED_MANIFEST_BACKUP_PATH)
        if not backup_records:
            return

        if NORM_DATABASE_MANIFEST_PATH.exists():
            records_to_restore = read_manifest_records(NORM_DATABASE_MANIFEST_PATH)
        else:
            shutil.copy2(UPLOADED_MANIFEST_BACKUP_PATH, NORM_DATABASE_MANIFEST_PATH)
            records_to_restore = backup_records

        if (
            UPLOADED_NORM_TABLES_BACKUP_PATH.exists()
            and not NORM_DATABASE_WORKBOOK_PATH.exists()
        ):
            shutil.copy2(UPLOADED_NORM_TABLES_BACKUP_PATH, NORM_DATABASE_WORKBOOK_PATH)

        for record in records_to_restore:
            dataset_id = normalize_answer(record.get("dataset_id"))
            if not dataset_id:
                continue
            source_path = uploaded_dataset_workbook_path(dataset_id)
            destination_path = norm_database_dataset_path(dataset_id)
            if source_path.exists() and not destination_path.exists():
                shutil.copy2(source_path, destination_path)
    except OSError:
        return


def backup_norm_database_to_uploaded_datasets(
    records: list[dict],
    record: dict | None = None,
    raw_workbook_bytes: bytes | None = None,
    rules: dict | None = None,
) -> None:
    ensure_uploaded_dataset_dirs()

    try:
        UPLOADED_MANIFEST_BACKUP_PATH.write_text(
            json.dumps(records, indent=2, sort_keys=True) + "\n"
        )
        if NORM_DATABASE_WORKBOOK_PATH.exists():
            shutil.copy2(NORM_DATABASE_WORKBOOK_PATH, UPLOADED_NORM_TABLES_BACKUP_PATH)

        for saved_record in records:
            dataset_id = normalize_answer(saved_record.get("dataset_id"))
            if not dataset_id:
                continue
            source_dataset_path = norm_database_dataset_path(dataset_id)
            if source_dataset_path.exists():
                shutil.copy2(
                    source_dataset_path,
                    uploaded_dataset_workbook_path(dataset_id),
                )

        if not record:
            return

        dataset_id = normalize_answer(record.get("dataset_id"))
        if not dataset_id:
            return

        if raw_workbook_bytes:
            uploaded_dataset_raw_path(record).write_bytes(raw_workbook_bytes)

        if rules is not None:
            uploaded_dataset_rules_path(dataset_id).write_text(
                json.dumps(rules, indent=2, sort_keys=True) + "\n"
            )
    except (OSError, TypeError):
        return


def streamlit_secret(section: str, key: str, default=None):
    try:
        section_values = st.secrets.get(section, {})
        if hasattr(section_values, "get"):
            return section_values.get(key, default)
    except Exception:
        return default
    return default


def truthy_config(value, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on"}


def github_autocommit_config() -> dict:
    token = (
        os.environ.get("GITHUB_AUTOCOMMIT_TOKEN")
        or os.environ.get("GITHUB_TOKEN")
        or streamlit_secret("github_autocommit", "token")
        or streamlit_secret("github", "token")
    )
    repo = (
        os.environ.get("GITHUB_AUTOCOMMIT_REPO")
        or os.environ.get("GITHUB_REPOSITORY")
        or streamlit_secret("github_autocommit", "repo")
        or streamlit_secret("github", "repo")
    )
    branch = (
        os.environ.get("GITHUB_AUTOCOMMIT_BRANCH")
        or os.environ.get("GITHUB_BRANCH")
        or streamlit_secret("github_autocommit", "branch")
        or streamlit_secret("github", "branch")
        or "main"
    )
    data_path = (
        os.environ.get("GITHUB_AUTOCOMMIT_DATA_PATH")
        or streamlit_secret("github_autocommit", "data_path")
        or streamlit_secret("github", "data_path")
        or UPLOADED_DATASETS_DIR.name
    )
    enabled_value = (
        os.environ.get("GITHUB_AUTOCOMMIT_ENABLED")
        or streamlit_secret("github_autocommit", "enabled")
        or streamlit_secret("github", "autocommit")
    )
    enabled = truthy_config(enabled_value, default=bool(token and repo))
    if repo:
        repo = str(repo).strip()
        repo = re.sub(r"^https://github\.com/", "", repo).strip("/")
        repo = re.sub(r"\.git$", "", repo)

    return {
        "enabled": enabled,
        "token": token,
        "repo": repo,
        "branch": branch,
        "data_path": str(data_path).strip("/"),
    }


def github_api_request(
    method: str,
    path: str,
    token: str,
    payload: dict | None = None,
) -> dict:
    data = json.dumps(payload).encode("utf-8") if payload is not None else None
    request = urllib.request.Request(
        f"https://api.github.com{path}",
        data=data,
        method=method,
        headers={
            "Accept": "application/vnd.github+json",
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
            "X-GitHub-Api-Version": "2022-11-28",
            "User-Agent": "bls-norms-streamlit-app",
        },
    )
    try:
        with urllib.request.urlopen(request, timeout=30) as response:
            response_body = response.read().decode("utf-8")
    except urllib.error.HTTPError as exc:
        error_body = exc.read().decode("utf-8", errors="replace")
        try:
            error_message = json.loads(error_body).get("message", error_body)
        except json.JSONDecodeError:
            error_message = error_body
        raise RuntimeError(f"GitHub API error {exc.code}: {error_message}") from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f"GitHub API connection error: {exc.reason}") from exc

    return json.loads(response_body) if response_body else {}


def uploaded_dataset_commit_files(data_path: str) -> list[tuple[str, bytes]]:
    files: list[tuple[str, bytes]] = []
    if not UPLOADED_DATASETS_DIR.exists():
        return files

    for path in sorted(UPLOADED_DATASETS_DIR.rglob("*")):
        if not path.is_file() or path.name == ".DS_Store":
            continue
        relative_path = path.relative_to(UPLOADED_DATASETS_DIR).as_posix()
        files.append((f"{data_path}/{relative_path}", path.read_bytes()))

    return files


def github_autocommit_uploaded_datasets(commit_message: str) -> tuple[bool, str]:
    config = github_autocommit_config()
    if not config["enabled"]:
        return False, "GitHub autocommit is not configured."
    if not config["token"] or not config["repo"]:
        return False, "GitHub autocommit needs a token and repo in Streamlit secrets."

    files = uploaded_dataset_commit_files(config["data_path"])
    if not files:
        return False, "No uploaded dataset backup files were available to commit."

    repo_path = f"/repos/{config['repo']}"
    branch = config["branch"]

    try:
        ref = github_api_request(
            "GET",
            f"{repo_path}/git/ref/heads/{branch}",
            config["token"],
        )
        parent_sha = ref["object"]["sha"]
        parent_commit = github_api_request(
            "GET",
            f"{repo_path}/git/commits/{parent_sha}",
            config["token"],
        )
        base_tree_sha = parent_commit["tree"]["sha"]

        tree_entries = []
        for file_path, content in files:
            blob = github_api_request(
                "POST",
                f"{repo_path}/git/blobs",
                config["token"],
                {
                    "content": base64.b64encode(content).decode("ascii"),
                    "encoding": "base64",
                },
            )
            tree_entries.append(
                {
                    "path": file_path,
                    "mode": "100644",
                    "type": "blob",
                    "sha": blob["sha"],
                }
            )

        tree = github_api_request(
            "POST",
            f"{repo_path}/git/trees",
            config["token"],
            {
                "base_tree": base_tree_sha,
                "tree": tree_entries,
            },
        )
        if tree["sha"] == base_tree_sha:
            return True, "GitHub autocommit found no file changes."

        commit = github_api_request(
            "POST",
            f"{repo_path}/git/commits",
            config["token"],
            {
                "message": commit_message,
                "tree": tree["sha"],
                "parents": [parent_sha],
            },
        )
        github_api_request(
            "PATCH",
            f"{repo_path}/git/refs/heads/{branch}",
            config["token"],
            {
                "sha": commit["sha"],
                "force": False,
            },
        )
    except (KeyError, RuntimeError) as exc:
        return False, f"GitHub autocommit failed: {exc}"

    return True, f"GitHub autocommit saved `{config['data_path']}` to `{config['repo']}`."


def save_message_with_github_status(
    base_message: str,
    commit_ok: bool,
    commit_message: str,
) -> str:
    if commit_message == "GitHub autocommit found no file changes.":
        return base_message
    if commit_message == "GitHub autocommit is not configured.":
        return (
            f"{base_message} GitHub backup is not configured, so this may not "
            "survive a Streamlit Cloud restart."
        )
    if commit_message == "GitHub autocommit needs a token and repo in Streamlit secrets.":
        return (
            f"{base_message} GitHub backup needs a token and repo in Streamlit "
            "secrets before saves are permanent across restarts."
        )
    if commit_ok:
        return f"{base_message} {commit_message}"
    return f"{base_message} {commit_message}"


def migrate_legacy_norm_database_if_needed() -> None:
    if NORM_DATABASE_DIR == NORM_DATABASE_LEGACY_DIR:
        restore_norm_database_from_uploaded_datasets_if_needed()
        return
    if not NORM_DATABASE_LEGACY_DIR.exists() or NORM_DATABASE_MANIFEST_PATH.exists():
        restore_norm_database_from_uploaded_datasets_if_needed()
        return

    try:
        NORM_DATABASE_DIR.mkdir(parents=True, exist_ok=True)
        shutil.copytree(
            NORM_DATABASE_LEGACY_DIR,
            NORM_DATABASE_DIR,
            dirs_exist_ok=True,
        )
    except OSError:
        return
    restore_norm_database_from_uploaded_datasets_if_needed()


def ensure_norm_database_dirs() -> None:
    migrate_legacy_norm_database_if_needed()
    NORM_DATABASE_DATASETS_DIR.mkdir(parents=True, exist_ok=True)


@contextmanager
def norm_database_write_lock(timeout_seconds: int = 30):
    migrate_legacy_norm_database_if_needed()
    NORM_DATABASE_DIR.mkdir(parents=True, exist_ok=True)
    lock_fd = None
    start_time = time.time()

    while True:
        try:
            lock_fd = os.open(
                NORM_DATABASE_LOCK_PATH,
                os.O_CREAT | os.O_EXCL | os.O_WRONLY,
            )
            os.write(
                lock_fd,
                f"{os.getpid()} {datetime.now().isoformat()}".encode("utf-8"),
            )
            break
        except FileExistsError:
            try:
                lock_age = time.time() - NORM_DATABASE_LOCK_PATH.stat().st_mtime
                if lock_age > 300:
                    NORM_DATABASE_LOCK_PATH.unlink()
                    continue
            except FileNotFoundError:
                continue

            if time.time() - start_time > timeout_seconds:
                raise TimeoutError("Norms database is busy. Try again in a moment.")
            time.sleep(0.1)

    try:
        yield
    finally:
        if lock_fd is not None:
            os.close(lock_fd)
        try:
            NORM_DATABASE_LOCK_PATH.unlink()
        except FileNotFoundError:
            pass


def load_norm_database_manifest() -> list[dict]:
    migrate_legacy_norm_database_if_needed()
    records = read_manifest_records(NORM_DATABASE_MANIFEST_PATH)
    if records and not UPLOADED_MANIFEST_BACKUP_PATH.exists():
        backup_norm_database_to_uploaded_datasets(records)
    return records


def save_norm_database_manifest(records: list[dict]) -> None:
    ensure_norm_database_dirs()
    temp_path = NORM_DATABASE_MANIFEST_PATH.with_suffix(".json.tmp")
    temp_path.write_text(json.dumps(records, indent=2, sort_keys=True) + "\n")
    temp_path.replace(NORM_DATABASE_MANIFEST_PATH)


def norm_upload_fingerprint(workbook_bytes: bytes, data_sheet: str | None) -> str:
    digest = hashlib.sha256()
    digest.update(workbook_bytes)
    digest.update(b"\0")
    digest.update((data_sheet or "").encode("utf-8"))
    return digest.hexdigest()


def normalized_hash(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def respondent_id_column_for_data(
    df: pd.DataFrame,
    question_labels: dict[str, str] | None = None,
) -> str | None:
    question_labels = question_labels or {}
    candidate_keys = {
        normalize_alias_key(candidate)
        for candidate in RESPONDENT_ID_CANDIDATES
        if normalize_alias_key(candidate)
    }

    for column in df.columns:
        if normalize_alias_key(column) in candidate_keys:
            return column

    for column in df.columns:
        label = question_labels.get(column)
        if label and normalize_alias_key(label) in candidate_keys:
            return column

    return None


def respondent_id_hashes_for_data(
    df: pd.DataFrame,
    question_labels: dict[str, str] | None = None,
    selected_column: str | None = None,
) -> tuple[str | None, list[str]]:
    if selected_column == NO_RESPONDENT_ID_OPTION:
        return None, []

    if selected_column is not None:
        column = selected_column if selected_column in df.columns else None
    else:
        column = respondent_id_column_for_data(df, question_labels)

    if column is None or column not in df.columns:
        return None, []

    normalized_values = [
        normalize_answer(value)
        for value in df[column].tolist()
    ]
    unique_values = sorted({value for value in normalized_values if value})
    return column, [normalized_hash(value) for value in unique_values]


def normalize_respondent_id_selection(value: object, columns: list[str]) -> str | None:
    selected_value = normalize_answer(value)
    if not selected_value or selected_value == NO_RESPONDENT_ID_OPTION:
        return None
    return selected_value if selected_value in columns else None


def display_variable_with_label(variable: str, question_labels: dict[str, str]) -> str:
    label = display_question_label(variable, question_labels)
    return variable if label == variable else f"{variable} - {label}"


def render_respondent_id_selector(
    df: pd.DataFrame,
    question_labels: dict[str, str],
    key: str,
    current_column: str | None = None,
) -> str | None:
    columns = list(df.columns)
    detected_column = respondent_id_column_for_data(df, question_labels)
    current_column = normalize_respondent_id_selection(current_column, columns)
    default_column = current_column or detected_column
    options = [NO_RESPONDENT_ID_OPTION, *columns]
    default_index = options.index(default_column) if default_column in options else 0

    selected_column = st.selectbox(
        "Respondent ID variable",
        options,
        index=default_index,
        key=key,
        format_func=lambda value: (
            NO_RESPONDENT_ID_OPTION
            if value == NO_RESPONDENT_ID_OPTION
            else display_variable_with_label(value, question_labels)
        ),
        help=(
            "Used to detect duplicate uploads. Select the respondent ID field if "
            "the app cannot detect it automatically."
        ),
    )
    respondent_id_column = (
        NO_RESPONDENT_ID_OPTION
        if selected_column == NO_RESPONDENT_ID_OPTION
        else normalize_respondent_id_selection(selected_column, columns)
    )
    _column, respondent_hashes = respondent_id_hashes_for_data(
        df,
        question_labels,
        respondent_id_column,
    )

    if respondent_id_column and respondent_id_column != NO_RESPONDENT_ID_OPTION:
        st.caption(
            f"Duplicate check will use `{respondent_id_column}` "
            f"with {len(respondent_hashes):,} unique respondent IDs."
        )
    elif detected_column:
        st.caption(
            f"Detected `{detected_column}`, but duplicate check is set to no respondent ID."
        )
    else:
        st.caption(
            "No respondent ID selected. Duplicate detection will only catch exact cleaned-data matches."
        )

    return respondent_id_column


def dataframe_content_fingerprint(
    df: pd.DataFrame,
    ignore_project_metadata: bool = True,
) -> str:
    columns = [
        column
        for column in df.columns
        if not ignore_project_metadata or not is_project_metadata_variable(column)
    ]
    normalized_df = df.loc[:, columns].copy()
    for column in normalized_df.columns:
        normalized_df[column] = normalized_df[column].map(
            lambda value: normalize_answer(value) or ""
        )
    payload = normalized_df.to_json(orient="split", date_format="iso", default_handler=str)
    return normalized_hash(payload)


def respondent_id_overlap_stats(
    new_hashes: list[str],
    saved_hashes: list[str],
) -> dict[str, float | int]:
    new_set = set(new_hashes)
    saved_set = set(saved_hashes)
    overlap_count = len(new_set & saved_set)
    new_count = len(new_set)
    saved_count = len(saved_set)
    new_overlap = overlap_count / new_count if new_count else 0
    saved_overlap = overlap_count / saved_count if saved_count else 0
    return {
        "overlap_count": overlap_count,
        "new_count": new_count,
        "saved_count": saved_count,
        "new_overlap": new_overlap,
        "saved_overlap": saved_overlap,
        "max_overlap": max(new_overlap, saved_overlap),
    }


def duplicate_match_label(match: dict | None) -> str:
    if not match:
        return ""

    record = match.get("record", {})
    if match.get("method") == "respondent_id_overlap":
        return (
            f"{match.get('overlap_percent', 0):.1f}% respondent-ID overlap with "
            f"{record.get('uploaded_file', NOT_AVAILABLE)} "
            f"({record.get('dataset_id', NOT_AVAILABLE)})"
        )

    return (
        "Exact data-content match with "
        f"{record.get('uploaded_file', NOT_AVAILABLE)} "
        f"({record.get('dataset_id', NOT_AVAILABLE)})"
    )


def norm_database_dataset_path(dataset_id: str) -> Path:
    return NORM_DATABASE_DATASETS_DIR / f"{dataset_id}.xlsx"


def find_norm_database_duplicate_record(
    records: list[dict],
    new_record: dict,
    overlap_threshold: float = DUPLICATE_RESPONDENT_ID_OVERLAP_THRESHOLD,
) -> dict | None:
    new_hashes = new_record.get("respondent_id_hashes") or []
    new_content_fingerprint = new_record.get("data_content_fingerprint")

    for saved_record in records:
        saved_hashes = saved_record.get("respondent_id_hashes") or []
        if new_hashes and saved_hashes:
            stats = respondent_id_overlap_stats(new_hashes, saved_hashes)
            if stats["max_overlap"] >= overlap_threshold:
                return {
                    "record": saved_record,
                    "method": "respondent_id_overlap",
                    "overlap_percent": stats["max_overlap"] * 100,
                    **stats,
                }
            continue

        saved_content_fingerprints = {
            saved_record.get("data_content_fingerprint"),
            saved_record.get("full_data_content_fingerprint"),
        }
        if new_content_fingerprint and new_content_fingerprint in saved_content_fingerprints:
            return {
                "record": saved_record,
                "method": "data_content_fingerprint",
                "overlap_percent": 100,
            }

        if (
            new_record.get("upload_fingerprint")
            and saved_record.get("upload_fingerprint") == new_record.get("upload_fingerprint")
        ):
            return {
                "record": saved_record,
                "method": "legacy_upload_fingerprint",
                "overlap_percent": 100,
            }

    return None


def norm_database_record_for_upload(
    workbook_bytes: bytes,
    uploaded_file_name: str | None,
    data_sheet: str | None,
    data: pd.DataFrame,
    group_column: str,
    control_label: str,
    test_label: str,
    norm_questions: list[str],
    included_mappings: dict[str, str],
    denominator_settings: dict[str, str],
    question_labels: dict[str, str] | None = None,
    respondent_id_column: str | None = None,
) -> dict:
    upload_fingerprint = norm_upload_fingerprint(workbook_bytes, data_sheet)
    respondent_id_column, respondent_id_hashes = respondent_id_hashes_for_data(
        data,
        question_labels,
        respondent_id_column,
    )
    content_fingerprint = dataframe_content_fingerprint(data)
    full_content_fingerprint = dataframe_content_fingerprint(data, ignore_project_metadata=False)
    dataset_id = (
        normalized_hash("|".join(respondent_id_hashes))[:16]
        if respondent_id_hashes
        else content_fingerprint[:16]
    )
    metadata_values = project_metadata_values(data)
    return {
        "dataset_id": dataset_id,
        "upload_fingerprint": upload_fingerprint,
        "data_content_fingerprint": content_fingerprint,
        "full_data_content_fingerprint": full_content_fingerprint,
        "respondent_id_column": respondent_id_column or NOT_AVAILABLE,
        "respondent_id_count": len(respondent_id_hashes),
        "respondent_id_hashes": respondent_id_hashes,
        "duplicate_overlap_threshold": DUPLICATE_RESPONDENT_ID_OVERLAP_THRESHOLD,
        "uploaded_file": uploaded_file_name or NOT_AVAILABLE,
        "data_sheet": data_sheet or NOT_AVAILABLE,
        "row_count": int(len(data)),
        "column_count": int(len(data.columns)),
        "norm_count": int(len(norm_questions)),
        "group_column": group_column,
        "control_label": control_label,
        "test_label": test_label,
        "metadata_values": metadata_values,
        "norm_mappings": {
            question: included_mappings.get(question, question)
            for question in norm_questions
        },
        "denominator_settings": {
            question: denominator_settings.get(question, DEFAULT_DENOMINATOR)
            for question in norm_questions
        },
    }


def norm_database_duplicate_probe_for_data(
    workbook_bytes: bytes,
    uploaded_file_name: str | None,
    data_sheet: str | None,
    data: pd.DataFrame,
    question_labels: dict[str, str] | None = None,
    respondent_id_column: str | None = None,
) -> dict:
    respondent_id_column, respondent_id_hashes = respondent_id_hashes_for_data(
        data,
        question_labels,
        respondent_id_column,
    )
    content_fingerprint = dataframe_content_fingerprint(data)
    full_content_fingerprint = dataframe_content_fingerprint(data, ignore_project_metadata=False)
    return {
        "dataset_id": (
            normalized_hash("|".join(respondent_id_hashes))[:16]
            if respondent_id_hashes
            else content_fingerprint[:16]
        ),
        "upload_fingerprint": norm_upload_fingerprint(workbook_bytes, data_sheet),
        "data_content_fingerprint": content_fingerprint,
        "full_data_content_fingerprint": full_content_fingerprint,
        "respondent_id_column": respondent_id_column or NOT_AVAILABLE,
        "respondent_id_count": len(respondent_id_hashes),
        "respondent_id_hashes": respondent_id_hashes,
        "uploaded_file": uploaded_file_name or NOT_AVAILABLE,
        "data_sheet": data_sheet or NOT_AVAILABLE,
        "row_count": int(len(data)),
        "column_count": int(len(data.columns)),
    }


def norm_database_rules_for_upload(
    source_questions: list[str],
    norm_questions: list[str],
    included_mappings: dict[str, str],
    group_column: str,
    control_label: str,
    test_label: str,
    denominator_settings: dict[str, str],
    split_multi_select: bool,
    delimiter: str,
    response_labels: dict[str, dict[str, str]],
    question_labels: dict[str, str],
    box_score_settings: dict[str, list[str]],
    saved_box_score_settings: dict[str, list[str]],
    question_type_settings: dict[str, str],
    saved_question_type_settings: dict[str, str],
) -> dict:
    norm_rules = []
    norm_question_set = set(norm_questions)
    for question in source_questions:
        is_included = question in norm_question_set
        norm_rules.append(
            {
                "Source variable": question,
                "Include": is_included,
                "Norm / benchmark": (
                    included_mappings.get(question, question)
                    if is_included
                    else NA_NORM_OPTION
                ),
                "Denominator": denominator_settings.get(question, DEFAULT_DENOMINATOR),
                "Question Type": question_type_settings.get(
                    question,
                    saved_question_type_settings.get(question, "Single-Select"),
                ),
                "Box scores": ";".join(
                    box_score_settings.get(
                        question,
                        saved_box_score_settings.get(question, []),
                    )
                ),
            }
        )

    return {
        "version": "1",
        "group_column": group_column,
        "control_label": control_label,
        "test_label": test_label,
        "split_multi_select": split_multi_select,
        "delimiter": delimiter,
        "norm_rules": norm_rules,
        "response_labels": response_labels,
        "question_labels": question_labels,
    }


def flatten_norm_database_record(record: dict) -> dict:
    metadata_values = record.get("metadata_values", {})
    row = {
        "Dataset ID": record.get("dataset_id", NOT_AVAILABLE),
        "Saved at": record.get("saved_at", NOT_AVAILABLE),
        "Uploaded file": record.get("uploaded_file", NOT_AVAILABLE),
        "Data sheet": record.get("data_sheet", NOT_AVAILABLE),
        "Rows": record.get("row_count", NOT_AVAILABLE),
        "Columns": record.get("column_count", NOT_AVAILABLE),
        "Norms": record.get("norm_count", NOT_AVAILABLE),
        "Control/test group column": record.get("group_column", NOT_AVAILABLE),
        "Control label": record.get("control_label", NOT_AVAILABLE),
        "Test label": record.get("test_label", NOT_AVAILABLE),
        "Respondent ID column": record.get("respondent_id_column", NOT_AVAILABLE),
        "Respondent IDs": record.get("respondent_id_count", NOT_AVAILABLE),
    }
    for variable in PROJECT_METADATA_VARIABLES:
        row[variable] = metadata_values.get(variable, NOT_AVAILABLE)
    return row


def add_norm_database_context(
    table: pd.DataFrame,
    record: dict,
    metric: str,
    source_question: str,
) -> pd.DataFrame:
    metadata_values = record.get("metadata_values", {})
    context = {
        "Dataset ID": record.get("dataset_id", NOT_AVAILABLE),
        "Saved at": record.get("saved_at", NOT_AVAILABLE),
        "Uploaded file": record.get("uploaded_file", NOT_AVAILABLE),
        "Data sheet": record.get("data_sheet", NOT_AVAILABLE),
        "Metric": metric,
        "Source variable": source_question,
        "Control/test group column": record.get("group_column", NOT_AVAILABLE),
        "Control label": record.get("control_label", NOT_AVAILABLE),
        "Test label": record.get("test_label", NOT_AVAILABLE),
    }
    for variable in PROJECT_METADATA_VARIABLES:
        context[variable] = metadata_values.get(variable, NOT_AVAILABLE)

    context_rows = pd.DataFrame([context] * len(table))
    return pd.concat([context_rows, table.reset_index(drop=True)], axis=1)


def contextualize_norm_database_tables(
    tables: dict[str, pd.DataFrame],
    record: dict,
) -> pd.DataFrame:
    contextualized_tables = []
    for table_key, table in tables.items():
        metric, source_question = (
            table_key.split("__", 1)
            if "__" in table_key
            else (table_key, table_key)
        )
        contextualized_tables.append(
            add_norm_database_context(table, record, metric, source_question)
        )

    if not contextualized_tables:
        return pd.DataFrame()
    return pd.concat(contextualized_tables, ignore_index=True)


def write_norm_dataset_rules(
    writer: pd.ExcelWriter,
    rules: dict,
) -> None:
    pd.DataFrame(
        [
            {"Rule": "version", "Value": rules.get("version", "1")},
            {"Rule": "group_column", "Value": rules.get("group_column", "")},
            {"Rule": "control_label", "Value": rules.get("control_label", "")},
            {"Rule": "test_label", "Value": rules.get("test_label", "")},
            {"Rule": "split_multi_select", "Value": str(rules.get("split_multi_select", False))},
            {"Rule": "delimiter", "Value": rules.get("delimiter", ";")},
        ]
    ).to_excel(writer, sheet_name=NORM_DATASET_RULES_SHEET, index=False)

    pd.DataFrame(
        rules.get("norm_rules", []),
        columns=[
            "Source variable",
            "Include",
            "Norm / benchmark",
            "Denominator",
            "Question Type",
            "Box scores",
        ],
    ).to_excel(writer, sheet_name=NORM_DATASET_NORM_RULES_SHEET, index=False)

    pd.DataFrame(
        [
            {"Source variable": question, "Question label": label}
            for question, label in rules.get("question_labels", {}).items()
        ],
        columns=["Source variable", "Question label"],
    ).to_excel(writer, sheet_name=NORM_DATASET_QUESTION_LABELS_SHEET, index=False)

    response_label_rows = []
    for question, labels in rules.get("response_labels", {}).items():
        for value, label in labels.items():
            response_label_rows.append(
                {
                    "Source variable": question,
                    "Response value": value,
                    "Response label": label,
                }
            )
    pd.DataFrame(
        response_label_rows,
        columns=["Source variable", "Response value", "Response label"],
    ).to_excel(writer, sheet_name=NORM_DATASET_RESPONSE_LABELS_SHEET, index=False)


def read_norm_dataset_rules(dataset_path: Path) -> dict | None:
    try:
        rules_df = pd.read_excel(dataset_path, sheet_name=NORM_DATASET_RULES_SHEET, dtype=object)
        norm_rules_df = pd.read_excel(
            dataset_path,
            sheet_name=NORM_DATASET_NORM_RULES_SHEET,
            dtype=object,
        )
    except Exception:
        return None

    rule_lookup = {
        normalize_answer(row.get("Rule")): normalize_answer(row.get("Value"))
        for _index, row in rules_df.iterrows()
        if normalize_answer(row.get("Rule"))
    }

    question_labels: dict[str, str] = {}
    try:
        question_labels_df = pd.read_excel(
            dataset_path,
            sheet_name=NORM_DATASET_QUESTION_LABELS_SHEET,
            dtype=object,
        )
        for _index, row in question_labels_df.iterrows():
            question = normalize_answer(row.get("Source variable"))
            label = normalize_answer(row.get("Question label"))
            if question and label:
                question_labels[question] = label
    except Exception:
        question_labels = {}

    response_labels: dict[str, dict[str, str]] = {}
    try:
        response_labels_df = pd.read_excel(
            dataset_path,
            sheet_name=NORM_DATASET_RESPONSE_LABELS_SHEET,
            dtype=object,
        )
        for _index, row in response_labels_df.iterrows():
            question = normalize_answer(row.get("Source variable"))
            value = normalize_answer(row.get("Response value"))
            label = normalize_answer(row.get("Response label"))
            if question and value and label:
                response_labels.setdefault(question, {})[value] = label
    except Exception:
        response_labels = {}

    norm_rules = []
    for _index, row in norm_rules_df.iterrows():
        source_variable = normalize_answer(row.get("Source variable"))
        if not source_variable:
            continue
        box_scores = [
            score
            for score in re.split(r"[;,]", normalize_answer(row.get("Box scores")) or "")
            if score in BOX_SCORE_OPTIONS
        ]
        norm_rules.append(
            {
                "Source variable": source_variable,
                "Include": str(row.get("Include", True)).lower() != "false",
                "Norm / benchmark": normalize_answer(row.get("Norm / benchmark")) or source_variable,
                "Denominator": (
                    normalize_answer(row.get("Denominator"))
                    if normalize_answer(row.get("Denominator")) in DENOMINATOR_OPTIONS
                    else DEFAULT_DENOMINATOR
                ),
                "Question Type": normalize_question_type(row.get("Question Type")),
                "Box scores": ";".join(box_scores),
            }
        )

    return {
        "version": rule_lookup.get("version", "1"),
        "group_column": rule_lookup.get("group_column", ""),
        "control_label": rule_lookup.get("control_label", ""),
        "test_label": rule_lookup.get("test_label", ""),
        "split_multi_select": str(rule_lookup.get("split_multi_select", "False")).lower() == "true",
        "delimiter": rule_lookup.get("delimiter", ";"),
        "norm_rules": norm_rules,
        "response_labels": response_labels,
        "question_labels": question_labels,
    }


def write_norm_dataset_workbook(
    record: dict,
    tables: dict[str, pd.DataFrame],
    source_data: pd.DataFrame | None = None,
    rules: dict | None = None,
) -> None:
    ensure_norm_database_dirs()
    dataset_path = norm_database_dataset_path(record["dataset_id"])
    contextualized_table = normalize_lift_output_table(
        contextualize_norm_database_tables(tables, record)
    )

    with pd.ExcelWriter(dataset_path, engine="openpyxl") as writer:
        pd.DataFrame([flatten_norm_database_record(record)]).to_excel(
            writer,
            sheet_name="Dataset",
            index=False,
        )
        contextualized_table.to_excel(writer, sheet_name="Norm Tables", index=False)
        if source_data is not None:
            source_data.to_excel(writer, sheet_name=NORM_DATASET_RESPONDENT_SHEET, index=False)
        if rules is not None:
            write_norm_dataset_rules(writer, rules)


def refresh_norm_database_workbook(records: list[dict]) -> None:
    ensure_norm_database_dirs()
    dataset_rows = [flatten_norm_database_record(record) for record in records]
    norm_tables = []

    for record in records:
        dataset_path = norm_database_dataset_path(record.get("dataset_id", ""))
        if not dataset_path.exists():
            continue
        try:
            norm_tables.append(
                normalize_lift_output_table(
                    pd.read_excel(dataset_path, sheet_name="Norm Tables")
                )
            )
        except Exception:
            continue

    with pd.ExcelWriter(NORM_DATABASE_WORKBOOK_PATH, engine="openpyxl") as writer:
        pd.DataFrame(dataset_rows).to_excel(writer, sheet_name="Datasets", index=False)
        if norm_tables:
            saved_norm_tables = normalize_lift_output_table(
                pd.concat(norm_tables, ignore_index=True)
            )
            saved_norm_tables.to_excel(
                writer,
                sheet_name="All Norms",
                index=False,
            )
        else:
            pd.DataFrame().to_excel(writer, sheet_name="All Norms", index=False)


def save_norm_dataset_to_database(
    record: dict,
    tables: dict[str, pd.DataFrame],
    source_data: pd.DataFrame | None = None,
    rules: dict | None = None,
    raw_workbook_bytes: bytes | None = None,
    replace_existing: bool = False,
) -> tuple[bool, str]:
    if not tables:
        return False, "No norm tables are available to save."

    saved_record_for_commit = None
    try:
        with norm_database_write_lock():
            records = load_norm_database_manifest()
            duplicate_match = find_norm_database_duplicate_record(
                records,
                record,
                DUPLICATE_RESPONDENT_ID_OVERLAP_THRESHOLD,
            )
            existing_index = (
                records.index(duplicate_match["record"])
                if duplicate_match and duplicate_match.get("record") in records
                else None
            )
            if existing_index is not None and not replace_existing:
                return False, f"Possible duplicate upload: {duplicate_match_label(duplicate_match)}."

            if existing_index is not None:
                backup_norm_database_to_uploaded_datasets(records)
                create_norm_database_history_snapshot(
                    "Before replace dataset",
                    records,
                    records[existing_index],
                    "Automatic safety snapshot before replacing a saved dataset.",
                )

            saved_record = {
                **record,
                "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
            if existing_index is not None:
                saved_record["dataset_id"] = records[existing_index].get(
                    "dataset_id",
                    saved_record["dataset_id"],
                )
            write_norm_dataset_workbook(saved_record, tables, source_data, rules)

            if existing_index is None:
                records.append(saved_record)
            else:
                records[existing_index] = saved_record

            save_norm_database_manifest(records)
            refresh_norm_database_workbook(records)
            backup_norm_database_to_uploaded_datasets(
                records,
                saved_record,
                raw_workbook_bytes,
                rules,
            )
            create_norm_database_history_snapshot(
                "Replace dataset" if existing_index is not None else "Save dataset",
                records,
                saved_record,
            )
            saved_record_for_commit = saved_record
    except TimeoutError as exc:
        return False, str(exc)

    dataset_id = (
        normalize_answer(saved_record_for_commit.get("dataset_id"))
        if saved_record_for_commit
        else "dataset"
    )
    commit_ok, commit_message = github_autocommit_uploaded_datasets(
        f"Save BLS norms dataset {dataset_id}"
    )
    return True, save_message_with_github_status(
        "Dataset has been added.",
        commit_ok,
        commit_message,
    )


def render_norm_database_save_controls(
    record: dict,
    tables: dict[str, pd.DataFrame],
    source_data: pd.DataFrame,
    rules: dict,
    raw_workbook_bytes: bytes | None = None,
    standards_change_issues: list[dict] | None = None,
    standards_change_confirmed: bool = True,
) -> None:
    st.subheader("Save to norms database")
    st.caption(
        "Saves the full audited upload, not temporary table filters. "
        "Duplicate uploads are detected by respondent-ID overlap when available."
    )
    render_persistence_status()

    saved_records = load_norm_database_manifest()
    duplicate_match = find_norm_database_duplicate_record(
        saved_records,
        record,
        DUPLICATE_RESPONDENT_ID_OVERLAP_THRESHOLD,
    )
    existing_record = duplicate_match.get("record") if duplicate_match else None
    standards_change_issues = standards_change_issues or []
    standards_change_blocked = bool(standards_change_issues) and not standards_change_confirmed

    status_cols = st.columns(3)
    status_cols[0].metric("Saved datasets", f"{len(saved_records):,}")
    status_cols[1].metric("Rows to save", f"{record.get('row_count', 0):,}")
    status_cols[2].metric("Norms to save", f"{record.get('norm_count', 0):,}")
    if record.get("respondent_id_count", 0):
        st.caption(
            "Duplicate check: "
            f"`{record.get('respondent_id_column')}` respondent IDs, "
            f"{DUPLICATE_RESPONDENT_ID_OVERLAP_THRESHOLD:.0%} overlap threshold."
        )
    else:
        st.caption(
            "Duplicate check: no respondent ID field found, so the app will only "
            "block exact cleaned-data matches."
        )

    if standards_change_issues:
        if standards_change_blocked:
            st.warning(
                "This upload changes box-score rules from previous saved norms. "
                "Confirm the standards change on the Survey Question Audit page before saving."
            )
        else:
            st.info(
                "This upload has confirmed box-score rule changes. After saving, review "
                "Saved datasets and update previous datasets so the metrics stay consistent."
            )

    if existing_record:
        st.warning(
            "Possible duplicate upload detected: "
            f"{duplicate_match_label(duplicate_match)}. "
            f"Saved at {existing_record.get('saved_at', NOT_AVAILABLE)} "
            f"as dataset {existing_record.get('dataset_id', NOT_AVAILABLE)}."
        )
        replace_existing = st.checkbox(
            "Replace the existing saved dataset",
            key="replace_existing_norm_dataset",
        )
        if st.button(
            "Replace saved dataset",
            disabled=not replace_existing or standards_change_blocked,
            use_container_width=True,
        ):
            success, message = save_norm_dataset_to_database(
                record,
                tables,
                source_data,
                rules,
                raw_workbook_bytes,
                replace_existing=True,
            )
            if success:
                st.success("Dataset has been updated.")
            else:
                st.error(message)
    else:
        if st.button(
            "Save dataset to norms database",
            disabled=standards_change_blocked,
            use_container_width=True,
        ):
            success, message = save_norm_dataset_to_database(
                record,
                tables,
                source_data,
                rules,
                raw_workbook_bytes,
            )
            if success:
                st.success(message)
            else:
                st.error(message)


def saved_dataset_option_label(record: dict) -> str:
    return (
        f"{record.get('uploaded_file', NOT_AVAILABLE)} | "
        f"{record.get('data_sheet', NOT_AVAILABLE)} | "
        f"{record.get('dataset_id', NOT_AVAILABLE)}"
    )


def saved_dataset_rules_editor_frame(rules: dict, data_columns: list[str]) -> pd.DataFrame:
    rules_by_source = {
        normalize_answer(rule.get("Source variable")): rule
        for rule in rules.get("norm_rules", [])
        if normalize_answer(rule.get("Source variable"))
    }
    rows = []

    for source_variable in data_columns:
        rule = rules_by_source.get(source_variable, {})
        box_scores = [
            score
            for score in re.split(r"[;,]", normalize_answer(rule.get("Box scores")) or "")
            if score in BOX_SCORE_OPTIONS
        ]
        include = bool(rule.get("Include", False))
        norm = normalize_answer(rule.get("Norm / benchmark")) or source_variable
        rows.append(
            {
                "Include": include,
                "Source variable": source_variable,
                "Norm / benchmark": norm if include else NA_NORM_OPTION,
                "Denominator": (
                    rule.get("Denominator")
                    if rule.get("Denominator") in DENOMINATOR_OPTIONS
                    else DEFAULT_DENOMINATOR
                ),
                "Question Type": normalize_question_type(rule.get("Question Type")),
                "T2B": "T2B" in box_scores,
                "T3B": "T3B" in box_scores,
                "B2B": "B2B" in box_scores,
                "B3B": "B3B" in box_scores,
            }
        )

    return pd.DataFrame(rows)


def normalize_saved_dataset_rules_editor(
    editor_df: pd.DataFrame,
) -> tuple[dict[str, str], dict[str, str], dict[str, list[str]], dict[str, str], list[dict]]:
    included_mappings: dict[str, str] = {}
    denominator_settings: dict[str, str] = {}
    box_score_settings: dict[str, list[str]] = {}
    question_type_settings: dict[str, str] = {}
    norm_rules: list[dict] = []

    for _index, row in editor_df.iterrows():
        source_variable = normalize_answer(row.get("Source variable"))
        if not source_variable:
            continue

        include = bool(row.get("Include"))
        mapped_norm = normalize_answer(row.get("Norm / benchmark")) or source_variable
        denominator = normalize_answer(row.get("Denominator"))
        if denominator not in DENOMINATOR_OPTIONS:
            denominator = DEFAULT_DENOMINATOR

        question_type = normalize_question_type(row.get("Question Type"))
        selected_box_scores = [
            score
            for score in BOX_SCORE_OPTIONS
            if bool(row.get(score))
        ]

        if include and mapped_norm != NA_NORM_OPTION:
            included_mappings[source_variable] = mapped_norm
            denominator_settings[source_variable] = denominator
            box_score_settings[source_variable] = selected_box_scores
            question_type_settings[source_variable] = question_type

        norm_rules.append(
            {
                "Source variable": source_variable,
                "Include": include and mapped_norm != NA_NORM_OPTION,
                "Norm / benchmark": (
                    mapped_norm
                    if include and mapped_norm != NA_NORM_OPTION
                    else NA_NORM_OPTION
                ),
                "Denominator": denominator,
                "Question Type": question_type,
                "Box scores": ";".join(selected_box_scores),
            }
        )

    return (
        included_mappings,
        denominator_settings,
        box_score_settings,
        question_type_settings,
        norm_rules,
    )


def update_saved_norm_dataset_rules(
    record: dict,
    source_data: pd.DataFrame,
    updated_rules: dict,
    tables: dict[str, pd.DataFrame],
    respondent_id_column: str | None = None,
) -> tuple[bool, str]:
    updated_record_for_commit = None
    try:
        with norm_database_write_lock():
            records = load_norm_database_manifest()
            dataset_id = record.get("dataset_id")
            record_index = next(
                (
                    index
                    for index, saved_record in enumerate(records)
                    if saved_record.get("dataset_id") == dataset_id
                ),
                None,
            )
            if record_index is None:
                return False, "Saved dataset record was not found in the manifest."

            backup_norm_database_to_uploaded_datasets(records)
            create_norm_database_history_snapshot(
                "Before update rules",
                records,
                records[record_index],
                "Automatic safety snapshot before updating saved dataset rules.",
            )

            included_mappings = {
                rule["Source variable"]: rule["Norm / benchmark"]
                for rule in updated_rules.get("norm_rules", [])
                if rule.get("Include") and rule.get("Norm / benchmark") != NA_NORM_OPTION
            }
            denominator_settings = {
                rule["Source variable"]: rule.get("Denominator", DEFAULT_DENOMINATOR)
                for rule in updated_rules.get("norm_rules", [])
                if rule.get("Include") and rule.get("Norm / benchmark") != NA_NORM_OPTION
            }
            updated_respondent_id_column, updated_respondent_id_hashes = (
                respondent_id_hashes_for_data(
                    source_data,
                    updated_rules.get("question_labels", {}),
                    respondent_id_column,
                )
            )
            updated_record = {
                **record,
                "row_count": int(len(source_data)),
                "column_count": int(len(source_data.columns)),
                "norm_count": int(len(included_mappings)),
                "group_column": updated_rules.get("group_column", record.get("group_column")),
                "control_label": updated_rules.get("control_label", record.get("control_label")),
                "test_label": updated_rules.get("test_label", record.get("test_label")),
                "metadata_values": project_metadata_values(source_data),
                "norm_mappings": included_mappings,
                "denominator_settings": denominator_settings,
                "respondent_id_column": updated_respondent_id_column or NOT_AVAILABLE,
                "respondent_id_count": len(updated_respondent_id_hashes),
                "respondent_id_hashes": updated_respondent_id_hashes,
                "duplicate_overlap_threshold": DUPLICATE_RESPONDENT_ID_OVERLAP_THRESHOLD,
                "rules_updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }

            write_norm_dataset_workbook(updated_record, tables, source_data, updated_rules)
            records[record_index] = updated_record
            save_norm_database_manifest(records)
            refresh_norm_database_workbook(records)
            backup_norm_database_to_uploaded_datasets(
                records,
                updated_record,
                None,
                updated_rules,
            )
            create_norm_database_history_snapshot(
                "Update rules",
                records,
                updated_record,
            )
            updated_record_for_commit = updated_record
    except TimeoutError as exc:
        return False, str(exc)

    dataset_id = (
        normalize_answer(updated_record_for_commit.get("dataset_id"))
        if updated_record_for_commit
        else "dataset"
    )
    _commit_ok, commit_message = github_autocommit_uploaded_datasets(
        f"Update BLS norms dataset rules {dataset_id}"
    )
    return True, save_message_with_github_status(
        "Saved dataset rules updated and norm tables regenerated.",
        _commit_ok,
        commit_message,
    )


def delete_norm_dataset_from_database(record: dict) -> tuple[bool, str]:
    dataset_id = normalize_answer(record.get("dataset_id"))
    if not dataset_id:
        return False, "Saved dataset record is missing a dataset ID."

    deleted_record_for_commit = None
    try:
        with norm_database_write_lock():
            records = load_norm_database_manifest()
            record_index = next(
                (
                    index
                    for index, saved_record in enumerate(records)
                    if saved_record.get("dataset_id") == dataset_id
                ),
                None,
            )
            if record_index is None:
                return False, "Saved dataset record was not found in the manifest."

            backup_norm_database_to_uploaded_datasets(records)
            create_norm_database_history_snapshot(
                "Before delete dataset",
                records,
                records[record_index],
                "Automatic safety snapshot before deleting a saved dataset.",
            )

            deleted_record = records.pop(record_index)
            norm_database_dataset_path(dataset_id).unlink(missing_ok=True)
            save_norm_database_manifest(records)
            refresh_norm_database_workbook(records)
            backup_norm_database_to_uploaded_datasets(records)
            create_norm_database_history_snapshot(
                "Delete dataset",
                records,
                deleted_record,
                "Saved dataset removed from the active manifest.",
            )
            deleted_record_for_commit = deleted_record
    except (OSError, TimeoutError) as exc:
        return False, str(exc)

    commit_ok, commit_message = github_autocommit_uploaded_datasets(
        f"Delete BLS norms dataset {dataset_id}"
    )
    return True, save_message_with_github_status(
        "Saved dataset removed. A version-history restore point was created.",
        commit_ok,
        commit_message,
    )


def render_persistence_status() -> None:
    config = github_autocommit_config()
    if config.get("enabled") and config.get("token") and config.get("repo"):
        return

    st.warning(
        "Permanent cross-device storage is not configured. Saved datasets can "
        "disappear when Streamlit Cloud restarts unless `[github_autocommit]` "
        "secrets are set."
    )
    with st.expander("GitHub backup secrets needed", expanded=False):
        st.code(
            """[github_autocommit]
enabled = true
repo = "OWNER/REPO"
branch = "main"
token = "github_pat_or_classic_token"
data_path = "uploaded_datasets"
""",
            language="toml",
        )


def norm_history_option_label(entry: dict) -> str:
    return (
        f"{entry.get('created_at', NOT_AVAILABLE)} | "
        f"{entry.get('action', NOT_AVAILABLE)} | "
        f"{entry.get('dataset_count', 0)} datasets"
    )


def norm_history_manifest_path(entry: dict) -> Path:
    version_id = normalize_answer(entry.get("version_id"))
    return norm_history_version_dir(version_id) / "norm_settings" / "manifest.json"


def norm_database_records_display(records: list[dict]) -> pd.DataFrame:
    return (
        pd.DataFrame(flatten_norm_database_record(record) for record in records)
        .fillna(NOT_AVAILABLE)
        .astype(str)
    )


def render_norm_history_controls() -> None:
    st.subheader("Version history")

    history_entries = load_norm_history_index()
    if not history_entries:
        st.info("No dataset restore points are available yet.")
        return

    selected_version = st.selectbox(
        "Version to restore",
        history_entries,
        format_func=norm_history_option_label,
        key="norm_history_version_to_restore",
    )
    if not selected_version:
        return

    preview_manifest_path = norm_history_manifest_path(selected_version)
    preview_records = read_manifest_records(preview_manifest_path)
    st.markdown("**Saved datasets in selected version**")
    if preview_records:
        render_vn_table(norm_database_records_display(preview_records))
    elif preview_manifest_path.exists():
        st.info("This version has no saved datasets.")
    else:
        st.warning("This version is missing its saved-datasets manifest.")

    with st.expander("Restore saved datasets from selected version", expanded=False):
        st.warning(
            "Restoring replaces the active saved-datasets manifest with the selected "
            "version. A safety snapshot of the current state is created first."
        )
        confirm_restore = st.checkbox(
            "I understand this will replace the active saved datasets",
            key=f"confirm_restore_{selected_version.get('version_id')}",
        )
        if st.button(
            "Restore selected version",
            disabled=not confirm_restore,
            use_container_width=True,
        ):
            success, message = restore_norm_database_history_version(
                selected_version.get("version_id", "")
            )
            if success:
                st.success(message)
                st.rerun()
            else:
                st.error(message)


def ensure_norm_history_baseline(records: list[dict]) -> None:
    if not records or load_norm_history_index():
        return
    backup_norm_database_to_uploaded_datasets(records)
    create_norm_database_history_snapshot(
        "Baseline",
        records,
        note="Initial restore point for saved datasets that existed before version history.",
    )


def render_saved_datasets_tab() -> None:
    st.subheader("Saved datasets")
    st.write(
        "Review saved norm datasets and edit their saved calculation rules when "
        "standards change."
    )
    records = load_norm_database_manifest()
    ensure_norm_history_baseline(records)
    render_persistence_status()
    render_norm_history_controls()

    if not records:
        st.info("No saved norm datasets are available yet.")
        return

    render_vn_table(norm_database_records_display(records))

    selected_record = st.selectbox(
        "Dataset to edit",
        records,
        format_func=saved_dataset_option_label,
    )
    if not selected_record:
        return

    with st.expander("Remove selected saved dataset", expanded=False):
        st.warning(
            "Removing a dataset takes it out of the active saved norms database. "
            "A version-history restore point is created before deletion."
        )
        confirm_delete = st.checkbox(
            "I understand this will remove the selected saved dataset",
            key=f"delete_dataset_confirm_{selected_record.get('dataset_id')}",
        )
        if st.button(
            "Remove selected dataset",
            disabled=not confirm_delete,
            use_container_width=True,
        ):
            success, message = delete_norm_dataset_from_database(selected_record)
            if success:
                st.success(message)
                st.rerun()
            else:
                st.error(message)

    dataset_path = norm_database_dataset_path(selected_record.get("dataset_id", ""))
    if not dataset_path.exists():
        st.error("The saved dataset workbook is missing.")
        return

    rules = read_norm_dataset_rules(dataset_path)
    try:
        source_data = pd.read_excel(
            dataset_path,
            sheet_name=NORM_DATASET_RESPONDENT_SHEET,
            dtype=object,
        )
    except Exception:
        source_data = None

    if rules is None or source_data is None:
        st.warning(
            "This saved dataset was created before editable rules were stored. "
            "Re-save or replace the dataset from the original upload to enable rule edits."
        )
        return

    columns = list(source_data.columns)
    saved_group_column = rules.get("group_column") or selected_record.get("group_column")
    group_index = columns.index(saved_group_column) if saved_group_column in columns else 0
    group_column = st.selectbox(
        "Control/test group column",
        columns,
        index=group_index,
        key=f"saved_group_{selected_record.get('dataset_id')}",
    )
    group_labels = (
        source_data[group_column]
        .map(normalize_answer)
        .dropna()
        .drop_duplicates()
        .tolist()
    )
    if len(group_labels) < 2:
        st.error("The saved group column needs at least two non-empty labels.")
        return

    default_control = rules.get("control_label") or selected_record.get("control_label")
    default_test = rules.get("test_label") or selected_record.get("test_label")
    control_index = group_labels.index(default_control) if default_control in group_labels else 0
    test_index = group_labels.index(default_test) if default_test in group_labels else min(1, len(group_labels) - 1)
    control_col, test_col = st.columns(2)
    control_label = control_col.selectbox(
        "Control label",
        group_labels,
        index=control_index,
        key=f"saved_control_{selected_record.get('dataset_id')}",
    )
    test_label = test_col.selectbox(
        "Test label",
        group_labels,
        index=test_index,
        key=f"saved_test_{selected_record.get('dataset_id')}",
    )

    saved_respondent_id_column = normalize_answer(
        selected_record.get("respondent_id_column")
    )
    if saved_respondent_id_column == NOT_AVAILABLE:
        saved_respondent_id_column = None
    respondent_id_column = render_respondent_id_selector(
        source_data,
        rules.get("question_labels", {}),
        key=f"saved_respondent_id_{selected_record.get('dataset_id')}",
        current_column=saved_respondent_id_column,
    )

    data_columns = [
        column
        for column in columns
        if column != group_column and not is_project_metadata_variable(column)
    ]
    rules_frame = saved_dataset_rules_editor_frame(rules, data_columns)
    edited_rules = st.data_editor(
        rules_frame,
        key=f"saved_rules_editor_{selected_record.get('dataset_id')}",
        use_container_width=True,
        hide_index=True,
        num_rows="fixed",
        column_config={
            "Include": st.column_config.CheckboxColumn("Include", width="small"),
            "Source variable": st.column_config.TextColumn(
                "Source variable",
                disabled=True,
                width=180,
            ),
            "Norm / benchmark": st.column_config.TextColumn(
                "Norm / benchmark",
                width=220,
            ),
            "Denominator": st.column_config.SelectboxColumn(
                "Denominator",
                options=DENOMINATOR_OPTIONS,
                required=True,
                width=150,
            ),
            "Question Type": st.column_config.SelectboxColumn(
                "Question Type",
                options=QUESTION_TYPES,
                required=True,
                width=170,
            ),
        },
    )

    (
        included_mappings,
        denominator_settings,
        box_score_settings,
        question_type_settings,
        norm_rules,
    ) = normalize_saved_dataset_rules_editor(edited_rules)
    st.caption(f"Included norms after edit: {len(included_mappings):,}")

    if st.button("Save updated dataset rules", use_container_width=True):
        if control_label == test_label:
            st.error("Control and test labels must be different.")
            return
        if not included_mappings:
            st.error("At least one saved rule must be included.")
            return

        updated_rules = {
            **rules,
            "group_column": group_column,
            "control_label": control_label,
            "test_label": test_label,
            "norm_rules": norm_rules,
        }
        tables = build_norm_tables(
            source_data,
            list(included_mappings.keys()),
            included_mappings,
            group_column,
            control_label,
            test_label,
            denominator_settings,
            rules.get("split_multi_select", False),
            rules.get("delimiter", ";"),
            rules.get("response_labels", {}),
            rules.get("question_labels", {}),
            box_score_settings,
            {},
            question_type_settings,
            {},
        )
        success, message = update_saved_norm_dataset_rules(
            selected_record,
            source_data,
            updated_rules,
            tables,
            respondent_id_column,
        )
        if success:
            st.success(message)
        else:
            st.error(message)


def load_saved_norm_tables() -> pd.DataFrame:
    if not NORM_DATABASE_WORKBOOK_PATH.exists():
        return pd.DataFrame()

    try:
        return normalize_lift_output_table(
            pd.read_excel(NORM_DATABASE_WORKBOOK_PATH, sheet_name="All Norms", dtype=object)
        )
    except Exception:
        return pd.DataFrame()


def saved_norm_database_download(
    saved_tables: pd.DataFrame,
    saved_records: list[dict],
) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(
            [flatten_norm_database_record(record) for record in saved_records]
        ).to_excel(writer, sheet_name="Datasets", index=False)
        normalize_lift_output_table(saved_tables).to_excel(
            writer,
            sheet_name="All Norms",
            index=False,
        )
    return output.getvalue()


def render_saved_norm_tables_review() -> None:
    st.subheader("Saved norm tables")
    st.caption(
        "These tables recalculate against the saved norms database as one combined "
        "read. Upload a workbook only when you want to audit and add or replace a dataset."
    )
    saved_records = load_norm_database_manifest()

    if not saved_records:
        metric_cols = st.columns(3)
        metric_cols[0].metric("Saved datasets", "0")
        metric_cols[1].metric("Respondent rows", "0")
        metric_cols[2].metric("Saved metrics", "0")
        st.info("No saved norm tables are available yet.")
        return

    selected_filters = render_saved_norm_filter_controls(saved_records)
    filter_totals = saved_norm_filter_totals(saved_records, selected_filters)
    combined_tables, source_variables_by_metric = build_combined_saved_norm_tables(
        saved_records,
        selected_filters,
    )

    metric_cols = st.columns(5)
    metric_cols[0].metric(
        "Saved datasets",
        f"{filter_totals['datasets']:,}",
    )
    metric_cols[1].metric("Filtered rows", f"{filter_totals['rows']:,}")
    metric_cols[2].metric("Control rows", f"{filter_totals['control_rows']:,}")
    metric_cols[3].metric("Test rows", f"{filter_totals['test_rows']:,}")
    metric_cols[4].metric("Saved metrics", f"{len(combined_tables):,}")

    if not combined_tables:
        st.info("No saved norm tables are available for the current filters.")
        return

    for metric, metric_table in combined_tables.items():
        st.subheader(str(metric))
        source_variables = source_variables_by_metric.get(metric, [])
        if source_variables:
            st.caption(f"Source variable: {', '.join(source_variables)}")
        render_norm_table_with_chart(metric_table.reset_index(drop=True))

    st.download_button(
        "Download saved norms Excel",
        data=norm_tables_to_excel(combined_tables),
        file_name="combined_saved_norm_tables.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def render_vn_table(table: pd.DataFrame) -> None:
    display_table = table.copy()
    table_html = display_table.to_html(
        index=False,
        border=0,
        classes="vn-norm-table",
        escape=True,
    )
    st.markdown(
        f'<div class="vn-norm-table-wrap">{table_html}</div>',
        unsafe_allow_html=True,
    )


def chart_bar_height(points: float) -> float:
    return max(0, min(100, points))


def lift_direction_class(points: float, significant: bool) -> str:
    if not significant:
        return "neutral"
    if points > 0:
        return "positive"
    if points < 0:
        return "negative"
    return "neutral"


def render_norm_bar_chart(table: pd.DataFrame) -> None:
    chart_rows = norm_chart_display_rows(table)
    if not chart_rows:
        st.caption("No chartable control/test percentages are available.")
        return

    groups = []
    for row in chart_rows:
        control_height = chart_bar_height(row["control_points"])
        test_height = chart_bar_height(row["test_points"])
        lift_label = format_lift_points(row["lift_points"]).replace("pts", "")
        lift_class = lift_direction_class(row["lift_points"], row["significant"])
        groups.append(
            '<div class="vn-chart-group">'
            '<div class="vn-chart-plot">'
            '<div class="vn-chart-bar-column">'
            f'<div class="vn-chart-value control">{escape(format_percent_points(row["control_points"]))}</div>'
            f'<div class="vn-chart-bar control" style="height:{control_height:.1f}%"></div>'
            "</div>"
            '<div class="vn-chart-bar-column">'
            f'<div class="vn-chart-value test">{escape(format_percent_points(row["test_points"]))}</div>'
            f'<div class="vn-chart-bar test" style="height:{test_height:.1f}%"></div>'
            "</div>"
            f'<div class="vn-lift-bubble {lift_class}">{escape(lift_label)}</div>'
            "</div>"
            f'<div class="vn-chart-label">{escape(row["chart_label"])}</div>'
            "</div>"
        )

    st.markdown(
        '<div class="vn-chart-card">'
        '<div class="vn-chart-header">'
        '<div class="vn-chart-title">Control vs test</div>'
        '<div class="vn-chart-legend">'
        '<span class="vn-chart-legend-item">'
        '<span class="vn-chart-legend-swatch control"></span>Control'
        "</span>"
        '<span class="vn-chart-legend-item">'
        '<span class="vn-chart-legend-swatch test"></span>Test'
        "</span>"
        '<span class="vn-chart-legend-item">'
        '<span class="vn-chart-legend-bubble">+/-</span>Lift'
        "</span>"
        "</div>"
        "</div>"
        '<div class="vn-chart-scroll">'
        f'{"".join(groups)}'
        "</div>"
        "</div>",
        unsafe_allow_html=True,
    )


def render_norm_table(table: pd.DataFrame) -> None:
    render_vn_table(normalize_lift_output_table(table))


def render_norm_table_with_chart(table: pd.DataFrame) -> None:
    render_norm_table(table)
    render_norm_bar_chart(table)


def main() -> None:
    st.set_page_config(
        page_title="BLS Norms Database",
        layout="wide",
        initial_sidebar_state="collapsed",
    )
    apply_bls_theme()
    render_bls_header()

    audit_tab, mapping_tab, norms_tab, saved_datasets_tab = st.tabs(
        PAGE_NAMES
    )

    uploaded_file = None
    uploaded_file_name = None
    workbook_bytes = None
    excel_file = None
    data_sheet = None
    data = None
    label_sheet = NO_LABEL_SHEET
    review_saved_norms = st.session_state.get("review_saved_norms_mode", False)
    data_layout = SMART_TABLES_LAYOUT
    response_labels: dict[str, dict[str, str]] = {}
    question_labels: dict[str, str] = {}
    candidate_questions: list[str] = []
    norm_questions: list[str] = []
    norm_mappings: dict[str, str] = {}
    included_mappings: dict[str, str] = {}
    saved_norm_mappings = load_norm_mapping_settings()
    saved_box_score_settings = load_box_score_settings()
    saved_question_type_settings = load_question_type_settings()
    saved_na_alias_settings = load_na_alias_settings()
    saved_denominator_settings = load_denominator_settings()
    prior_norm_rules = load_saved_norm_rule_history()
    box_score_settings: dict[str, list[str]] = {}
    question_type_settings: dict[str, str] = {}
    audit_consistency_issues: list[dict] = []
    audit_consistency_confirmed = True
    respondent_id_column = None
    group_column = None
    control_label = None
    test_label = None
    split_multi_select = False
    delimiter = ";"

    with audit_tab:
        render_page_navigation(0, review_saved_norms)
        st.subheader("Survey Question Audit")
        st.write(
            "Upload a workbook, confirm the control/test setup, then review "
            "which norm or benchmark each question should map to."
        )

        review_saved_norms = st.checkbox(
            "Review saved norms without uploading a workbook",
            help="Use this when you want to inspect saved norm, denominator, and box-score decisions without updating them from a new workbook.",
            key="review_saved_norms_mode",
        )

        if review_saved_norms:
            norm_mappings = saved_norm_mappings
            box_score_settings = saved_box_score_settings
            question_type_settings = saved_question_type_settings
            included_mappings = included_norm_mappings(norm_mappings)
            norm_questions = list(included_mappings.keys())
        else:
            uploaded_file = st.file_uploader(
                "Survey Excel workbook",
                type=["xlsx", "xlsm", "xls"],
                key=UPLOAD_WORKBOOK_SESSION_KEY,
                on_change=cache_uploaded_workbook_in_session,
            )

            if uploaded_file is not None:
                uploaded_file_name = uploaded_file.name
                workbook_bytes, excel_file = read_excel_workbook(uploaded_file)
                if workbook_bytes is not None:
                    st.session_state[ACTIVE_WORKBOOK_NAME_SESSION_KEY] = uploaded_file_name
                    st.session_state[ACTIVE_WORKBOOK_BYTES_SESSION_KEY] = workbook_bytes
            else:
                active_workbook = active_uploaded_workbook_from_session()
                if active_workbook is not None:
                    uploaded_file_name, workbook_bytes = active_workbook
                    try:
                        excel_file = pd.ExcelFile(BytesIO(workbook_bytes))
                    except Exception as exc:
                        st.error(f"Could not read Excel workbook: {exc}")
                        clear_uploaded_workbook_session()
                        workbook_bytes = None
                        excel_file = None

            if workbook_bytes is None or excel_file is None:
                st.info("Upload a workbook to begin setup, or use the saved-norm review option above.")
            else:
                st.caption(f"Workbook loaded: `{uploaded_file_name}`")

        if not review_saved_norms and workbook_bytes is not None and excel_file is not None:
            data_sheet = st.selectbox("Respondent data sheet", excel_file.sheet_names)
            data_layout = SMART_TABLES_LAYOUT

            respondent_sheet = read_respondent_sheet(workbook_bytes, data_sheet, data_layout)
            if respondent_sheet is None:
                st.stop()

            data = respondent_sheet.dataframe
            question_labels.update(respondent_sheet.question_labels)

            if data is None or data.empty:
                st.warning("The selected respondent data sheet is empty or unavailable.")
            else:
                metric_col_1, metric_col_2, metric_col_3, metric_col_4 = st.columns(4)
                metric_col_1.metric("Respondent rows", f"{len(data):,}")
                metric_col_2.metric("Question columns", f"{len(data.columns):,}")
                metric_col_3.metric(
                    "Metadata rows removed",
                    f"{respondent_sheet.metadata_rows_removed:,}",
                )
                metric_col_4.metric(
                    "Metadata variables defaulted NA",
                    f"{respondent_sheet.metadata_columns_default_na:,}",
                )

                if respondent_sheet.metadata_columns:
                    with st.expander("Metadata variables defaulted NA", expanded=False):
                        st.write(", ".join(str(column) for column in respondent_sheet.metadata_columns))

                label_sheet, response_labels, label_question_labels = auto_detect_response_labels(
                    workbook_bytes,
                    excel_file,
                    data_sheet,
                )
                question_labels.update(label_question_labels)
                st.subheader("Respondent ID setup")
                respondent_id_column = render_respondent_id_selector(
                    data,
                    question_labels,
                    key=f"upload_respondent_id_{uploaded_file_name}_{data_sheet}",
                )
                upload_duplicate_probe = norm_database_duplicate_probe_for_data(
                    workbook_bytes,
                    uploaded_file_name,
                    data_sheet,
                    data,
                    question_labels,
                    respondent_id_column,
                )
                upload_duplicate_match = find_norm_database_duplicate_record(
                    load_norm_database_manifest(),
                    upload_duplicate_probe,
                    DUPLICATE_RESPONDENT_ID_OVERLAP_THRESHOLD,
                )
                if upload_duplicate_match:
                    st.warning("Dataset already added to norms.")

                data, question_labels = prompt_for_missing_project_metadata(
                    data,
                    question_labels,
                )

                st.subheader("Sample setup")
                columns = list(data.columns)
                group_column = st.selectbox(
                    "Control/test group column",
                    columns,
                    index=default_group_column_index(columns),
                )
                candidate_questions = [
                    column
                    for column in columns
                    if column != group_column and not is_project_metadata_variable(column)
                ]
                audit_signature = hashlib.sha1(
                    json.dumps([str(column) for column in candidate_questions]).encode("utf-8")
                ).hexdigest()
                if st.session_state.get("norm_audit_signature") != audit_signature:
                    st.session_state.norm_audit_signature = audit_signature
                    st.session_state.norm_audit_pending_box_score_settings = {}
                    st.session_state.norm_audit_pending_question_type_settings = {}
                    st.session_state.norm_audit_editor_version = 0

                group_labels = (
                    data[group_column]
                    .map(normalize_answer)
                    .dropna()
                    .drop_duplicates()
                    .tolist()
                )

                if len(group_labels) < 2:
                    st.error(
                        "The selected group column needs at least two non-empty labels."
                    )
                else:
                    control_col, test_col = st.columns(2)
                    control_label = control_col.selectbox(
                        "Control label",
                        group_labels,
                        format_func=lambda value: display_response_label(
                            group_column,
                            value,
                            response_labels,
                        ),
                    )
                    default_test_index = 1 if len(group_labels) > 1 else 0
                    test_label = test_col.selectbox(
                        "Test label",
                        group_labels,
                        index=default_test_index,
                        format_func=lambda value: display_response_label(
                            group_column,
                            value,
                            response_labels,
                        ),
                    )

                    if control_label == test_label:
                        st.error("Control and test labels must be different.")

                st.subheader("Norm / benchmark audit")
                st.caption(
                    "The app suggests the closest norm or benchmark from the available "
                    "norm list. For the first project this defaults to each question's "
                    "own variable name. Smart Tables metadata variables default to NA. "
                    "Change the dropdown when needed, check NA to exclude a question, "
                    "or select T2B/T3B/B2B/B3B to add top/bottom-box norm rows."
                )

                norm_catalog = build_norm_catalog(
                    candidate_questions,
                    saved_norm_mappings,
                    prior_norm_rules,
                )
                norm_options = [NA_NORM_OPTION, *norm_catalog]
                pending_box_score_settings = st.session_state.get(
                    "norm_audit_pending_box_score_settings",
                    {},
                )
                pending_question_type_settings = st.session_state.get(
                    "norm_audit_pending_question_type_settings",
                    {},
                )
                effective_box_score_settings = {
                    **saved_box_score_settings,
                    **pending_box_score_settings,
                }
                effective_question_type_settings = {
                    **saved_question_type_settings,
                    **pending_question_type_settings,
                }
                audit_frame = build_norm_audit_frame(
                    data,
                    candidate_questions,
                    question_labels,
                    response_labels,
                    saved_norm_mappings,
                    effective_box_score_settings,
                    effective_question_type_settings,
                    saved_na_alias_settings,
                    split_multi_select,
                    delimiter,
                    prior_norm_rules,
                )
                if "norm_audit_editor_version" not in st.session_state:
                    st.session_state.norm_audit_editor_version = 0

                scale_action_cols = st.columns(5)
                bulk_selection: list[str] | None = None
                if scale_action_cols[0].button("All scale T2B", use_container_width=True):
                    bulk_selection = ["T2B"]
                if scale_action_cols[1].button("All scale T3B", use_container_width=True):
                    bulk_selection = ["T3B"]
                if scale_action_cols[2].button("All scale B2B", use_container_width=True):
                    bulk_selection = ["B2B"]
                if scale_action_cols[3].button("All scale B3B", use_container_width=True):
                    bulk_selection = ["B3B"]
                if scale_action_cols[4].button("Clear scale boxes", use_container_width=True):
                    bulk_selection = []

                if bulk_selection is not None:
                    audit_frame = apply_scale_box_score_selection(audit_frame, bulk_selection)
                    st.session_state.norm_audit_pending_box_score_settings = (
                        normalize_box_score_audit_editor(audit_frame)
                    )
                    st.session_state.norm_audit_pending_question_type_settings = (
                        normalize_question_type_audit_editor(audit_frame)
                    )
                    st.session_state.norm_audit_editor_version += 1

                edited_audit = st.data_editor(
                    audit_frame,
                    key=f"norm_audit_editor_{st.session_state.norm_audit_editor_version}",
                    use_container_width=True,
                    hide_index=True,
                    num_rows="fixed",
                    column_config={
                        "NA": st.column_config.CheckboxColumn(
                            "NA",
                            help="Exclude this question from the norms database.",
                            width="small",
                        ),
                        "Variable Name": st.column_config.TextColumn(
                            "Variable Name",
                            disabled=True,
                            width=180,
                        ),
                        "Question Text": st.column_config.TextColumn(
                            "Question Text",
                            disabled=True,
                            width=420,
                        ),
                        "Question Type": st.column_config.SelectboxColumn(
                            "Question Type",
                            options=QUESTION_TYPES,
                            required=True,
                            width=170,
                        ),
                        "Suggested norm/benchmark": st.column_config.TextColumn(
                            "Suggested norm/benchmark",
                            disabled=True,
                            width=220,
                        ),
                        "Norm / benchmark": st.column_config.SelectboxColumn(
                            "Norm / benchmark",
                            options=norm_options,
                            required=True,
                            width=220,
                        ),
                        "T2B": st.column_config.CheckboxColumn(
                            "T2B",
                            help="Add a Top 2 Box row using the first two ordered scale options.",
                            width="small",
                        ),
                        "T3B": st.column_config.CheckboxColumn(
                            "T3B",
                            help="Add a Top 3 Box row using the first three ordered scale options.",
                            width="small",
                        ),
                        "B2B": st.column_config.CheckboxColumn(
                            "B2B",
                            help="Add a Bottom 2 Box row using the last two ordered scale options.",
                            width="small",
                        ),
                        "B3B": st.column_config.CheckboxColumn(
                            "B3B",
                            help="Add a Bottom 3 Box row using the last three ordered scale options.",
                            width="small",
                        ),
                        "Answer Choices Count": st.column_config.NumberColumn(
                            "Answer Choices Count",
                            disabled=True,
                            width=160,
                        ),
                        "Answer Choices": st.column_config.TextColumn(
                            "Answer Choices",
                            disabled=True,
                            width=620,
                        ),
                    },
                )
                norm_mappings = normalize_norm_audit_editor(edited_audit)
                box_score_settings = normalize_box_score_audit_editor(edited_audit)
                question_type_settings = normalize_question_type_audit_editor(edited_audit)
                st.session_state.norm_audit_pending_box_score_settings = box_score_settings
                st.session_state.norm_audit_pending_question_type_settings = question_type_settings
                included_mappings = included_norm_mappings(norm_mappings)
                norm_questions = list(included_mappings.keys())

                metric_included, metric_excluded, metric_box_rows = st.columns(3)
                metric_included.metric("Included in norms", f"{len(norm_questions):,}")
                metric_excluded.metric(
                    "Marked NA",
                    f"{len(candidate_questions) - len(norm_questions):,}",
                )
                metric_box_rows.metric(
                    "Box score rows selected",
                    f"{sum(len(box_score_settings.get(question, [])) for question in norm_questions):,}",
                )

                audit_consistency_issues = norm_audit_prior_rule_issues(
                    edited_audit,
                    prior_norm_rules,
                )
                audit_consistency_confirmed = True
                if audit_consistency_issues:
                    st.warning(
                        "Some box-score selections do not match previous saved norms."
                    )
                    st.dataframe(
                        pd.DataFrame(audit_consistency_issues),
                        use_container_width=True,
                        hide_index=True,
                    )
                    st.info(
                        "Confirm this only when the standard is intentionally changing. "
                        "After saving, update previous saved datasets on the Saved datasets "
                        "page so each metric is measured consistently."
                    )
                    audit_consistency_confirmed = st.checkbox(
                        "Confirm this standards change for this upload",
                        key=f"norm_audit_consistency_confirmed_{audit_signature}",
                    )

                if st.button(
                    "Save audit mapping",
                    disabled=bool(audit_consistency_issues) and not audit_consistency_confirmed,
                ):
                    save_norm_mapping_settings(norm_mappings)
                    save_box_score_settings(box_score_settings)
                    save_question_type_settings(question_type_settings)
                    saved_na_alias_settings = merge_na_alias_settings(
                        normalize_na_aliases_from_audit_editor(edited_audit)
                    )
                    saved_norm_mappings = load_norm_mapping_settings()
                    saved_box_score_settings = load_box_score_settings()
                    saved_question_type_settings = load_question_type_settings()
                    st.session_state.norm_audit_pending_box_score_settings = saved_box_score_settings
                    st.session_state.norm_audit_pending_question_type_settings = saved_question_type_settings
                    st.success("Norm / benchmark audit mapping saved.")

                if not norm_questions:
                    st.info("At least one question must map to a norm or benchmark before tables can be calculated.")

        render_page_navigation(0, review_saved_norms)

    settings = load_denominator_settings()
    selected_denominators: dict[str, str] = {}
    setup_complete = (
        data is not None
        and not data.empty
        and group_column is not None
        and control_label is not None
        and test_label is not None
        and control_label != test_label
        and bool(norm_questions)
    )

    with mapping_tab:
        render_page_navigation(1, review_saved_norms)
        st.subheader("Denominator settings")
        if review_saved_norms:
            st.info(
                "Review mode skips denominator setup. Use Norm tables to review "
                "saved norms, or Saved datasets to edit saved rules."
            )
        elif not norm_questions:
            st.info("Complete the Survey Question Audit and include at least one mapped norm.")
        else:
            st.caption(
                "Default denominator: Total answering. Total sample uses the same "
                "base as the Smart Tables Total Base section."
            )

            for question in norm_questions:
                current_setting = settings.get(question, DEFAULT_DENOMINATOR)
                if current_setting not in DENOMINATOR_OPTIONS:
                    current_setting = DEFAULT_DENOMINATOR

                selected_denominators[question] = st.selectbox(
                    display_norm_variable_mapping_label(
                        question,
                        included_mappings.get(question, question),
                    ),
                    DENOMINATOR_OPTIONS,
                    index=DENOMINATOR_OPTIONS.index(current_setting),
                    key=setting_key(question),
                )

            if st.button("Save denominator settings"):
                persist_denominator_changes(settings, selected_denominators)
                settings = load_denominator_settings()
                st.success("Denominator settings saved.")

        render_page_navigation(1, review_saved_norms)

    with norms_tab:
        render_page_navigation(2, review_saved_norms)
        if not review_saved_norms and data is not None and not setup_complete:
            st.info(
                "Complete the Survey Question Audit to enable saving this upload. "
                "The saved norms database is shown below."
            )
        elif not review_saved_norms and setup_complete:
            effective_denominator_settings = {
                question: selected_denominators.get(
                    question,
                    settings.get(question, DEFAULT_DENOMINATOR),
                )
                for question in norm_questions
            }
            if workbook_bytes is not None and data_sheet is not None:
                full_tables = build_norm_tables(
                    data,
                    norm_questions,
                    included_mappings,
                    group_column,
                    control_label,
                    test_label,
                    effective_denominator_settings,
                    split_multi_select,
                    delimiter,
                    response_labels,
                    question_labels,
                    box_score_settings,
                    saved_box_score_settings,
                    question_type_settings,
                    saved_question_type_settings,
                )
                database_record = norm_database_record_for_upload(
                    workbook_bytes,
                    uploaded_file_name,
                    data_sheet,
                    data,
                    group_column,
                    control_label,
                    test_label,
                    norm_questions,
                    included_mappings,
                    effective_denominator_settings,
                    question_labels,
                    respondent_id_column,
                )
                database_rules = norm_database_rules_for_upload(
                    candidate_questions or norm_questions,
                    norm_questions,
                    included_mappings,
                    group_column,
                    control_label,
                    test_label,
                    effective_denominator_settings,
                    split_multi_select,
                    delimiter,
                    response_labels,
                    question_labels,
                    box_score_settings,
                    saved_box_score_settings,
                    question_type_settings,
                    saved_question_type_settings,
                )
                render_norm_database_save_controls(
                    database_record,
                    full_tables,
                    data,
                    database_rules,
                    workbook_bytes,
                    audit_consistency_issues,
                    audit_consistency_confirmed,
                )

        render_saved_norm_tables_review()

        render_page_navigation(2, review_saved_norms)

    with saved_datasets_tab:
        render_page_navigation(3, review_saved_norms)
        render_saved_datasets_tab()
        render_page_navigation(3, review_saved_norms)


if __name__ == "__main__":
    main()
